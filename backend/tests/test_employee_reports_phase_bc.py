"""
Tests for HRMS Employee Reporting Module (Phase B + C)
Endpoints:
  - GET /api/employees/reports/pf-ecr
  - GET /api/employees/reports/esi-return
  - GET /api/employees/reports/tds
  - GET /api/employees/reports/form-16/{employee_id}
  - GET /api/employees/reports/leave-balance
  - GET /api/employees/reports/ytd-earnings/{employee_id}
"""
import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    try:
        with open("/app/frontend/.env") as f:
            for line in f:
                if line.startswith("REACT_APP_BACKEND_URL="):
                    BASE_URL = line.strip().split("=", 1)[1].rstrip("/")
                    break
    except Exception:
        pass

API = f"{BASE_URL}/api/employees/reports"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME = "application/pdf"
TEST_MONTH = 4
TEST_YEAR = 2026
EMP_ID = "EMP-0001"
FY_START = 2025


def _all_cells(ws, max_row=200):
    """Yield all cell values in top-left region for scanning header keywords."""
    out = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
        if row:
            out.append(list(row))
    return out


# ==================== 1. PF ECR ====================
class TestPfEcr:
    def test_pf_ecr_valid_xlsx(self):
        r = requests.get(f"{API}/pf-ecr", params={"month": TEST_MONTH, "year": TEST_YEAR}, timeout=60)
        assert r.status_code == 200, r.text
        assert XLSX_MIME in r.headers.get("Content-Type", "")
        cd = r.headers.get("Content-Disposition", "")
        assert "attachment" in cd and "PF_ECR" in cd and ".xlsx" in cd
        # Valid xlsx zip signature
        assert r.content[:2] == b"PK"

        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        rows = _all_cells(ws)
        flat = [v for row in rows for v in row if v is not None]
        # Expected PF ECR headers
        for h in ["UAN", "Member Name", "Gross Wages", "EPF Wages", "EPS Wages",
                  "EDLI Wages", "EPF Contribution (12%)", "EPS Contribution (8.33%)",
                  "EPF Employer Share", "NCP Days", "Refund of Advances"]:
            assert h in flat, f"Header missing: {h}"
        # TOTAL row
        assert "TOTAL" in flat, "TOTAL row missing"

    def test_pf_ecr_wage_capping_15000(self):
        """EPF/EPS/EDLI wages must be capped at 15000 regardless of basic."""
        r = requests.get(f"{API}/pf-ecr", params={"month": TEST_MONTH, "year": TEST_YEAR}, timeout=60)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Find header row
        header_idx = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row and "UAN" in row:
                header_idx = i
                hdrs = list(row)
                break
        assert header_idx
        epf_col = hdrs.index("EPF Wages") + 1
        eps_col = hdrs.index("EPS Wages") + 1
        edli_col = hdrs.index("EDLI Wages") + 1
        # Data rows until TOTAL
        r_idx = header_idx + 1
        while True:
            v = ws.cell(row=r_idx, column=1).value
            if v is None or v == "TOTAL":
                break
            for col in (epf_col, eps_col, edli_col):
                cv = ws.cell(row=r_idx, column=col).value
                if isinstance(cv, (int, float)):
                    assert cv <= 15000, f"PF wage cap violated at row {r_idx} col {col}: {cv}"
            r_idx += 1

    def test_pf_ecr_invalid_month(self):
        r = requests.get(f"{API}/pf-ecr", params={"month": 13, "year": TEST_YEAR}, timeout=15)
        assert r.status_code == 422

    def test_pf_ecr_invalid_year(self):
        r = requests.get(f"{API}/pf-ecr", params={"month": 4, "year": 1999}, timeout=15)
        assert r.status_code == 422


# ==================== 2. ESI RETURN ====================
class TestEsiReturn:
    def test_esi_return_valid_xlsx(self):
        r = requests.get(f"{API}/esi-return", params={"month": TEST_MONTH, "year": TEST_YEAR}, timeout=60)
        assert r.status_code == 200, r.text
        assert XLSX_MIME in r.headers.get("Content-Type", "")
        cd = r.headers.get("Content-Disposition", "")
        assert "ESI_Return" in cd and ".xlsx" in cd
        assert r.content[:2] == b"PK"

        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        flat = [v for row in _all_cells(ws) for v in row if v is not None]
        for h in ["ESI IP Number", "Name", "Working Days", "Total Wages",
                  "ESI Wages", "Employee Contrib (0.75%)",
                  "Employer Contrib (3.25%)", "Total ESI"]:
            assert h in flat, f"Missing ESI header: {h}"

    def test_esi_return_handles_over_ceiling(self):
        """EMP-0001 has monthly_salary=25000 (>21000). Should NOT appear and
        the 'No employees under ESI wage ceiling' note should be present
        OR a covered-count TOTAL. Since it's the only active emp, expect the empty-state note."""
        r = requests.get(f"{API}/esi-return", params={"month": TEST_MONTH, "year": TEST_YEAR}, timeout=60)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        flat_strs = [str(v) for row in _all_cells(ws) for v in row if v is not None]
        joined = " | ".join(flat_strs)
        # Either the empty-state note OR TOTAL (covered) exists
        assert ("No employees under ESI wage ceiling" in joined) or ("TOTAL" in joined), (
            "Expected empty-state note or TOTAL row"
        )

    def test_esi_invalid_month(self):
        r = requests.get(f"{API}/esi-return", params={"month": 0, "year": TEST_YEAR}, timeout=15)
        assert r.status_code == 422


# ==================== 3. TDS ====================
class TestTdsReport:
    def test_tds_valid_xlsx(self):
        r = requests.get(f"{API}/tds", params={"fy_start_year": FY_START}, timeout=90)
        assert r.status_code == 200, r.text
        assert XLSX_MIME in r.headers.get("Content-Type", "")
        cd = r.headers.get("Content-Disposition", "")
        assert "TDS_FY" in cd and ".xlsx" in cd
        assert r.content[:2] == b"PK"

        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        flat = [v for row in _all_cells(ws) for v in row if v is not None]
        for h in ["Emp ID", "Name", "PAN", "Annual Gross", "Annual Basic",
                  "Total TDS (FY)", "Avg Monthly TDS", "Months Processed"]:
            assert h in flat, f"Missing TDS header: {h}"
        assert "GRAND TOTAL" in flat, "GRAND TOTAL row missing"

    def test_tds_months_processed_is_int(self):
        r = requests.get(f"{API}/tds", params={"fy_start_year": FY_START}, timeout=90)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        header_idx = None
        hdrs = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row and "Emp ID" in row:
                header_idx = i
                hdrs = list(row)
                break
        assert header_idx
        mp_col = hdrs.index("Months Processed") + 1
        # First data row
        v = ws.cell(row=header_idx + 1, column=mp_col).value
        assert isinstance(v, int), f"Months Processed must be int, got {type(v)}"
        # Recent joining (2026-04-18) + FY 2025 -> should be 0 or 1
        assert 0 <= v <= 12

    def test_tds_missing_param(self):
        r = requests.get(f"{API}/tds", timeout=15)
        assert r.status_code == 422


# ==================== 4. FORM 16 PDF ====================
class TestForm16:
    def test_form16_valid_pdf(self):
        r = requests.get(f"{API}/form-16/{EMP_ID}", params={"fy_start_year": FY_START}, timeout=90)
        assert r.status_code == 200, r.text
        assert PDF_MIME in r.headers.get("Content-Type", "")
        cd = r.headers.get("Content-Disposition", "")
        assert "Form16" in cd and ".pdf" in cd
        # PDF magic
        assert r.content[:4] == b"%PDF", "Not a valid PDF"
        # At least a few KB (real PDF with tables)
        assert len(r.content) > 2000

    def test_form16_nonexistent_employee_404(self):
        r = requests.get(f"{API}/form-16/EMP-NONEXIST-999",
                         params={"fy_start_year": FY_START}, timeout=30)
        assert r.status_code == 404

    def test_form16_missing_fy(self):
        r = requests.get(f"{API}/form-16/{EMP_ID}", timeout=15)
        assert r.status_code == 422


# ==================== 5. LEAVE BALANCE ====================
class TestLeaveBalance:
    def test_leave_balance_valid_xlsx(self):
        r = requests.get(f"{API}/leave-balance", timeout=60)
        assert r.status_code == 200, r.text
        assert XLSX_MIME in r.headers.get("Content-Type", "")
        cd = r.headers.get("Content-Disposition", "")
        assert "Leave_Balance" in cd and ".xlsx" in cd
        assert r.content[:2] == b"PK"

        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        flat = [v for row in _all_cells(ws) for v in row if v is not None]
        for h in ["Emp ID", "Name", "Department", "DOJ",
                  "CL Total", "CL Used", "CL Balance",
                  "SL Total", "SL Used", "SL Balance", "EL Balance"]:
            assert h in flat, f"Missing leave balance header: {h}"

    def test_leave_balance_has_emp_row(self):
        r = requests.get(f"{API}/leave-balance", timeout=60)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        flat = [str(v) for row in _all_cells(ws) for v in row if v is not None]
        assert EMP_ID in flat, f"{EMP_ID} row missing in leave-balance"


# ==================== 6. YTD EARNINGS ====================
class TestYtdEarnings:
    def test_ytd_valid_json(self):
        r = requests.get(f"{API}/ytd-earnings/{EMP_ID}",
                         params={"fy_start_year": FY_START}, timeout=60)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("success") is True
        # Top-level structure
        for k in ["employee", "financial_year", "monthly", "totals",
                  "ytd_days", "months_processed"]:
            assert k in data, f"Missing YTD key: {k}"
        # Employee info
        emp = data["employee"]
        assert emp["employee_id"] == EMP_ID
        assert "name" in emp
        # Totals shape
        tot = data["totals"]
        for k in ["gross", "net", "pf", "esi", "pt", "tds"]:
            assert k in tot, f"Missing totals.{k}"
            assert isinstance(tot[k], (int, float))
        # monthly is list
        assert isinstance(data["monthly"], list)
        # months_processed equals len(monthly)
        assert data["months_processed"] == len(data["monthly"])
        # financial_year string like "2025-26"
        assert data["financial_year"].startswith(str(FY_START))

    def test_ytd_monthly_entry_structure(self):
        r = requests.get(f"{API}/ytd-earnings/{EMP_ID}",
                         params={"fy_start_year": FY_START}, timeout=60)
        assert r.status_code == 200
        data = r.json()
        if data["monthly"]:
            m = data["monthly"][0]
            for k in ["month", "year", "gross", "net", "pf", "esi", "pt",
                      "tds", "lop", "days_paid"]:
                assert k in m, f"Missing monthly key: {k}"
            assert isinstance(m["year"], int)
            assert m["year"] in (FY_START, FY_START + 1)

    def test_ytd_nonexistent_employee_404(self):
        r = requests.get(f"{API}/ytd-earnings/EMP-NONEXIST-999",
                         params={"fy_start_year": FY_START}, timeout=30)
        assert r.status_code == 404

    def test_ytd_missing_fy(self):
        r = requests.get(f"{API}/ytd-earnings/{EMP_ID}", timeout=15)
        assert r.status_code == 422


# ==================== HEADERS / CONTENT TYPE SANITY ====================
class TestHeadersPhaseBC:
    def test_pf_ecr_filename(self):
        r = requests.get(f"{API}/pf-ecr", params={"month": 4, "year": 2026}, timeout=60)
        assert r.status_code == 200
        assert 'filename="PF_ECR_April_2026.xlsx"' in r.headers.get("Content-Disposition", "")

    def test_esi_filename(self):
        r = requests.get(f"{API}/esi-return", params={"month": 4, "year": 2026}, timeout=60)
        assert r.status_code == 200
        assert 'filename="ESI_Return_April_2026.xlsx"' in r.headers.get("Content-Disposition", "")

    def test_tds_filename(self):
        r = requests.get(f"{API}/tds", params={"fy_start_year": 2025}, timeout=90)
        assert r.status_code == 200
        assert 'filename="TDS_FY2025_26.xlsx"' in r.headers.get("Content-Disposition", "")
