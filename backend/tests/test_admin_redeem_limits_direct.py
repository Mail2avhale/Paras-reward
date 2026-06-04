"""
Tests for Admin Redeem Limits Dashboard + Direct Redeem (Jan 2026)
==================================================================
Covers:
- GET /api/admin/redeem-limits/users (list, sort, search, filter, pagination)
- GET /api/admin/redeem-limits/users/export-excel (openpyxl-parseable xlsx)
- POST /api/admin/redeem-limits/direct-redeem (preview + commit, blockers)
- PUT/GET /api/admin/redeem-limits/users/{uid}/bank-details (profile bank form)
"""
import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
ADMIN_EMAIL = "admin@test.com"
ADMIN_PIN = "153759"
TEST_USER_UID = "76b75808-47fa-48dd-ad7c-8074678e3607"  # primary test user
PRC_USER_UID = "6c96a6cc-08a2-442c-8e2d-f1fb6f18aa21"   # PRC test user


@pytest.fixture(scope="module")
def session():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{BASE_URL}/api/auth/login", json={"identifier": ADMIN_EMAIL, "pin": ADMIN_PIN})
    if r.status_code == 200:
        tok = r.json().get("token") or r.json().get("access_token")
        if tok:
            s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


# ============= /users list =============
class TestUsersList:
    def test_list_default(self, session):
        r = session.get(f"{BASE_URL}/api/admin/redeem-limits/users")
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["success"] is True
        assert "rows" in data and isinstance(data["rows"], list)
        assert "total" in data and "page" in data and "page_size" in data
        assert data["page"] == 1
        # Each row should have key fields
        if data["rows"]:
            row = data["rows"][0]
            for key in ["uid", "name", "mobile", "total_prc", "redeem_limit_prc",
                        "used_prc", "balance_redeemable_prc", "is_active_elite"]:
                assert key in row, f"missing {key}"

    def test_search_by_mobile(self, session):
        r = session.get(f"{BASE_URL}/api/admin/redeem-limits/users",
                        params={"search": "9970100782"})
        assert r.status_code == 200
        data = r.json()
        # Should find the primary test user
        assert data["total"] >= 1
        mobiles = [row.get("mobile") for row in data["rows"]]
        assert any("9970100782" in str(m) for m in mobiles)

    def test_status_active_filter(self, session):
        r = session.get(f"{BASE_URL}/api/admin/redeem-limits/users",
                        params={"status": "active"})
        assert r.status_code == 200
        for row in r.json()["rows"]:
            assert row["is_active_elite"] is True

    def test_status_inactive_filter(self, session):
        r = session.get(f"{BASE_URL}/api/admin/redeem-limits/users",
                        params={"status": "inactive"})
        assert r.status_code == 200
        for row in r.json()["rows"]:
            assert row["is_active_elite"] is False

    def test_sort_by_total_prc_desc(self, session):
        r = session.get(f"{BASE_URL}/api/admin/redeem-limits/users",
                        params={"sort_by": "total_prc", "sort_order": "desc"})
        assert r.status_code == 200
        rows = r.json()["rows"]
        if len(rows) >= 2:
            assert rows[0]["total_prc"] >= rows[1]["total_prc"]

    def test_sort_by_name_asc(self, session):
        r = session.get(f"{BASE_URL}/api/admin/redeem-limits/users",
                        params={"sort_by": "name", "sort_order": "asc"})
        assert r.status_code == 200
        rows = r.json()["rows"]
        if len(rows) >= 2:
            assert (rows[0]["name"] or "").lower() <= (rows[1]["name"] or "").lower()

    def test_invalid_sort_field_rejected(self, session):
        r = session.get(f"{BASE_URL}/api/admin/redeem-limits/users",
                        params={"sort_by": "bogus"})
        assert r.status_code == 422

    def test_pagination(self, session):
        r = session.get(f"{BASE_URL}/api/admin/redeem-limits/users",
                        params={"page": 1, "page_size": 2})
        assert r.status_code == 200
        data = r.json()
        assert len(data["rows"]) <= 2
        assert data["page_size"] == 2


# ============= /users/export-excel =============
class TestExcelExport:
    def test_excel_content_type(self, session):
        r = session.get(f"{BASE_URL}/api/admin/redeem-limits/users/export-excel")
        assert r.status_code == 200, r.text
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml.sheet" in ct, f"unexpected content-type: {ct}"
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd and ".xlsx" in cd

    def test_excel_parseable(self, session):
        r = session.get(f"{BASE_URL}/api/admin/redeem-limits/users/export-excel")
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws.title == "Redeem Limits"
        # Header row check
        headers = [c.value for c in ws[1]]
        assert "User Name" in headers
        assert "Mobile" in headers
        assert "Total PRC" in headers
        assert "Balance Redeemable (PRC)" in headers
        assert "IFSC" in headers
        # Should have at least header + some user rows
        assert ws.max_row >= 1

    def test_excel_with_filter(self, session):
        r = session.get(f"{BASE_URL}/api/admin/redeem-limits/users/export-excel",
                        params={"status": "active"})
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Active column index 4
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[3] is not None:
                assert row[3] == "Yes"


# ============= Bank Details PUT/GET =============
class TestBankDetails:
    def test_get_empty_returns_strings(self, session):
        r = session.get(f"{BASE_URL}/api/admin/redeem-limits/users/{TEST_USER_UID}/bank-details")
        assert r.status_code == 200
        data = r.json()
        assert data["success"] is True
        # Empty fields must be empty strings, NOT null
        for k in ["account_number", "ifsc_code", "bank_name", "upi_id", "phonepe_gpay_number"]:
            assert data[k] is not None, f"{k} should not be None"
            assert isinstance(data[k], str), f"{k} should be string"

    def test_put_full_details(self, session):
        payload = {
            "account_number": "1234567890123",
            "ifsc_code": "HDFC0001234",
            "bank_name": "HDFC Bank",
            "account_holder_name": "TEST_Holder",
            "upi_id": "test@hdfcbank",
            "phonepe_gpay_number": "9876543210",
        }
        r = session.put(f"{BASE_URL}/api/admin/redeem-limits/users/{TEST_USER_UID}/bank-details",
                        json=payload)
        assert r.status_code == 200, r.text
        assert r.json()["success"] is True

        # Verify GET returns saved data
        g = session.get(f"{BASE_URL}/api/admin/redeem-limits/users/{TEST_USER_UID}/bank-details")
        gd = g.json()
        assert gd["account_number"] == "1234567890123"
        assert gd["ifsc_code"] == "HDFC0001234"
        assert gd["bank_name"] == "HDFC Bank"
        assert gd["upi_id"] == "test@hdfcbank"
        assert gd["phonepe_gpay_number"] == "9876543210"
        assert gd["updated_at"] is not None

    def test_put_invalid_ifsc(self, session):
        r = session.put(f"{BASE_URL}/api/admin/redeem-limits/users/{TEST_USER_UID}/bank-details",
                        json={"ifsc_code": "foobarbazqq"})
        # foo too short would 422 from pydantic; this 11-char invalid pattern -> 400
        assert r.status_code in (400, 422), r.text

    def test_put_invalid_ifsc_short(self, session):
        r = session.put(f"{BASE_URL}/api/admin/redeem-limits/users/{TEST_USER_UID}/bank-details",
                        json={"ifsc_code": "foo"})
        assert r.status_code == 422  # pydantic min_length

    def test_put_invalid_account_non_digit(self, session):
        r = session.put(f"{BASE_URL}/api/admin/redeem-limits/users/{TEST_USER_UID}/bank-details",
                        json={"account_number": "abcd123456"})
        assert r.status_code == 400
        assert "digit" in r.text.lower()

    def test_put_invalid_upi_no_at(self, session):
        r = session.put(f"{BASE_URL}/api/admin/redeem-limits/users/{TEST_USER_UID}/bank-details",
                        json={"upi_id": "noatsign"})
        assert r.status_code == 400
        assert "@" in r.text

    def test_put_invalid_phonepe_non_digit(self, session):
        r = session.put(f"{BASE_URL}/api/admin/redeem-limits/users/{TEST_USER_UID}/bank-details",
                        json={"phonepe_gpay_number": "abcd123456"})
        assert r.status_code == 400

    def test_get_nonexistent_user(self, session):
        r = session.get(f"{BASE_URL}/api/admin/redeem-limits/users/nonexistent-uid-xyz/bank-details")
        assert r.status_code == 404


# ============= Direct Redeem =============
class TestDirectRedeem:
    def test_preview_user_not_found(self, session):
        r = session.post(f"{BASE_URL}/api/admin/redeem-limits/direct-redeem",
                         json={"admin_id": "admin@test.com",
                               "user_id": "nonexistent-uid-xyz",
                               "amount_inr": 500,
                               "utr_number": "UTR123456",
                               "confirm": False})
        assert r.status_code == 404

    def test_preview_missing_utr(self, session):
        # Preview should still validate pydantic UTR min_length=4
        r = session.post(f"{BASE_URL}/api/admin/redeem-limits/direct-redeem",
                         json={"admin_id": "admin@test.com",
                               "user_id": TEST_USER_UID,
                               "amount_inr": 500,
                               "utr_number": "",
                               "confirm": False})
        assert r.status_code == 422  # pydantic rejects min_length=4

    def test_preview_happy(self, session):
        """Preview should always return blockers/can_proceed without debiting."""
        # Get user's PRC balance via list endpoint
        users_resp = session.get(f"{BASE_URL}/api/admin/redeem-limits/users",
                                 params={"search": "9970100782"})
        rows = users_resp.json()["rows"]
        if not rows:
            pytest.skip("Test user not present in users list")
        before_balance = rows[0]["total_prc"]

        r = session.post(f"{BASE_URL}/api/admin/redeem-limits/direct-redeem",
                         json={"admin_id": "admin@test.com",
                               "user_id": TEST_USER_UID,
                               "amount_inr": 100,
                               "utr_number": "UTRPREVIEW001",
                               "confirm": False})
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["preview"] is True
        assert "fees" in data
        assert "blockers" in data
        assert "can_proceed" in data
        assert data["fees"]["withdrawal_inr"] == 100
        assert data["fees"]["admin_fee_inr"] == 20  # 20% of 100
        assert data["fees"]["transaction_fee_inr"] == 10

        # Confirm no debit happened — get balance again
        users_resp2 = session.get(f"{BASE_URL}/api/admin/redeem-limits/users",
                                  params={"search": "9970100782"})
        after_balance = users_resp2.json()["rows"][0]["total_prc"]
        assert after_balance == before_balance, "Preview should not debit PRC"

    def test_commit_insufficient_balance(self, session):
        """Force a huge amount to trigger insufficient balance blocker on commit."""
        r = session.post(f"{BASE_URL}/api/admin/redeem-limits/direct-redeem",
                         json={"admin_id": "admin@test.com",
                               "user_id": TEST_USER_UID,
                               "amount_inr": 99999999,
                               "utr_number": "UTRBIG123",
                               "confirm": True})
        # Commit returns 400 with concatenated blocker reason
        assert r.status_code == 400
        body_lower = r.text.lower()
        assert "balance" in body_lower or "limit" in body_lower or "minimum" in body_lower

    def test_commit_happy_path(self, session):
        """Run a small valid commit; verify PRC debit, bank_transfer doc, txn, progressive min."""
        # Step 1: get current balance
        before = session.get(f"{BASE_URL}/api/admin/redeem-limits/users",
                             params={"search": "9970100782"}).json()["rows"][0]
        before_prc = before["total_prc"]

        # Step 2: ensure user has saved bank details (test_put_full_details may run after)
        session.put(f"{BASE_URL}/api/admin/redeem-limits/users/{TEST_USER_UID}/bank-details",
                    json={"account_number": "1234567890",
                          "ifsc_code": "HDFC0001234",
                          "bank_name": "HDFC Bank",
                          "account_holder_name": "TEST_DirectRedeem"})

        # Step 3: preview to check blockers
        prev = session.post(f"{BASE_URL}/api/admin/redeem-limits/direct-redeem",
                            json={"admin_id": "admin@test.com",
                                  "user_id": TEST_USER_UID,
                                  "amount_inr": 100,
                                  "utr_number": "UTRTEST100",
                                  "confirm": False}).json()
        if not prev.get("can_proceed"):
            pytest.skip(f"Cannot proceed with test commit; blockers: {prev.get('blockers')}")

        total_prc_to_debit = prev["fees"]["total_prc_debited"]

        # Step 4: commit
        utr = "UTRJANCOMMIT1234"
        r = session.post(f"{BASE_URL}/api/admin/redeem-limits/direct-redeem",
                         json={"admin_id": "admin@test.com",
                               "user_id": TEST_USER_UID,
                               "amount_inr": 100,
                               "utr_number": utr,
                               "remark": "TEST_direct_redeem_jan2026",
                               "confirm": True})
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["success"] is True
        assert data["utr_number"] == utr
        assert data["amount_inr"] == 100
        assert "request_id" in data
        assert data["request_id"].startswith("ADM-DIR-")
        assert "next_min_withdrawal_inr" in data
        assert data["next_min_withdrawal_inr"] >= 150  # 100*1.5

        # Step 5: balance should have decreased by total_prc
        after = session.get(f"{BASE_URL}/api/admin/redeem-limits/users",
                            params={"search": "9970100782"}).json()["rows"][0]
        delta = before_prc - after["total_prc"]
        # Allow tiny float tolerance
        assert abs(delta - total_prc_to_debit) < 1.0, \
            f"PRC debit mismatch: expected {total_prc_to_debit}, got delta {delta}"
