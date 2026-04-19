"""
Tests for HRMS Employee Reporting Module (Phase A)
- Salary Register (Excel)
- Salary Slip (PDF)
- Attendance Report (Excel)
- Pool Distribution (Excel)
- HR Analytics (JSON)
"""
import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    # Fallback read from frontend/.env for safety
    try:
        with open("/app/frontend/.env") as f:
            for line in f:
                if line.startswith("REACT_APP_BACKEND_URL="):
                    BASE_URL = line.strip().split("=", 1)[1].rstrip("/")
                    break
    except Exception:
        pass

API = f"{BASE_URL}/api/employees/reports"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME = "application/pdf"
TEST_MONTH = 4
TEST_YEAR = 2026
EMP_ID = "EMP-0001"


# ==================== ANALYTICS ====================
class TestAnalytics:
    def test_analytics_returns_valid_json(self):
        r = requests.get(f"{API}/analytics", timeout=30)
        assert r.status_code == 200, r.text
        data = r.json()
        # Structure
        assert data.get("success") is True
        for key in ["headcount", "departments", "salary", "pool",
                    "top_earners_this_month", "leave_usage_this_year",
                    "attendance_this_month"]:
            assert key in data, f"Missing key {key}"

        # Headcount validation
        hc = data["headcount"]
        assert isinstance(hc["active"], int)
        assert isinstance(hc["total"], int)
        assert "attrition_rate" in hc

        # Salary stats
        sal = data["salary"]
        for k in ["total_monthly_cost", "avg_salary", "min_salary", "max_salary"]:
            assert k in sal
            assert isinstance(sal[k], (int, float))

        # Pool
        pool = data["pool"]
        for k in ["current_balance", "this_month_distributed",
                  "this_month_distribution_count", "prc_to_inr_rate"]:
            assert k in pool

        # Departments array
        assert isinstance(data["departments"], list)


# ==================== SALARY REGISTER ====================
class TestSalaryRegister:
    def test_salary_register_returns_valid_xlsx(self):
        r = requests.get(
            f"{API}/salary-register",
            params={"month": TEST_MONTH, "year": TEST_YEAR},
            timeout=60,
        )
        assert r.status_code == 200, r.text
        assert XLSX_MIME in r.headers.get("Content-Type", "")
        cd = r.headers.get("Content-Disposition", "")
        assert "attachment" in cd
        assert ".xlsx" in cd
        assert "Salary_Register" in cd

        # Parse bytes as real Excel
        wb = load_workbook(io.BytesIO(r.content), data_only=False)
        ws = wb.active
        # Find header row containing 'Emp ID'
        found = False
        net_col = None
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
            vals = [c.value for c in row]
            if vals and "Emp ID" in vals:
                found = True
                assert "Net Salary" in vals
                assert "Basic" in vals
                assert "PF (E)" in vals
                assert "TDS" in vals
                net_col = vals.index("Net Salary") + 1
                break
        assert found, "Header row with 'Emp ID' not found"
        assert net_col is not None

    def test_salary_register_invalid_month_zero(self):
        r = requests.get(
            f"{API}/salary-register",
            params={"month": 0, "year": TEST_YEAR},
            timeout=15,
        )
        assert r.status_code == 422

    def test_salary_register_invalid_month_13(self):
        r = requests.get(
            f"{API}/salary-register",
            params={"month": 13, "year": TEST_YEAR},
            timeout=15,
        )
        assert r.status_code == 422

    def test_salary_register_missing_params(self):
        r = requests.get(f"{API}/salary-register", timeout=15)
        assert r.status_code == 422


# ==================== SALARY SLIP PDF ====================
class TestSalarySlipPdf:
    def test_salary_slip_valid_employee(self):
        r = requests.get(
            f"{API}/salary-slip-pdf/{EMP_ID}",
            params={"month": TEST_MONTH, "year": TEST_YEAR},
            timeout=60,
        )
        assert r.status_code == 200, r.text
        assert PDF_MIME in r.headers.get("Content-Type", "")
        cd = r.headers.get("Content-Disposition", "")
        assert "attachment" in cd
        assert ".pdf" in cd
        # PDF signature
        assert r.content[:4] == b"%PDF", "Not a valid PDF stream"
        # Reasonable size (slip should be at least a few KB)
        assert len(r.content) > 1000

    def test_salary_slip_invalid_employee_returns_404(self):
        r = requests.get(
            f"{API}/salary-slip-pdf/EMP-NONEXISTENT-999",
            params={"month": TEST_MONTH, "year": TEST_YEAR},
            timeout=30,
        )
        assert r.status_code == 404

    def test_salary_slip_invalid_month(self):
        r = requests.get(
            f"{API}/salary-slip-pdf/{EMP_ID}",
            params={"month": 13, "year": TEST_YEAR},
            timeout=15,
        )
        assert r.status_code == 422


# ==================== ATTENDANCE REPORT ====================
class TestAttendanceReport:
    def test_attendance_returns_valid_xlsx(self):
        r = requests.get(
            f"{API}/attendance",
            params={"month": TEST_MONTH, "year": TEST_YEAR},
            timeout=60,
        )
        assert r.status_code == 200, r.text
        assert XLSX_MIME in r.headers.get("Content-Type", "")
        cd = r.headers.get("Content-Disposition", "")
        assert "Attendance" in cd and ".xlsx" in cd

        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Look for header row with 'Emp ID' and day columns
        header_row_idx = None
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if row and "Emp ID" in row:
                header_row_idx = idx
                vals = list(row)
                assert "Name" in vals
                assert "Department" in vals
                # Should contain day columns "01".."30"
                assert "01" in vals
                # April has 30 days → should find "30"
                assert "30" in vals
                # Summary
                assert "Present" in vals
                assert "Absent" in vals
                break
        assert header_row_idx is not None, "Attendance header row not found"

    def test_attendance_empty_month_does_not_crash(self):
        # A month far in future - no data but should still return valid xlsx (matrix of blanks)
        r = requests.get(
            f"{API}/attendance",
            params={"month": 12, "year": 2099},
            timeout=30,
        )
        # Could be 200 (empty matrix) or 404 (no active employees)
        assert r.status_code in (200, 404)
        if r.status_code == 200:
            wb = load_workbook(io.BytesIO(r.content))
            assert wb.active is not None

    def test_attendance_invalid_month(self):
        r = requests.get(
            f"{API}/attendance",
            params={"month": 0, "year": TEST_YEAR},
            timeout=15,
        )
        assert r.status_code == 422


# ==================== POOL DISTRIBUTION ====================
class TestPoolDistribution:
    def test_pool_distribution_returns_valid_xlsx(self):
        r = requests.get(
            f"{API}/pool-distribution",
            params={"from_date": "2026-04-01", "to_date": "2026-04-30"},
            timeout=30,
        )
        assert r.status_code == 200, r.text
        assert XLSX_MIME in r.headers.get("Content-Type", "")
        cd = r.headers.get("Content-Disposition", "")
        assert "Pool_Distribution" in cd

        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        found_header = False
        found_total_row = False
        for row in ws.iter_rows(min_row=1, max_row=60, values_only=True):
            if row and "Emp ID" in (row or ()):
                found_header = True
                vals = list(row)
                assert "Total PRC Earned" in vals
                assert "Distribution Count" in vals
                assert "Name" in vals
            if row and any(v == "TOTAL PRC DISTRIBUTED" for v in (row or ())):
                found_total_row = True
        assert found_header, "Header row missing in pool distribution"
        assert found_total_row, "TOTAL PRC DISTRIBUTED row missing"

    def test_pool_distribution_empty_range(self):
        # Date range with no data
        r = requests.get(
            f"{API}/pool-distribution",
            params={"from_date": "2020-01-01", "to_date": "2020-01-05"},
            timeout=30,
        )
        assert r.status_code == 200, r.text
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Should contain 'No distributions' note
        found_empty_note = False
        for row in ws.iter_rows(values_only=True):
            if row and any(isinstance(v, str) and "No distributions" in v for v in row if v):
                found_empty_note = True
                break
        assert found_empty_note, "Empty-state note not present for zero-row range"

    def test_pool_distribution_missing_params(self):
        r = requests.get(f"{API}/pool-distribution", timeout=15)
        assert r.status_code == 422


# ==================== CONTENT-TYPE / DISPOSITION ====================
class TestHeaders:
    def test_salary_register_headers(self):
        r = requests.get(
            f"{API}/salary-register",
            params={"month": TEST_MONTH, "year": TEST_YEAR},
            timeout=30,
        )
        assert r.status_code == 200
        assert XLSX_MIME in r.headers.get("Content-Type", "")
        cd = r.headers.get("Content-Disposition", "")
        assert 'filename="Salary_Register_April_2026.xlsx"' in cd

    def test_salary_slip_headers(self):
        r = requests.get(
            f"{API}/salary-slip-pdf/{EMP_ID}",
            params={"month": TEST_MONTH, "year": TEST_YEAR},
            timeout=30,
        )
        assert r.status_code == 200
        assert PDF_MIME in r.headers.get("Content-Type", "")
        cd = r.headers.get("Content-Disposition", "")
        assert ".pdf" in cd
