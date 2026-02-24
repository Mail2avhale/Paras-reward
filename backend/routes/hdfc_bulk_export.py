"""
HDFC Bank Bulk Payment Excel Export
Generates Excel files in HDFC Bank bulk upload format for:
- Bank Redeem (Withdrawal)
- EMI Payments  
- Savings Vault Redeem

HDFC Bank Bulk Payment Format:
- NEFT/RTGS format with beneficiary details
- Required fields: Account No, IFSC, Beneficiary Name, Amount, Narration

Created: February 2026
"""

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
from typing import Optional, List
import io
import logging

router = APIRouter(prefix="/admin/hdfc-export", tags=["HDFC Bulk Export"])

db = None

def set_db(database):
    global db
    db = database


def sanitize_text(text: str, max_length: int = 30) -> str:
    """
    Sanitize text for HDFC Bank format:
    - Remove special characters
    - Limit length
    - Convert to uppercase
    """
    if not text:
        return ""
    # Remove special characters, keep only alphanumeric and spaces
    clean = ''.join(c for c in str(text) if c.isalnum() or c == ' ')
    # Limit length and uppercase
    return clean[:max_length].upper().strip()


def format_amount(amount: float) -> str:
    """Format amount with 2 decimal places as required by HDFC"""
    return f"{float(amount):.2f}"


def generate_hdfc_excel(data: List[dict], payment_type: str):
    """
    Generate HDFC Bank bulk payment Excel file
    
    HDFC NEFT/RTGS Bulk Format Columns:
    1. Sr No
    2. Beneficiary Account Number
    3. Beneficiary Name
    4. IFSC Code
    5. Bank Name
    6. Branch Name
    7. Amount
    8. Narration/Remarks
    9. Email (optional)
    10. Mobile (optional)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"HDFC_{payment_type}"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="004B87", end_color="004B87", fill_type="solid")  # HDFC Blue
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center")
    amount_alignment = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers as per HDFC NEFT/RTGS format
    headers = [
        "SR NO",
        "BENEFICIARY ACCOUNT NO",
        "BENEFICIARY NAME",
        "IFSC CODE",
        "BANK NAME",
        "BRANCH NAME", 
        "AMOUNT",
        "NARRATION",
        "EMAIL",
        "MOBILE",
        "REQUEST ID",
        "USER ID"
    ]
    
    # Column widths
    column_widths = [8, 22, 30, 15, 25, 25, 15, 35, 30, 15, 20, 15]
    
    # Write headers
    for col, (header, width) in enumerate(zip(headers, column_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Write data rows
    total_amount = 0
    for row_num, item in enumerate(data, 2):
        # Extract bank details
        bank_details = item.get("bank_details", {})
        if not bank_details:
            bank_details = {
                "account_number": item.get("account_number", ""),
                "account_holder_name": item.get("account_holder_name", item.get("user_name", "")),
                "ifsc_code": item.get("ifsc_code", ""),
                "bank_name": item.get("bank_name", ""),
                "branch_name": item.get("branch_name", "")
            }
        
        amount = float(item.get("amount_inr", 0) or item.get("net_amount", 0) or item.get("amount", 0) or 0)
        total_amount += amount
        
        # Generate narration based on payment type
        request_id = item.get("request_id", item.get("rd_id", item.get("_id", "")))
        if payment_type == "BANK_REDEEM":
            narration = f"PRC REDEEM {str(request_id)[-8:].upper()}"
        elif payment_type == "EMI_PAY":
            narration = f"EMI PAYMENT {str(request_id)[-8:].upper()}"
        elif payment_type == "SAVINGS_VAULT":
            narration = f"VAULT REDEEM {str(request_id)[-8:].upper()}"
        else:
            narration = f"PARAS REWARD {str(request_id)[-8:].upper()}"
        
        row_data = [
            row_num - 1,  # SR NO
            str(bank_details.get("account_number", "")).strip(),  # ACCOUNT NO
            sanitize_text(bank_details.get("account_holder_name", ""), 30),  # BENEFICIARY NAME
            str(bank_details.get("ifsc_code", "")).upper().strip(),  # IFSC
            sanitize_text(bank_details.get("bank_name", ""), 25),  # BANK NAME
            sanitize_text(bank_details.get("branch_name", ""), 25),  # BRANCH NAME
            format_amount(amount),  # AMOUNT
            sanitize_text(narration, 35),  # NARRATION
            str(item.get("email", "")).strip()[:30],  # EMAIL
            str(item.get("mobile", item.get("phone", ""))).strip()[:15],  # MOBILE
            str(request_id)[-12:],  # REQUEST ID (for reference)
            str(item.get("uid", item.get("user_id", "")))[:15]  # USER ID
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col == 7:  # Amount column
                cell.alignment = amount_alignment
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = cell_alignment
    
    # Add summary row
    summary_row = len(data) + 2
    ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=summary_row, column=6, value=f"{len(data)} Records").font = Font(bold=True)
    total_cell = ws.cell(row=summary_row, column=7, value=format_amount(total_amount))
    total_cell.font = Font(bold=True, color="004B87")
    total_cell.alignment = amount_alignment
    
    # Add metadata row
    meta_row = summary_row + 2
    ws.cell(row=meta_row, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}")
    ws.cell(row=meta_row, column=4, value=f"Payment Type: {payment_type}")
    ws.cell(row=meta_row, column=7, value="HDFC Bank Bulk Format")
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, total_amount


@router.get("/bank-redeem")
async def export_bank_redeem_hdfc(
    status: str = Query("approved", description="Filter by status: pending, approved, rejected, paid"),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(500, le=2000)
):
    """
    Export Bank Redeem requests in HDFC Bulk Payment format
    
    - Default exports 'approved' requests ready for payment
    - Use status='approved' for requests approved but not yet paid
    """
    try:
        query = {"status": status}
        
        if date_from:
            query["created_at"] = {"$gte": date_from}
        if date_to:
            if "created_at" in query:
                query["created_at"]["$lte"] = date_to
            else:
                query["created_at"] = {"$lte": date_to}
        
        # Fetch requests with user details
        requests = await db.bank_redeem_requests.find(query).sort("created_at", -1).limit(limit).to_list(limit)
        
        if not requests:
            raise HTTPException(status_code=404, detail="No requests found for export")
        
        # Enrich with user bank details
        export_data = []
        for req in requests:
            uid = req.get("uid", req.get("user_id"))
            user = await db.users.find_one({"uid": uid}, {"_id": 0, "email": 1, "mobile": 1, "phone": 1, "bank_details": 1, "name": 1})
            
            export_item = {
                "request_id": req.get("request_id", str(req.get("_id", ""))),
                "uid": uid,
                "user_name": req.get("user_name", user.get("name", "") if user else ""),
                "email": user.get("email", "") if user else "",
                "mobile": user.get("mobile", user.get("phone", "")) if user else "",
                "bank_details": req.get("bank_details", user.get("bank_details", {}) if user else {}),
                "amount_inr": req.get("amount_inr", req.get("amount", 0)),
                "net_amount": req.get("net_amount", 0),
                "created_at": req.get("created_at", "")
            }
            export_data.append(export_item)
        
        # Generate Excel
        excel_file, total = generate_hdfc_excel(export_data, "BANK_REDEEM")
        
        # Generate filename
        date_str = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"HDFC_BankRedeem_{status}_{date_str}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Total-Records": str(len(export_data)),
                "X-Total-Amount": str(total)
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error exporting bank redeem: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/emi-payments")
async def export_emi_payments_hdfc(
    status: str = Query("approved", description="Filter by status"),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(500, le=2000)
):
    """
    Export EMI Payment requests in HDFC Bulk Payment format
    """
    try:
        query = {"status": status}
        
        if date_from:
            query["created_at"] = {"$gte": date_from}
        if date_to:
            if "created_at" in query:
                query["created_at"]["$lte"] = date_to
            else:
                query["created_at"] = {"$lte": date_to}
        
        # Fetch EMI requests
        requests = await db.emi_requests.find(query).sort("created_at", -1).limit(limit).to_list(limit)
        
        if not requests:
            # Try bill_payment_requests with emi type
            requests = await db.bill_payment_requests.find({
                **query,
                "payment_type": {"$in": ["emi", "EMI"]}
            }).sort("created_at", -1).limit(limit).to_list(limit)
        
        if not requests:
            raise HTTPException(status_code=404, detail="No EMI requests found for export")
        
        # Enrich with user details
        export_data = []
        for req in requests:
            uid = req.get("uid", req.get("user_id"))
            user = await db.users.find_one({"uid": uid}, {"_id": 0, "email": 1, "mobile": 1, "phone": 1, "bank_details": 1, "name": 1})
            
            export_item = {
                "request_id": req.get("request_id", str(req.get("_id", ""))),
                "uid": uid,
                "user_name": req.get("user_name", user.get("name", "") if user else ""),
                "email": user.get("email", "") if user else "",
                "mobile": user.get("mobile", user.get("phone", "")) if user else "",
                "bank_details": req.get("bank_details", user.get("bank_details", {}) if user else {}),
                "amount_inr": req.get("amount_inr", req.get("amount", 0)),
                "net_amount": req.get("net_amount", req.get("amount", 0)),
                "created_at": req.get("created_at", "")
            }
            export_data.append(export_item)
        
        # Generate Excel
        excel_file, total = generate_hdfc_excel(export_data, "EMI_PAY")
        
        date_str = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"HDFC_EMI_{status}_{date_str}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Total-Records": str(len(export_data)),
                "X-Total-Amount": str(total)
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error exporting EMI payments: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/savings-vault")
async def export_savings_vault_hdfc(
    status: str = Query("approved", description="Filter by status: pending, approved, rejected, completed"),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(500, le=2000)
):
    """
    Export Savings Vault (RD) Redeem requests in HDFC Bulk Payment format
    """
    try:
        query = {"status": status}
        
        if date_from:
            query["created_at"] = {"$gte": date_from}
        if date_to:
            if "created_at" in query:
                query["created_at"]["$lte"] = date_to
            else:
                query["created_at"] = {"$lte": date_to}
        
        # Fetch RD redeem requests
        requests = await db.rd_redeem_requests.find(query).sort("created_at", -1).limit(limit).to_list(limit)
        
        if not requests:
            raise HTTPException(status_code=404, detail="No Savings Vault redeem requests found for export")
        
        # Enrich with user details
        export_data = []
        for req in requests:
            uid = req.get("uid", req.get("user_id"))
            user = await db.users.find_one({"uid": uid}, {"_id": 0, "email": 1, "mobile": 1, "phone": 1, "bank_details": 1, "name": 1})
            
            export_item = {
                "request_id": req.get("request_id", req.get("rd_id", str(req.get("_id", "")))),
                "uid": uid,
                "user_name": req.get("user_name", user.get("name", "") if user else ""),
                "email": user.get("email", "") if user else "",
                "mobile": user.get("mobile", user.get("phone", "")) if user else "",
                "bank_details": req.get("bank_details", user.get("bank_details", {}) if user else {}),
                "amount_inr": req.get("amount_inr", 0),
                "net_amount": req.get("net_amount", 0),
                "created_at": req.get("created_at", "")
            }
            export_data.append(export_item)
        
        # Generate Excel
        excel_file, total = generate_hdfc_excel(export_data, "SAVINGS_VAULT")
        
        date_str = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"HDFC_SavingsVault_{status}_{date_str}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Total-Records": str(len(export_data)),
                "X-Total-Amount": str(total)
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error exporting savings vault: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/combined")
async def export_combined_hdfc(
    status: str = Query("approved", description="Filter by status"),
    include_bank_redeem: bool = True,
    include_emi: bool = True,
    include_savings: bool = True,
    limit: int = Query(500, le=2000)
):
    """
    Export ALL payment types combined in single HDFC Bulk Payment file
    Useful for processing all approved payments at once
    """
    try:
        all_data = []
        
        # Bank Redeem
        if include_bank_redeem:
            bank_requests = await db.bank_redeem_requests.find({"status": status}).limit(limit).to_list(limit)
            for req in bank_requests:
                uid = req.get("uid", req.get("user_id"))
                user = await db.users.find_one({"uid": uid}, {"_id": 0, "email": 1, "mobile": 1, "bank_details": 1, "name": 1})
                all_data.append({
                    "request_id": f"BR-{str(req.get('request_id', ''))[-8:]}",
                    "uid": uid,
                    "user_name": req.get("user_name", user.get("name", "") if user else ""),
                    "email": user.get("email", "") if user else "",
                    "mobile": user.get("mobile", "") if user else "",
                    "bank_details": req.get("bank_details", user.get("bank_details", {}) if user else {}),
                    "amount_inr": req.get("amount_inr", req.get("amount", 0)),
                    "payment_type": "BANK_REDEEM"
                })
        
        # EMI
        if include_emi:
            emi_requests = await db.bill_payment_requests.find({
                "status": status,
                "payment_type": {"$in": ["emi", "EMI", "loan_emi"]}
            }).limit(limit).to_list(limit)
            for req in emi_requests:
                uid = req.get("uid", req.get("user_id"))
                user = await db.users.find_one({"uid": uid}, {"_id": 0, "email": 1, "mobile": 1, "bank_details": 1, "name": 1})
                all_data.append({
                    "request_id": f"EMI-{str(req.get('request_id', ''))[-8:]}",
                    "uid": uid,
                    "user_name": req.get("user_name", user.get("name", "") if user else ""),
                    "email": user.get("email", "") if user else "",
                    "mobile": user.get("mobile", "") if user else "",
                    "bank_details": req.get("bank_details", user.get("bank_details", {}) if user else {}),
                    "amount_inr": req.get("amount_inr", req.get("amount", 0)),
                    "payment_type": "EMI_PAY"
                })
        
        # Savings Vault
        if include_savings:
            rd_requests = await db.rd_redeem_requests.find({"status": status}).limit(limit).to_list(limit)
            for req in rd_requests:
                uid = req.get("uid", req.get("user_id"))
                user = await db.users.find_one({"uid": uid}, {"_id": 0, "email": 1, "mobile": 1, "bank_details": 1, "name": 1})
                all_data.append({
                    "request_id": f"SV-{str(req.get('rd_id', req.get('request_id', '')))[-8:]}",
                    "uid": uid,
                    "user_name": req.get("user_name", user.get("name", "") if user else ""),
                    "email": user.get("email", "") if user else "",
                    "mobile": user.get("mobile", "") if user else "",
                    "bank_details": req.get("bank_details", user.get("bank_details", {}) if user else {}),
                    "amount_inr": req.get("amount_inr", req.get("net_amount", 0)),
                    "payment_type": "SAVINGS_VAULT"
                })
        
        if not all_data:
            raise HTTPException(status_code=404, detail="No requests found for export")
        
        # Generate Excel
        excel_file, total = generate_hdfc_excel(all_data, "COMBINED")
        
        date_str = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"HDFC_AllPayments_{status}_{date_str}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Total-Records": str(len(all_data)),
                "X-Total-Amount": str(total)
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error exporting combined payments: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/preview")
async def preview_export_data(
    payment_type: str = Query(..., description="Type: bank_redeem, emi, savings_vault, combined"),
    status: str = Query("approved"),
    limit: int = Query(10, le=50)
):
    """
    Preview export data before downloading
    Returns JSON with sample records and summary
    """
    try:
        data = []
        
        if payment_type in ["bank_redeem", "combined"]:
            requests = await db.bank_redeem_requests.find({"status": status}).limit(limit).to_list(limit)
            for req in requests:
                data.append({
                    "type": "Bank Redeem",
                    "request_id": req.get("request_id", ""),
                    "amount": req.get("amount_inr", 0),
                    "user": req.get("user_name", ""),
                    "bank": req.get("bank_details", {}).get("bank_name", "N/A")
                })
        
        if payment_type in ["emi", "combined"]:
            requests = await db.bill_payment_requests.find({
                "status": status,
                "payment_type": {"$in": ["emi", "EMI", "loan_emi"]}
            }).limit(limit).to_list(limit)
            for req in requests:
                data.append({
                    "type": "EMI Payment",
                    "request_id": req.get("request_id", ""),
                    "amount": req.get("amount_inr", req.get("amount", 0)),
                    "user": req.get("user_name", ""),
                    "bank": req.get("bank_details", {}).get("bank_name", "N/A")
                })
        
        if payment_type in ["savings_vault", "combined"]:
            requests = await db.rd_redeem_requests.find({"status": status}).limit(limit).to_list(limit)
            for req in requests:
                data.append({
                    "type": "Savings Vault",
                    "request_id": req.get("rd_id", req.get("request_id", "")),
                    "amount": req.get("amount_inr", 0),
                    "user": req.get("user_name", ""),
                    "bank": req.get("bank_details", {}).get("bank_name", "N/A")
                })
        
        total_amount = sum(item.get("amount", 0) for item in data)
        
        return {
            "preview": data,
            "summary": {
                "total_records": len(data),
                "total_amount": round(total_amount, 2),
                "status_filter": status,
                "payment_type": payment_type
            },
            "download_url": f"/api/admin/hdfc-export/{payment_type.replace('_', '-')}?status={status}"
        }
        
    except Exception as e:
        logging.error(f"Error previewing export: {e}")
        raise HTTPException(status_code=500, detail=str(e))
