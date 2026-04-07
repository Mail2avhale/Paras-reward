"""
PARAS REWARD - EKO BBPS MULTI SERVICE BACKEND
============================================
Clean implementation following official EKO documentation with standard error handling.

Services:
1. Electricity Bill Payment
2. DTH Recharge
3. FASTag Recharge
4. Loan / EMI Payment
5. Mobile Recharge
6. Water Bill
7. Credit Card
8. Insurance
9. And more...

Error Handling follows Eko Developer Documentation:
- HTTP Response Codes (200, 403, 404, 405, 415, 500)
- Eko Status Codes (0=Success, 463=User not found, 347=Insufficient balance, etc.)
- Transaction Status (tx_status: 0=Success, 1=Failed, 2=Pending, 3=Refund Pending, 4=Refunded, 5=On Hold)
"""

from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from pydantic import BaseModel, validator
from typing import Optional, Dict, Any, List
import httpx  # ASYNC HTTP - replaces blocking 'requests' library
import base64
import hashlib
import hmac
import time
import json
import logging
import re
import os
from datetime import datetime, timezone, timedelta

# Eko error handler was removed - define necessary constants/functions inline

# Database reference (set by main server)
db = None
check_redeem_limit_func = None

def set_db(database):
    global db
    db = database

def set_redeem_limit_check(func):
    global check_redeem_limit_func
    check_redeem_limit_func = func

# Cooldown check for BBPS services (24 hours)
async def check_bbps_cooldown(user_id: str) -> dict:
    """Check if user can make a BBPS bill payment request (24 hour cooldown)"""
    if not db or not user_id:
        return {"allowed": True, "wait_hours": 0}
    

# ==================== SUBSCRIPTION CHECK FOR REDEEM SERVICES ====================
async def check_subscription_for_redeem(user_id: str, service_name: str = "bill payment") -> dict:
    """
    CRITICAL: Check if user has active paid subscription before allowing redeem.
    Only users with active startup/growth/elite can use redeem services.
    """
    if not db or not user_id:
        return {"allowed": False, "reason": "Invalid user"}
    
    user = await db.users.find_one({"uid": user_id})
    if not user:
        return {"allowed": False, "reason": "User not found"}
    
    # Check subscription plan
    subscription_plan = (user.get("subscription_plan") or "explorer").lower()
    
    # Free/Explorer users cannot redeem
    if subscription_plan in ["explorer", "free", "", None]:
        return {
            "allowed": False,
            "reason": f"Paid subscription required for {service_name}. Please upgrade to Startup, Growth or Elite plan.",
            "requires_subscription": True
        }
    
    # Check if subscription is expired
    expiry = user.get("subscription_expiry") or user.get("subscription_expires") or user.get("vip_expiry")
    if expiry:
        try:
            if isinstance(expiry, str):
                expiry_dt = datetime.fromisoformat(expiry.replace('Z', '+00:00'))
            else:
                expiry_dt = expiry
            
            if expiry_dt.tzinfo is None:
                expiry_dt = expiry_dt.replace(tzinfo=timezone.utc)
            
            now = datetime.now(timezone.utc)
            
            if expiry_dt < now:
                days_expired = (now - expiry_dt).days
                return {
                    "allowed": False,
                    "reason": f"Your subscription expired {days_expired} days ago. Please renew to use {service_name}.",
                    "is_expired": True,
                    "days_expired": days_expired
                }
        except Exception as e:
            logging.warning(f"[SUBSCRIPTION-CHECK] Expiry parse error for {user_id}: {e}")
    
    return {"allowed": True, "subscription_plan": subscription_plan}
    
    cooldown_hours = 24
    cooldown_delta = timedelta(hours=cooldown_hours)
    now = datetime.now(timezone.utc)
    cutoff_time = now - cooldown_delta
    
    # Find last successful/pending request
    last_request = await db.bill_payment_requests.find_one(
        {
            "user_id": user_id,
            "status": {"$nin": ["rejected", "failed", "cancelled"]}
        },
        sort=[("created_at", -1)]
    )
    
    if not last_request:
        return {"allowed": True, "wait_hours": 0}
    
    last_time = last_request.get("created_at")
    if isinstance(last_time, str):
        try:
            last_time = datetime.fromisoformat(last_time.replace('Z', '+00:00'))
        except:
            return {"allowed": True, "wait_hours": 0}
    
    if not last_time:
        return {"allowed": True, "wait_hours": 0}
    
    if last_time.tzinfo is None:
        last_time = last_time.replace(tzinfo=timezone.utc)
    
    if last_time > cutoff_time:
        time_passed = now - last_time
        remaining = cooldown_delta - time_passed
        remaining_hours = remaining.total_seconds() / 3600
        return {
            "allowed": False,
            "wait_hours": round(remaining_hours, 1),
            "last_request": last_time.isoformat()
        }
    
    return {"allowed": True, "wait_hours": 0}
EKO_ERROR_MESSAGES = {
    403: "Authentication failed",
    404: "Service not found",
    500: "Internal server error"
}

TX_STATUS_MESSAGES = {
    0: "Success",
    1: "Failed",
    2: "Pending",
    3: "Refund Pending",
    4: "Refunded",
    5: "Hold"
}

class EkoTxStatus:
    SUCCESS = 0
    FAILED = 1
    PENDING = 2
    REFUND_PENDING = 3
    REFUNDED = 4
    HOLD = 5

def sanitize_sender_name(name: str) -> str:
    """
    Sanitize sender name for Eko BBPS API.
    
    Eko BBPS requirement: "Sender Name should contain only letters"
    - Remove ALL numbers, special characters, AND SPACES
    - Keep only alphabets (a-z, A-Z)
    - Return "Customer" as fallback if empty
    
    Reference: https://developers.eko.in/v1/docs/bbps-faqs
    """
    if not name:
        return "Customer"
    
    # Remove EVERYTHING except letters (a-z, A-Z) - NO SPACES
    sanitized = re.sub(r'[^a-zA-Z]', '', name)
    
    # If result is empty after sanitization, return default
    if not sanitized or len(sanitized) == 0:
        return "Customer"
    
    # Limit to reasonable length (Eko may have max length)
    return sanitized[:50]


def get_common_error_message(msg_or_code):
    """Get common error message with better fallback handling"""
    # If it's a known error code
    if isinstance(msg_or_code, int):
        return EKO_ERROR_MESSAGES.get(msg_or_code, f"Error code: {msg_or_code}")
    
    # Handle specific Eko error messages
    if isinstance(msg_or_code, str):
        msg_lower = msg_or_code.lower()
        
        # "No key for Response" - operator doesn't support bill fetch
        if "no key for response" in msg_lower:
            return "This provider doesn't support automatic bill fetch. Please enter the amount manually."
        
        # Consumer validation errors
        if "invalid consumer" in msg_lower or "consumer not found" in msg_lower:
            return "Invalid consumer number. Please verify and try again."
        
        if "consumer belongs to prepaid" in msg_lower:
            return "This number is registered as prepaid, not postpaid."
        
        # Service unavailable
        if "service temporarily" in msg_lower or "try again later" in msg_lower:
            return "Service temporarily unavailable. Please try again later."
        
        # Return the original message if no specific handling
        return msg_or_code
    
    return "Unknown error"

def log_eko_transaction(data):
    """Log Eko transaction"""
    logging.info(f"[BBPS] Transaction: {data}")

router = APIRouter(prefix="/bbps", tags=["BBPS Services"])

# Global async HTTP client
_bbps_http_client: Optional[httpx.AsyncClient] = None

def get_bbps_http_client() -> httpx.AsyncClient:
    """Get or create async HTTP client"""
    global _bbps_http_client
    if _bbps_http_client is None:
        _bbps_http_client = httpx.AsyncClient(timeout=httpx.Timeout(60.0, connect=10.0))
    return _bbps_http_client

async def bbps_get(url: str, headers: dict = None, timeout: int = 60) -> httpx.Response:
    """Non-blocking async GET request for BBPS"""
    client = get_bbps_http_client()
    return await client.get(url, headers=headers, timeout=timeout)

async def bbps_post(url: str, headers: dict = None, data: dict = None, json_body: dict = None, timeout: int = 60) -> httpx.Response:
    """Non-blocking async POST request for BBPS - supports both form data and JSON"""
    client = get_bbps_http_client()
    if json_body:
        return await client.post(url, headers=headers, json=json_body, timeout=timeout)
    return await client.post(url, headers=headers, data=data, timeout=timeout)


# ==================== EKO PRODUCTION CONFIG (ALL FROM ENV) ====================

BASE_URL = os.environ.get("EKO_BASE_URL", "https://api.eko.in:25002/ekoicici")
DEVELOPER_KEY = os.environ.get("EKO_DEVELOPER_KEY")
INITIATOR_ID = os.environ.get("EKO_INITIATOR_ID")
AUTH_KEY = os.environ.get("EKO_AUTHENTICATOR_KEY")
USER_CODE = os.environ.get("EKO_USER_CODE")
SOURCE_IP = os.environ.get("EKO_SOURCE_IP", "34.44.149.98")
DEFAULT_LATLONG = "19.9975,73.7898"

# Request timeout in seconds
REQUEST_TIMEOUT = 60


def validate_bbps_config():
    """Validate all required Eko configuration is present"""
    missing = []
    if not DEVELOPER_KEY:
        missing.append("EKO_DEVELOPER_KEY")
    if not INITIATOR_ID:
        missing.append("EKO_INITIATOR_ID")
    if not AUTH_KEY:
        missing.append("EKO_AUTHENTICATOR_KEY")
    if not USER_CODE:
        missing.append("EKO_USER_CODE")
    
    if missing:
        logging.error(f"[BBPS] Missing required environment variables: {', '.join(missing)}")
        return False
    return True


# ==================== AUTHENTICATION ====================

def generate_headers() -> Dict[str, str]:
    """
    Generate authentication headers as per EKO BBPS documentation.
    
    BBPS uses DIFFERENT algorithm than DMT:
    1. timestamp = current time in MILLISECONDS (not seconds like DMT)
    2. encoded_key = Base64(authenticator_key) as BYTES
    3. secret_key = Base64(HMAC_SHA256(encoded_key, timestamp))
    
    Reference: https://developers.eko.in/reference/pay-bills-api
    """
    timestamp = str(round(time.time() * 1000))  # MILLISECONDS for BBPS
    
    # encoded_key must be BYTES for hmac
    encoded_key = base64.b64encode(AUTH_KEY.encode())  # Returns bytes
    
    secret_key = base64.b64encode(
        hmac.new(
            encoded_key,  # Use bytes directly
            timestamp.encode(),
            hashlib.sha256
        ).digest()
    ).decode()
    
    return {
        "developer_key": DEVELOPER_KEY,
        "secret-key": secret_key,
        "secret-key-timestamp": timestamp,
        "Content-Type": "application/json"
    }


def generate_headers_for_payment(timestamp: str) -> Dict[str, str]:
    """
    Generate authentication headers for bill PAYMENT.
    Exactly as per Eko documentation.
    """
    # encoded_key must be BYTES for hmac, not string!
    encoded_key = base64.b64encode(AUTH_KEY.encode())  # Returns bytes
    
    secret_key = base64.b64encode(
        hmac.new(
            encoded_key,  # Use bytes directly
            timestamp.encode(),
            hashlib.sha256
        ).digest()
    ).decode()
    
    return {
        "developer_key": DEVELOPER_KEY,
        "secret-key": secret_key,
        "secret-key-timestamp": timestamp,
        "Content-Type": "application/json",
        "Connection": "Keep-Alive",
        "Accept-Encoding": "gzip",
        "User-Agent": "okhttp/3.9.0"
    }


def generate_request_hash(timestamp: str, account: str, amount: str) -> str:
    """
    Generate request_hash for payment transactions.
    
    Formula: Base64(HMAC_SHA256(encoded_key, timestamp + account + amount + user_code))
    
    IMPORTANT: encoded_key must be BYTES for hmac!
    """
    # encoded_key must be BYTES for hmac
    encoded_key = base64.b64encode(AUTH_KEY.encode())  # Returns bytes
    
    concatenated = f"{timestamp}{account}{amount}{USER_CODE}"
    
    request_hash = base64.b64encode(
        hmac.new(
            encoded_key,  # Use bytes directly
            concatenated.encode(),
            hashlib.sha256
        ).digest()
    ).decode()
    
    return request_hash


# ==================== REQUEST MODELS ====================

class FetchBillRequest(BaseModel):
    """Request model for bill fetch - supports operator-specific parameters"""
    operator_id: str
    account: str          # utility_acc_no (or last 4 digits for Credit Card)
    mobile: str           # confirmation_mobile_no
    sender_name: Optional[str] = "Customer"
    # Operator-specific optional parameters
    source_ip: Optional[str] = None
    postalcode: Optional[str] = None  # For MSEB electricity
    cycle_number: Optional[str] = None  # For MSEB
    authenticator: Optional[str] = None  # For MSEB
    dob: Optional[str] = None  # For LIC - DD/MM/YYYY format
    mobile_number: Optional[str] = None  # For Credit Card BBPS - registered mobile
    extra_params: Optional[Dict[str, str]] = None  # For any other operator-specific params
    
    @validator('mobile')
    def validate_mobile(cls, v):
        if not v or not v.isdigit() or len(v) != 10:
            raise ValueError('Mobile number must be 10 digits')
        return v
    
    @validator('account')
    def validate_account(cls, v):
        if not v or not v.strip():
            raise ValueError('Account number is required')
        return v.strip()


class PayBillRequest(BaseModel):
    """Request model for bill payment"""
    operator_id: str
    account: str
    amount: str
    mobile: str
    user_id: Optional[str] = None  # For cooldown check
    sender_name: Optional[str] = "Customer"
    bill_fetch_response: Optional[str] = None  # Required when fetchBill=1
    payment_amount_breakup: Optional[str] = None  # For Credit Card BBPS - JSON string with billid and amount
    hc_channel: Optional[str] = None  # For Credit Card BBPS only - don't send by default
    # Operator-specific parameters
    recharge_plan_id: Optional[str] = None  # Required for Jio Prepaid (operator_id=90)
    cycle_number: Optional[str] = None  # Required for MSEDCL (operator_id=62) - BU number
    registered_mobile_number: Optional[str] = None  # Required for Credit Card BBPS operators
    extra_params: Optional[Dict[str, str]] = None  # For any other operator-specific params
    
    @validator('mobile')
    def validate_mobile(cls, v):
        if not v or not v.isdigit() or len(v) != 10:
            raise ValueError('Mobile number must be 10 digits')
        return v
    
    @validator('amount')
    def validate_amount(cls, v):
        try:
            amt = float(v)
            if amt <= 0:
                raise ValueError('Amount must be greater than 0')
            if amt > 100000:
                raise ValueError('Maximum amount is ₹1,00,000')
        except (ValueError, TypeError):
            raise ValueError('Invalid amount')
        return v


# ==================== STANDARD RESPONSE FORMAT ====================

def create_success_response(data: Dict, message: str = "Success") -> Dict:
    """Create standardized success response"""
    return {
        "success": True,
        "status": "SUCCESS",
        "message": message,
        "data": data
    }


def create_error_response(
    error_code: int,
    message: str,
    user_message: str = None,
    data: Dict = None
) -> Dict:
    """Create standardized error response"""
    return {
        "success": False,
        "status": "FAILED",
        "error_code": error_code,
        "message": message,
        "user_message": user_message or EKO_ERROR_MESSAGES.get(error_code, message),
        "data": data or {}
    }


def create_pending_response(
    tx_status: int,
    tid: str,
    message: str,
    data: Dict = None
) -> Dict:
    """Create response for pending/processing transactions"""
    status_map = {
        2: "PENDING",
        3: "REFUND_PENDING",
        5: "ON_HOLD"
    }
    return {
        "success": True,
        "status": status_map.get(tx_status, "PROCESSING"),
        "tx_status": tx_status,
        "tid": tid,
        "message": message,
        "user_message": TX_STATUS_MESSAGES.get(tx_status, message),
        "requires_status_check": True,
        "data": data or {}
    }


# ==================== HEALTH CHECK ====================

@router.get("/health")
def health():
    """Health check endpoint"""
    return {
        "status": "PARAS REWARD BBPS RUNNING",
        "version": "2.1",
        "services": ["electricity", "dth", "fastag", "emi", "mobile_prepaid", "water", "credit_card", "insurance"]
    }


# ==================== GET OPERATOR PARAMETERS API ====================

@router.get("/wallet-balance")
async def get_eko_wallet_balance():
    """
    Get Eko wallet/settlement account balance.
    Used by admin dashboard to monitor available funds for BBPS payments.
    
    Reference: https://developers.eko.in/reference/wallet-balance-inquiry
    """
    if not validate_bbps_config():
        return {"success": False, "balance": 0, "error": "Service configuration error"}
    
    try:
        timestamp = str(round(time.time() * 1000))
        
        # Generate secret key
        encoded_key = base64.b64encode(AUTH_KEY.encode())
        secret_key = base64.b64encode(
            hmac.new(encoded_key, timestamp.encode(), hashlib.sha256).digest()
        ).decode()
        
        # Correct endpoint: /v2/customers/mobile_number:{initiator_id}/balance
        url = f"{BASE_URL}/v2/customers/mobile_number:{INITIATOR_ID}/balance"
        
        headers = {
            "developer_key": DEVELOPER_KEY,
            "secret-key": secret_key,
            "secret-key-timestamp": timestamp
        }
        
        params = {
            "initiator_id": INITIATOR_ID,
            "user_code": USER_CODE
        }
        
        logging.info("[EKO BALANCE] Checking wallet balance")
        
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(url, headers=headers, params=params)
        
        logging.info(f"[EKO BALANCE] HTTP Status: {response.status_code}")
        logging.info(f"[EKO BALANCE] Response: {response.text[:500] if response.text else 'empty'}")
        
        if response.status_code != 200:
            return {"success": False, "balance": 0, "error": f"HTTP {response.status_code}"}
        
        result = response.json()
        
        # Check for success (status 0 or response_status_id 0)
        if result.get("status") == 0 or result.get("response_status_id") == 0:
            balance_str = result.get("data", {}).get("balance", "0")
            try:
                balance = float(balance_str)
            except (ValueError, TypeError):
                balance = 0
            
            return {
                "success": True,
                "balance": balance,
                "locked": result.get("data", {}).get("locked_amount", 0),
                "currency": result.get("data", {}).get("currency", "INR"),
                "last_used_okekey": result.get("data", {}).get("last_used_okekey"),
                "message": result.get("message")
            }
        else:
            return {
                "success": False,
                "balance": 0,
                "error": result.get("message", "Failed to get balance"),
                "raw": result
            }
            
    except Exception as e:
        logging.error(f"[EKO BALANCE] Error: {str(e)}")
        return {"success": False, "balance": 0, "error": str(e)}


@router.get("/operator-params/{operator_id}")
async def get_operator_parameters(operator_id: str):
    """
    Get required parameters for a specific operator.
    MUST call this before fetch/pay to know what params are needed.
    
    Reference: https://developers.eko.in/reference/bbps-operator-parameters
    
    Returns:
    - param_name: Name of parameter to pass in fetch/pay API
    - param_label: Label to show to user
    - param_type: Numeric, Decimal, AlphaNumeric, or List
    - regex: Validation regex
    - error_message: Error message if validation fails
    - fetchBill: 1 = Must call fetch before pay, 0 = Can pay directly
    """
    if not validate_bbps_config():
        return create_error_response(500, "Service configuration error", "Service temporarily unavailable.")
    
    try:
        url = f"{BASE_URL}/v2/billpayments/operators/{operator_id}"
        headers = generate_headers()
        
        logging.info(f"[BBPS OPERATOR PARAMS] Getting params for operator_id={operator_id}")
        
        response = await bbps_get(url, headers=headers, timeout=REQUEST_TIMEOUT)
        
        logging.info(f"[BBPS OPERATOR PARAMS] HTTP Status: {response.status_code}")
        logging.info(f"[BBPS OPERATOR PARAMS] Response: {response.text[:500] if response.text else 'empty'}")
        
        if response.status_code != 200:
            return create_error_response(
                response.status_code,
                f"Failed to get operator parameters: {response.text}",
                "Could not fetch operator details. Please try again."
            )
        
        result = response.json()
        
        # Return full response for debugging (Eko returns different status codes)
        eko_status = result.get("status")
        
        # Handle various Eko response patterns
        if eko_status == 0 or result.get("operator_name"):
            return {
                "success": True,
                "operator_id": operator_id,
                "operator_name": result.get("operator_name"),
                "parameters": result.get("data", []),
                "fetch_bill_required": result.get("fetchBill", 0) == 1,
                "is_bbps": result.get("BBPS", 0) == 1,
                "raw_response": result
            }
        else:
            # Return raw response for debugging
            return {
                "success": False,
                "operator_id": operator_id,
                "error_code": eko_status,
                "message": result.get("message", "Unknown error"),
                "raw_response": result
            }
    
    except Exception as e:
        logging.error(f"[BBPS OPERATOR PARAMS] Error: {e}")
        return create_error_response(500, str(e), "Service temporarily unavailable.")


# ==================== FETCH BILL ====================

@router.post("/fetch")
async def fetch_bill(data: FetchBillRequest):
    """
    Fetch bill details from EKO BBPS API.
    
    CRITICAL: Uses POST /v2/billpayments/fetchbill (NOT GET v3!)
    Testing confirmed that v3 GET returns 500 error while v2 POST works correctly.
    
    Works for: Electricity, DTH, FASTag, EMI, Water, etc.
    """
    if not validate_bbps_config():
        return create_error_response(500, "Service configuration error", "Service temporarily unavailable.")
    
    client_ref_id = f"FETCH{int(time.time() * 1000)}"
    
    try:
        # Build request body for POST v2 endpoint
        # Reference: Testing showed POST v2 returns HTTP 200 while GET v3 returns 500
        # CRITICAL: Sanitize sender_name - Eko requires "only letters" (no numbers/special chars)
        clean_sender_name = sanitize_sender_name(data.sender_name)
        
        request_body = {
            "user_code": USER_CODE,
            "client_ref_id": client_ref_id,
            "utility_acc_no": data.account,
            "confirmation_mobile_no": data.mobile,
            "sender_name": clean_sender_name,
            "operator_id": data.operator_id,
            "source_ip": data.source_ip or SOURCE_IP,
            "latlong": DEFAULT_LATLONG
        }
        
        # Add operator-specific parameters if provided
        if data.postalcode:
            request_body["postalcode"] = data.postalcode
        if data.cycle_number:
            request_body["cycle_number"] = data.cycle_number
        if data.authenticator:
            request_body["authenticator"] = data.authenticator
        if data.dob:
            request_body["dob"] = data.dob
        
        # For Credit Card BBPS - registered mobile number
        if data.mobile_number:
            request_body["mobile_number"] = data.mobile_number
        
        # Add any extra operator-specific params
        if data.extra_params:
            for key, value in data.extra_params.items():
                if key not in request_body:
                    request_body[key] = value
        
        # Use POST v2 endpoint (confirmed working in test)
        # Endpoint: /v2/billpayments/fetchbill
        url = f"{BASE_URL}/v2/billpayments/fetchbill?initiator_id={INITIATOR_ID}"
        
        logging.info(f"[BBPS FETCH] POST {url}")
        logging.info(f"[BBPS FETCH] operator={data.operator_id}, account=***{data.account[-4:]}")
        logging.info(f"[BBPS FETCH] Body keys: {list(request_body.keys())}")
        
        headers = generate_headers()
        
        # Use POST request with JSON body
        response = await bbps_post(url, headers=headers, json_body=request_body, timeout=REQUEST_TIMEOUT)
        
        logging.info(f"[BBPS FETCH] HTTP Status: {response.status_code}")
        logging.info(f"[BBPS FETCH] Response Body: {response.text[:500] if response.text else 'empty'}")
        
        # Handle HTTP-level errors
        if response.status_code == 403:
            logging.error("[BBPS FETCH] 403 Forbidden - Authentication failed")
            return create_error_response(
                403,
                "Authentication failed",
                "Service temporarily unavailable. Please try again later."
            )
        
        if response.status_code == 404:
            logging.error("[BBPS FETCH] 404 Not Found - Invalid endpoint")
            return create_error_response(404, "Service not found", "Service configuration error. Please contact support.")
        
        if response.status_code == 500:
            logging.error(f"[BBPS FETCH] 500 Server Error - Response: {response.text[:1000] if response.text else 'empty'}")
            # Try to parse error message from Eko
            try:
                error_data = response.json()
                error_msg = error_data.get("message", "Server error")
                logging.error(f"[BBPS FETCH] Eko Error: {error_msg}")
            except:
                error_msg = "Server error"
            return create_error_response(500, error_msg, "Eko server error. Please try again in a few minutes.")
        
        if response.status_code != 200:
            logging.error(f"[BBPS FETCH] Unexpected HTTP {response.status_code}: {response.text}")
            return create_error_response(
                response.status_code,
                response.text,
                f"Request failed. Please try again. (Error: {response.status_code})"
            )
        
        # Parse Eko response
        result = response.json()
        logging.info(f"[BBPS FETCH] Response: status={result.get('status')}, message={result.get('message')}")
        
        eko_status = result.get("status")
        eko_data = result.get("data", {})
        message = result.get("message", "")
        
        # SUCCESS: status = 0
        if eko_status == 0:
            # Handle billDetailsList if present (for EMI/Loan and some operators)
            bill_details_list = eko_data.get("billDetailsList", [])
            
            if bill_details_list and len(bill_details_list) > 0:
                # Use first bill from list
                first_bill = bill_details_list[0]
                bill_data = {
                    "bill_amount": first_bill.get("billAmount") or first_bill.get("netBillAmount") or eko_data.get("amount"),
                    "customer_name": first_bill.get("customer_name") or eko_data.get("utilitycustomername"),
                    "bill_date": first_bill.get("billDate") or eko_data.get("billdate"),
                    "due_date": first_bill.get("billDueDate") or eko_data.get("duedate"),
                    "bill_number": first_bill.get("billNumber") or eko_data.get("billnumber"),
                    "bill_fetch_response": eko_data.get("billfetchresponse"),
                    "operator_name": eko_data.get("billername") or eko_data.get("operator_name"),
                    "bbps_ref_id": first_bill.get("bharatBillReferenceNumber"),
                    "payment_status": first_bill.get("paymentStatus"),
                    "bill_period": first_bill.get("billperiod"),
                    "raw_response": result
                }
            else:
                # Standard response format
                bill_data = {
                    "bill_amount": eko_data.get("amount"),
                    "customer_name": eko_data.get("utilitycustomername"),
                    "bill_date": eko_data.get("billdate"),
                    "due_date": eko_data.get("duedate") or eko_data.get("billDueDate"),
                    "bill_number": eko_data.get("billnumber"),
                    "bill_fetch_response": eko_data.get("billfetchresponse"),
                    "operator_name": eko_data.get("billername") or eko_data.get("operator_name"),
                    "raw_response": result
                }
            
            logging.info(f"[BBPS FETCH] Success: amount={bill_data['bill_amount']}, customer={bill_data['customer_name']}")
            
            return {
                "success": True,
                "status": "SUCCESS",
                "message": "Bill fetched successfully",
                **bill_data
            }
        
        # ERROR: status != 0
        user_message = EKO_ERROR_MESSAGES.get(eko_status)
        if not user_message:
            user_message = get_common_error_message(message) if message else f"Unable to fetch bill (Error: {eko_status})"
        
        # Check for "no bill due" scenarios
        no_bill_keywords = ["no bill", "no due", "already paid", "no pending", "payment received"]
        is_no_bill = any(kw in message.lower() for kw in no_bill_keywords)
        
        if is_no_bill:
            return {
                "success": False,
                "status": "NO_BILL_DUE",
                "error_code": eko_status,
                "message": message,
                "user_message": "No pending bill found. You can enter amount manually for advance payment.",
                "raw_response": result
            }
        
        logging.warning(f"[BBPS FETCH] Error: status={eko_status}, message={message}")
        
        return {
            "success": False,
            "status": "FAILED",
            "error_code": eko_status,
            "message": message,
            "user_message": user_message,
            "raw_response": result
        }
        
    except httpx.TimeoutException:
        logging.error(f"[BBPS FETCH] Timeout after {REQUEST_TIMEOUT}s")
        return create_error_response(
            504,
            "Request timeout",
            "The service provider is taking too long to respond. Please try again."
        )
    
    except httpx.ConnectError as e:
        logging.error(f"[BBPS FETCH] Connection error: {e}")
        return create_error_response(
            503,
            "Connection failed",
            "Unable to connect to the service. Please check your internet and try again."
        )
    
    except Exception as e:
        logging.error(f"[BBPS FETCH] Unexpected error: {e}")
        return create_error_response(
            500,
            str(e),
            "An unexpected error occurred. Please try again."
        )


# ==================== PAY BILL ====================

@router.post("/pay")
async def pay_bill(data: PayBillRequest):
    """
    Pay bill via EKO BBPS API.
    
    Standard Process:
    1. Validate input parameters (amount, mobile, account)
    2. Generate authentication headers + request_hash
    3. Make API request to EKO
    4. Handle HTTP errors
    5. Parse Eko response status and tx_status
    6. Handle different transaction states:
       - tx_status=0: SUCCESS
       - tx_status=1: FAILED
       - tx_status=2: PENDING (requires status inquiry)
       - tx_status=3: REFUND_PENDING
       - tx_status=4: REFUNDED
       - tx_status=5: ON_HOLD (requires status inquiry)
    7. Return standardized response
    
    Works for: Electricity, DTH, FASTag, EMI, Water, Mobile, etc.
    """
    # ===== CRITICAL: CHECK SUBSCRIPTION FIRST =====
    if data.user_id:
        sub_check = await check_subscription_for_redeem(data.user_id, "bill payment")
        if not sub_check["allowed"]:
            return {
                "success": False,
                "status": 403,
                "message": sub_check["reason"],
                "data": None,
                "requires_subscription": sub_check.get("requires_subscription", False),
                "is_expired": sub_check.get("is_expired", False)
            }
    
    # ===== CHECK COOLDOWN: 24 hours between bill payments =====
    if data.user_id:
        cooldown = await check_bbps_cooldown(data.user_id)
        if not cooldown["allowed"]:
            return {
                "success": False,
                "status": 429,
                "message": f"Please wait {cooldown['wait_hours']:.0f} hours before making another bill payment.",
                "data": None
            }
    
    # ===== CHECK REDEEM LIMIT =====
    if data.user_id and check_redeem_limit_func and data.amount:
        try:
            limit_check = await check_redeem_limit_func(data.user_id, float(data.amount))
            if not limit_check.get("allowed"):
                limit_info = limit_check.get("limit_info", {})
                return {
                    "success": False,
                    "status": 403,
                    "message": f"Redeem limit exceeded. Your limit: {limit_info.get('total_limit', 0):,.0f} PRC, Used: {limit_info.get('total_redeemed', 0):,.0f} PRC, Remaining: {limit_info.get('remaining_limit', 0):,.0f} PRC",
                    "data": None
                }
        except Exception as e:
            logging.error(f"[BBPS PAY] Redeem limit check error: {e}")
    
    if not validate_bbps_config():
        return create_error_response(500, "Service configuration error", "Service temporarily unavailable.")
    
    client_ref_id = f"PAY{int(time.time() * 1000)}"
    
    # Pre-validation - just check required fields are present
    if not data.operator_id or not data.account or not data.amount or not data.mobile:
        return create_error_response(400, "Missing required fields", "Please fill all required fields")
    
    try:
        url = f"{BASE_URL}/v2/billpayments/paybill?initiator_id={INITIATOR_ID}"
        
        # Generate timestamp first for consistent hash generation
        timestamp = str(round(time.time() * 1000))
        
        # Generate headers with correct timestamp
        headers = generate_headers_for_payment(timestamp)
        
        # Generate request_hash using same timestamp
        request_hash = generate_request_hash(timestamp, data.account, data.amount)
        headers["request_hash"] = request_hash
        
        # CRITICAL: Sanitize sender_name - Eko requires "only letters" (no numbers/special chars)
        # Reference: https://developers.eko.in/v1/docs/bbps-faqs
        clean_sender_name = sanitize_sender_name(data.sender_name)
        
        body = {
            "initiator_id": INITIATOR_ID,  # Required in body as per Eko docs
            "source_ip": SOURCE_IP,  # Whitelisted production IP for Eko
            "user_code": USER_CODE,
            "amount": str(data.amount),  # String as per Eko docs
            "client_ref_id": client_ref_id,
            "utility_acc_no": str(data.account),
            "confirmation_mobile_no": str(data.mobile),
            "sender_name": clean_sender_name,
            "operator_id": str(data.operator_id),
            "latlong": DEFAULT_LATLONG
        }
        
        # Add bill_fetch_response if provided (required for fetchBill=1 operators)
        if data.bill_fetch_response:
            body["billfetchresponse"] = data.bill_fetch_response
        
        # Add Credit Card BBPS specific fields
        if data.payment_amount_breakup:
            body["payment_amount_breakup"] = data.payment_amount_breakup
        if data.hc_channel:
            body["hc_channel"] = data.hc_channel
        
        # Add Jio Prepaid specific field (operator_id=90 requires recharge_plan_id)
        if data.recharge_plan_id:
            body["recharge_plan_id"] = data.recharge_plan_id
        
        # Add MSEDCL (operator_id=62) specific field - BU number
        if data.cycle_number:
            body["cycle_number"] = data.cycle_number
        
        # Add Credit Card BBPS specific field - Registered mobile number
        # This is DIFFERENT from confirmation_mobile_no
        if data.registered_mobile_number:
            body["mobile_number"] = data.registered_mobile_number
        
        # Add any extra operator-specific params
        if data.extra_params:
            for key, value in data.extra_params.items():
                if key not in body and value:
                    body[key] = value
        
        logging.error(f"[BBPS PAY] client_ref={client_ref_id}, operator={data.operator_id}, amount={data.amount}")
        logging.error(f"[BBPS PAY] URL: {url}")
        logging.error(f"[BBPS PAY] Headers: developer_key={DEVELOPER_KEY[:10]}..., timestamp={timestamp}, request_hash={request_hash[:20]}...")
        logging.error(f"[BBPS PAY] Body: {body}")
        
        # Use JSON format for payment
        response = await bbps_post(url, headers=headers, json_body=body, timeout=REQUEST_TIMEOUT)
        
        logging.info(f"[BBPS PAY] HTTP Status: {response.status_code}")
        logging.info(f"[BBPS PAY] Response: {response.text[:500] if response.text else 'empty'}")
        
        # Handle HTTP-level errors
        if response.status_code == 403:
            logging.error("[BBPS PAY] 403 Forbidden - Authentication failed")
            return create_error_response(
                403,
                "Authentication failed",
                "Payment service temporarily unavailable. Your account has not been charged. Please try again later."
            )
        
        if response.status_code == 404:
            return create_error_response(404, "Service not found", "Payment service configuration error. Please contact support.")
        
        if response.status_code == 500:
            return create_error_response(
                500,
                "Server error",
                "Payment server is temporarily unavailable. Your account has not been charged. Please try again in a few minutes."
            )
        
        if response.status_code != 200:
            logging.error(f"[BBPS PAY] Unexpected HTTP {response.status_code}: {response.text}")
            return create_error_response(
                response.status_code,
                response.text,
                f"Payment request failed. Your account has not been charged. (Error: {response.status_code})"
            )
        
        # Parse Eko response
        result = response.json()
        eko_status = result.get("status")
        eko_data = result.get("data", {}) if isinstance(result.get("data"), dict) else {}
        message = result.get("message", "")
        
        # Extract transaction details
        tid = eko_data.get("tid")
        bbps_ref = eko_data.get("bbpstrxnrefid")
        tx_status = eko_data.get("tx_status")
        txstatus_desc = eko_data.get("txstatus_desc")
        
        # Convert tx_status to int if string
        if tx_status is not None and isinstance(tx_status, str):
            try:
                tx_status = int(tx_status)
            except ValueError:
                tx_status = None
        
        logging.info(f"[BBPS PAY] Response: status={eko_status}, tx_status={tx_status}, tid={tid}, message={message}")
        
        # SUCCESS: status = 0
        if eko_status == 0:
            # Check tx_status for final state
            if tx_status == EkoTxStatus.SUCCESS or tx_status == 0:
                # Transaction successful
                return {
                    "success": True,
                    "status": "SUCCESS",
                    "tx_status": 0,
                    "tid": tid,
                    "bbps_ref": bbps_ref,
                    "client_ref_id": client_ref_id,
                    "message": message or "Payment successful",
                    "user_message": "Your payment has been processed successfully!",
                    "amount": data.amount,
                    "raw_response": result
                }
            
            elif tx_status == EkoTxStatus.FAILED or tx_status == 1:
                # Transaction failed
                return {
                    "success": False,
                    "status": "FAILED",
                    "tx_status": 1,
                    "tid": tid,
                    "client_ref_id": client_ref_id,
                    "message": message or "Payment failed",
                    "user_message": txstatus_desc or "Transaction failed. Your account has not been charged.",
                    "raw_response": result
                }
            
            elif tx_status == EkoTxStatus.INITIATED or tx_status == 2:
                # Transaction pending/initiated
                return create_pending_response(
                    tx_status=2,
                    tid=tid,
                    message=message or "Payment is being processed",
                    data={"bbps_ref": bbps_ref, "raw_response": result}
                )
            
            elif tx_status == EkoTxStatus.REFUND_PENDING or tx_status == 3:
                # Refund pending
                return {
                    "success": False,
                    "status": "REFUND_PENDING",
                    "tx_status": 3,
                    "tid": tid,
                    "message": message or "Refund pending",
                    "user_message": "Transaction failed. Refund will be credited within 24-48 hours.",
                    "requires_status_check": True,
                    "raw_response": result
                }
            
            elif tx_status == EkoTxStatus.REFUNDED or tx_status == 4:
                # Already refunded
                return {
                    "success": False,
                    "status": "REFUNDED",
                    "tx_status": 4,
                    "tid": tid,
                    "message": message or "Amount refunded",
                    "user_message": "Transaction failed. Amount has been refunded to your account.",
                    "raw_response": result
                }
            
            elif tx_status == EkoTxStatus.ON_HOLD or tx_status == 5:
                # On hold - requires inquiry
                return {
                    "success": False,
                    "status": "ON_HOLD",
                    "tx_status": 5,
                    "tid": tid,
                    "message": message or "Transaction on hold",
                    "user_message": f"Transaction is on hold. Please contact support with Transaction ID: {tid}",
                    "requires_status_check": True,
                    "raw_response": result
                }
            
            else:
                # No tx_status or unknown - treat as pending
                if tid:
                    return create_pending_response(
                        tx_status=2,
                        tid=tid,
                        message=message or "Payment is being processed",
                        data={"bbps_ref": bbps_ref, "raw_response": result}
                    )
                else:
                    # Success without TID (rare case)
                    return {
                        "success": True,
                        "status": "SUCCESS",
                        "message": message or "Payment submitted",
                        "user_message": "Payment submitted successfully!",
                        "raw_response": result
                    }
        
        # ERROR: status != 0
        user_message = EKO_ERROR_MESSAGES.get(eko_status)
        if not user_message:
            user_message = get_common_error_message(message) if message else f"Payment failed (Error: {eko_status})"
        
        # Special handling for known errors
        if eko_status == 347:  # Insufficient balance
            user_message = "Insufficient balance in merchant account. Please contact support."
        elif eko_status == 944 or eko_status == 945:  # Limit exceeded
            user_message = "Transaction limit exceeded. Please try again tomorrow."
        elif eko_status == 544:  # Bank not available
            user_message = "Bank server is currently unavailable. Please try again in some time."
        
        logging.warning(f"[BBPS PAY] Failed: status={eko_status}, message={message}")
        
        return {
            "success": False,
            "status": "FAILED",
            "error_code": eko_status,
            "tid": tid,
            "client_ref_id": client_ref_id,
            "message": message,
            "user_message": user_message,
            "raw_response": result
        }
        
    except httpx.TimeoutException:
        logging.error(f"[BBPS PAY] Timeout after {REQUEST_TIMEOUT}s")
        return {
            "success": False,
            "status": "TIMEOUT",
            "error_code": 504,
            "message": "Request timeout",
            "user_message": "Payment is taking longer than expected. Please check your transaction history before retrying.",
            "requires_status_check": True,
            "client_ref_id": client_ref_id
        }
    
    except httpx.ConnectError as e:
        logging.error(f"[BBPS PAY] Connection error: {e}")
        return create_error_response(
            503,
            "Connection failed",
            "Unable to connect to payment service. Your account has not been charged. Please try again."
        )
    
    except Exception as e:
        logging.error(f"[BBPS PAY] Unexpected error: {e}")
        return create_error_response(
            500,
            str(e),
            "An unexpected error occurred. Please check your transaction history before retrying."
        )


# ==================== TRANSACTION STATUS INQUIRY ====================

def _parse_eko_status_response(result: dict, identifier: str, identifier_type: str = "tid") -> dict:
    """Parse EKO transaction status response into standardized format"""
    eko_status = result.get("status")
    eko_data = result.get("data", {})
    
    if eko_status == 0:
        tx_status = eko_data.get("tx_status")
        if isinstance(tx_status, str):
            try:
                tx_status = int(tx_status)
            except ValueError:
                pass
        
        status_map = {
            0: "SUCCESS",
            1: "FAILED",
            2: "PENDING",
            3: "REFUND_PENDING",
            4: "REFUNDED",
            5: "ON_HOLD"
        }
        
        return {
            "success": True,
            "tid": eko_data.get("tid", identifier if identifier_type == "tid" else None),
            "client_ref_id": eko_data.get("client_ref_id", identifier if identifier_type == "client_ref_id" else None),
            "tx_status": tx_status,
            "status": status_map.get(tx_status, "UNKNOWN"),
            "message": eko_data.get("txstatus_desc", TX_STATUS_MESSAGES.get(tx_status, "Unknown status")),
            "amount": eko_data.get("amount"),
            "bbps_ref": eko_data.get("bbpstrxnrefid"),
            "utr": eko_data.get("bank_ref_num") or eko_data.get("utr"),
            "refund_status": "REFUNDED" if tx_status == 4 else ("REFUND_PENDING" if tx_status == 3 else None),
            "raw_response": result
        }
    
    return {
        "success": False,
        "error_code": eko_status,
        "message": result.get("message", "Status check failed"),
        "user_message": "Transaction not found. Please verify the transaction ID.",
        "raw_response": result
    }


@router.get("/status/{tid}")
async def get_transaction_status(tid: str):
    """
    Check transaction status via EKO API using TID.
    
    Use this when:
    - tx_status = 2 (Pending/Initiated)
    - tx_status = 5 (On Hold)
    - Payment timeout occurred
    
    Reference: https://developers.eko.in/reference/transaction-inquiry
    """
    try:
        url = f"{BASE_URL}/v2/transactions/{tid}?initiator_id={INITIATOR_ID}"
        
        logging.info(f"[BBPS STATUS] Checking TID: {tid}")
        
        response = await bbps_get(url, headers=generate_headers(), timeout=30)
        
        if response.status_code != 200:
            return create_error_response(
                response.status_code,
                "Status check failed",
                "Unable to check transaction status. Please try again."
            )
        
        result = response.json()
        return _parse_eko_status_response(result, tid, "tid")
        
    except Exception as e:
        logging.error(f"[BBPS STATUS] Error: {e}")
        return create_error_response(500, str(e), "Unable to check status. Please try again.")


@router.get("/status-by-ref/{client_ref_id}")
async def get_transaction_status_by_ref(client_ref_id: str):
    """
    Check transaction status via EKO API using client_ref_id.
    
    Use when TID is N/A (transaction failed before EKO assigned a TID).
    This checks EKO's records using our client_ref_id to see if:
    - EKO debited the wallet
    - EKO refunded the amount
    - Transaction is still pending
    
    Reference: https://developers.eko.in/reference/transaction-inquiry
    """
    try:
        url = f"{BASE_URL}/v2/transactions/client_ref_id:{client_ref_id}?initiator_id={INITIATOR_ID}"
        
        logging.info(f"[BBPS STATUS BY REF] Checking client_ref_id: {client_ref_id}, URL: {url}")
        
        response = await bbps_get(url, headers=generate_headers(), timeout=30)
        
        logging.info(f"[BBPS STATUS BY REF] HTTP: {response.status_code}, Response: {response.text[:500]}")
        
        if response.status_code != 200:
            return create_error_response(
                response.status_code,
                "Status check failed",
                "Unable to check transaction status. Please try again."
            )
        
        result = response.json()
        parsed = _parse_eko_status_response(result, client_ref_id, "client_ref_id")
        parsed["lookup_type"] = "client_ref_id"
        return parsed
        
    except Exception as e:
        logging.error(f"[BBPS STATUS BY REF] Error: {e}")
        return create_error_response(500, str(e), "Unable to check status. Please try again.")


@router.get("/admin/check-eko-refund/{request_id}")
async def admin_check_eko_refund(request_id: str):
    """
    Admin endpoint: Check EKO refund status for a failed bill payment.
    
    Looks up the bill_payment_requests record, extracts client_ref_id or tid,
    and queries EKO for the current transaction + refund status.
    
    This helps admin verify if EKO wallet was debited and if refund was processed.
    """
    if db is None:
        return {"success": False, "error": "Database not initialized"}
    
    try:
        # Find the request in DB
        request_doc = await db.bill_payment_requests.find_one(
            {"request_id": request_id},
            {"_id": 0}
        )
        
        if not request_doc:
            # Try recharge_requests collection too
            request_doc = await db.recharge_requests.find_one(
                {"request_id": request_id},
                {"_id": 0}
            )
        
        if not request_doc:
            return {"success": False, "error": f"Request {request_id} not found"}
        
        eko_tid = request_doc.get("eko_tid") or request_doc.get("tid")
        client_ref = request_doc.get("client_ref_id") or request_doc.get("eko_client_ref_id")
        
        eko_result = None
        lookup_method = None
        
        # Try TID first (more reliable)
        if eko_tid and eko_tid not in ["N/A", "null", "", None]:
            url = f"{BASE_URL}/v2/transactions/{eko_tid}?initiator_id={INITIATOR_ID}"
            logging.info(f"[EKO REFUND CHECK] Checking TID: {eko_tid} for request: {request_id}")
            response = await bbps_get(url, headers=generate_headers(), timeout=30)
            if response.status_code == 200:
                eko_result = response.json()
                lookup_method = "tid"
        
        # Fallback to client_ref_id
        if not eko_result and client_ref:
            url = f"{BASE_URL}/v2/transactions/client_ref_id:{client_ref}?initiator_id={INITIATOR_ID}"
            logging.info(f"[EKO REFUND CHECK] Checking client_ref_id: {client_ref} for request: {request_id}")
            response = await bbps_get(url, headers=generate_headers(), timeout=30)
            if response.status_code == 200:
                eko_result = response.json()
                lookup_method = "client_ref_id"
        
        if not eko_result:
            return {
                "success": False,
                "request_id": request_id,
                "our_status": request_doc.get("status"),
                "eko_tid": eko_tid,
                "client_ref_id": client_ref,
                "error": "Could not find transaction in EKO system. No TID or client_ref_id available.",
                "suggestion": "Check EKO dashboard directly or contact EKO support."
            }
        
        parsed = _parse_eko_status_response(eko_result, eko_tid or client_ref, lookup_method)
        
        return {
            "success": True,
            "request_id": request_id,
            "our_status": request_doc.get("status"),
            "our_refund_status": request_doc.get("refund_status"),
            "prc_refunded": request_doc.get("prc_refunded") or request_doc.get("refund_amount"),
            "amount_inr": request_doc.get("amount") or request_doc.get("amount_inr"),
            "lookup_method": lookup_method,
            "eko_status": parsed.get("status"),
            "eko_tx_status": parsed.get("tx_status"),
            "eko_refund_status": parsed.get("refund_status"),
            "eko_tid": parsed.get("tid"),
            "eko_message": parsed.get("message"),
            "wallet_debited": parsed.get("tx_status") is not None,
            "eko_refunded": parsed.get("tx_status") in [3, 4],
            "raw_eko_response": parsed.get("raw_response")
        }
    
    except Exception as e:
        logging.error(f"[EKO REFUND CHECK] Error for {request_id}: {e}")
        return {"success": False, "request_id": request_id, "error": str(e)}


# ==================== GET CATEGORIES ====================

@router.get("/categories")
async def get_categories():
    """
    Get all BBPS categories from Eko API.
    
    Returns list of categories like:
    - Electricity, Water, Gas, DTH, Loan/EMI, Insurance, etc.
    """
    if not validate_bbps_config():
        return create_error_response(500, "Service configuration error", "Service temporarily unavailable.")
    
    try:
        url = f"{BASE_URL}/v3/customer/payment/bbps/categories?initiator_id={INITIATOR_ID}&user_code={USER_CODE}&client_ref_id=CAT{int(time.time())}"
        
        response = await bbps_get(url, headers=generate_headers(), timeout=30)
        
        logging.info(f"[BBPS CATEGORIES] HTTP Status: {response.status_code}")
        logging.info(f"[BBPS CATEGORIES] Response: {response.text[:500] if response.text else 'empty'}")
        
        if response.status_code != 200:
            return create_error_response(
                response.status_code,
                f"Failed to fetch categories: {response.text}",
                "Unable to load categories."
            )
        
        result = response.json()
        return {
            "success": True,
            "categories": result.get("data", result) if isinstance(result, dict) else result
        }
    
    except Exception as e:
        logging.error(f"[BBPS CATEGORIES] Error: {e}")
        return create_error_response(500, str(e), "Service temporarily unavailable.")


@router.get("/operators/search/{query}")
async def search_operators(query: str, category: Optional[str] = None):
    """
    Search operators by name across all categories or within a specific category.
    
    Example: /operators/search/IDFC?category=emi
    """
    if not validate_bbps_config():
        return create_error_response(500, "Service configuration error", "Service temporarily unavailable.")
    
    category_map = {
        "mobile_prepaid": 5, "mobile_postpaid": 10, "dth": 4,
        "electricity": 8, "water": 11, "landline": 9, "broadband": 1,
        "gas": 2, "lpg": 18, "emi": 21, "loan": 21, "credit_card": 7,
        "insurance": 20, "fastag": 22, "housing_society": 12,
        "municipal_tax": 15, "education": 14, "cable_tv": 17,
        "subscription": 13, "hospital": 19, "municipal_corp": 6,
        "loan_repayment": 25, "transport": 27
    }
    
    # If category specified, search only that category
    categories_to_search = [category_map.get(category.lower())] if category and category.lower() in category_map else list(set(category_map.values()))
    
    all_matches = []
    
    for cat_id in categories_to_search[:5]:  # Limit to 5 categories to avoid timeout
        try:
            url = f"{BASE_URL}/v2/billpayments/operators?initiator_id={INITIATOR_ID}&category={cat_id}"
            response = await bbps_get(url, headers=generate_headers(), timeout=15)
            
            if response.status_code == 200:
                result = response.json()
                operators = result.get("data", result) if isinstance(result, dict) else result
                
                if isinstance(operators, list):
                    for op in operators:
                        op_name = op.get("name", "") or ""
                        if query.lower() in op_name.lower():
                            all_matches.append({
                                "operator_id": op.get("operator_id"),
                                "name": op_name,
                                "category_id": cat_id,
                                "billFetchResponse": op.get("billFetchResponse", 0)
                            })
        except Exception as e:
            logging.warning(f"[BBPS SEARCH] Category {cat_id} failed: {e}")
            continue
    
    return {
        "success": True,
        "query": query,
        "matches": all_matches,
        "count": len(all_matches)
    }


# ==================== GET OPERATORS ====================

@router.get("/operators/{category}")
async def get_operators(category: str):
    """
    Get operators list for a category.
    
    EKO BBPS Categories (verified with production API):
    - 4: DTH (5 operators: Dish TV, Tata Sky, Airtel DTH, etc.)
    - 5: Mobile Prepaid (6 operators: Jio, Airtel, Vi, BSNL, MTNL)
    - 7: Credit Card (29 operators)
    - 8: Electricity (89 operators)
    - 9: Landline (5 operators)
    - 10: Mobile Postpaid (7 operators)
    - 11: Water (54 operators)
    - 12: Housing Society (105 operators)
    - 20: Insurance (40 operators)
    - 21: Loan/EMI (294 operators)
    - 22: FASTag (20 operators: IndusInd, Axis, BOB, etc.)
    - 1: Broadband (92 operators)
    """
    # Complete Eko BBPS Category Mapping (as of March 2026)
    # Based on direct API investigation
    category_map = {
        # Mobile Services
        "mobile_recharge": 5,   # Cat 5: Airtel, Jio, Vi, BSNL Prepaid (6 operators)
        "mobile_prepaid": 5,
        "mobile_postpaid": 10,  # Cat 10: Airtel, Jio, Vi Postpaid (7 operators)
        
        # Entertainment
        "dth": 4,               # Cat 4: Dish TV, Tata Sky, Airtel DTH (5 operators)
        "cable_tv": 17,         # Cat 17: Hathway, Asianet, INDigital (4 operators)
        "subscription": 13,     # Cat 13: Amazon Prime, JioHotstar, Hungama (17 operators)
        "ott": 13,
        
        # Utility Bills
        "electricity": 8,       # Cat 8: MSEDCL, BSES, Tata Power (89 operators)
        "water": 11,            # Cat 11: Municipal water boards (54 operators)
        "landline": 9,          # Cat 9: Airtel, BSNL, MTNL Landline (5 operators)
        "broadband": 1,         # Cat 1: Airtel, Jio, ACT Fibernet (93 operators)
        "gas": 2,               # Cat 2: Mahanagar Gas, Gujarat Gas, Adani (29 operators)
        "lpg": 18,              # Cat 18: Indane, HP Gas, Bharat Gas (3 operators)
        
        # Financial Services
        "credit_card": 7,       # Cat 7: HDFC, ICICI, Axis, etc. (29 operators)
        "emi": 21,              # Cat 21: Loan EMI payments (294 operators)
        "loan": 21,
        "loan_emi": 21,
        "loan_repayment": 25,   # Cat 25: Agent/Customer loan repayment (283 operators)
        "insurance": 20,        # Cat 20: Life & General Insurance (40 operators)
        
        # Transport
        "fastag": 22,           # Cat 22: Bank FASTag recharge (20 operators)
        "transport": 27,        # Cat 27: Transport Department services (5 operators)
        
        # Property & Housing
        "housing_society": 12,  # Cat 12: Society maintenance fees (105 operators)
        "municipal_tax": 15,    # Cat 15: Property Tax - MCGM, etc. (41 operators)
        "municipal_corp": 6,    # Cat 6: Metro Municipal Corps - KDMC, AMC (2 operators)
        "municipal_other": 23,  # Cat 23: Other municipal services (5 operators)
        
        # Education & Healthcare
        "education": 14,        # Cat 14: School/College Fees (1661 operators)
        "school_fees": 14,
        "college_fees": 14,
        "hospital": 19,         # Cat 19: Hospital payments (6 operators)
        "healthcare": 19
    }
    
    cat_id = category_map.get(category.lower())
    
    if not cat_id:
        return create_error_response(
            400,
            f"Unknown category: {category}",
            f"Invalid service category. Available: {', '.join(category_map.keys())}"
        )
    
    try:
        url = f"{BASE_URL}/v2/billpayments/operators?initiator_id={INITIATOR_ID}&category={cat_id}"
        
        response = await bbps_get(url, headers=generate_headers(), timeout=30)
        
        if response.status_code != 200:
            return create_error_response(
                response.status_code,
                f"Failed to fetch operators for {category}",
                "Unable to load service providers. Please try again."
            )
        
        result = response.json()
        
        # Eko operators API returns data directly as array OR with status wrapper
        # Handle both formats
        if isinstance(result, list):
            # Direct array response
            operators = result
        elif result.get("status") is not None and result.get("status") != 0:
            # Error response with status
            return create_error_response(
                result.get("status"),
                result.get("message", "Failed to fetch operators"),
                "Unable to load service providers. Please try again."
            )
        else:
            # Response with data wrapper
            operators = result.get("data", [])
        
        # Format operators for frontend
        formatted = []
        for op in operators:
            formatted.append({
                "operator_id": op.get("operator_id"),
                "name": op.get("name"),
                "category": op.get("category"),
                "billFetchResponse": op.get("billFetchResponse", 0),
                "supports_bill_fetch": op.get("billFetchResponse", 0) == 1
            })
        
        # Sort operators A to Z by name
        formatted.sort(key=lambda x: (x.get("name") or "").lower())
        
        return {
            "success": True,
            "category": category,
            "eko_category_id": cat_id,
            "count": len(formatted),
            "operators": formatted
        }
        
    except httpx.TimeoutException:
        return create_error_response(504, "Request timeout", "Service is slow. Please try again.")
    except Exception as e:
        logging.error(f"[BBPS OPERATORS] Error: {e}")
        return create_error_response(500, str(e), "Failed to load providers. Please refresh.")


# ==================== SERVICE STATUS ====================


# ==================== ERROR CODES REFERENCE ====================

@router.get("/error-codes")
def get_error_codes():
    """
    Get all Eko error codes and their meanings.
    Useful for debugging and support.
    """
    return {
        "http_codes": {
            "200": "OK - Check status and tx_status in response",
            "403": "Forbidden - Invalid authentication",
            "404": "Not Found - Invalid URL",
            "405": "Method Not Allowed",
            "415": "Unsupported Media Type",
            "500": "Server Error"
        },
        "status_codes": EKO_ERROR_MESSAGES,
        "tx_status": TX_STATUS_MESSAGES
    }


# ==================== SERVICE ACTIVATION ====================

@router.post("/activate-service/{service_code}")
async def activate_service(service_code: int = 53):
    """
    Activate a service for the merchant/agent
    
    Service Codes:
    - 53: BBPS (Bill Payments - Electricity, Gas, DTH, Mobile, etc.)
    - 45: DMT (Domestic Money Transfer)
    
    IMPORTANT: This must be called once to activate services before using them!
    """
    import asyncio
    
    try:
        # Generate fresh headers with correct Content-Type for form data
        timestamp = str(round(time.time() * 1000))
        encoded_key = base64.b64encode(AUTH_KEY.encode()).decode()
        secret_key = base64.b64encode(
            hmac.new(encoded_key.encode(), timestamp.encode(), hashlib.sha256).digest()
        ).decode()
        
        headers = {
            "developer_key": DEVELOPER_KEY,
            "secret-key": secret_key,
            "secret-key-timestamp": timestamp,
            "Content-Type": "application/x-www-form-urlencoded"
        }
        
        # URL for service activation
        url = f"{BASE_URL}/v1/user/service/activate"
        
        # Body - form data
        data = {
            "service_code": str(service_code),
            "initiator_id": INITIATOR_ID,
            "user_code": USER_CODE,
            "latlong": "19.0760,72.8777"
        }
        
        logging.info(f"[BBPS] Activating service {service_code} for user {USER_CODE}")
        logging.info(f"[BBPS] URL: {url}")
        
        # FIXED: Use httpx async client instead of sync requests
        async with httpx.AsyncClient() as client:
            response = await client.put(url, headers=headers, data=data, timeout=60)
        
        logging.info(f"[BBPS] Activation response: {response.status_code} - {response.text[:200]}")
        
        try:
            result = response.json()
        except:
            result = {"raw": response.text, "parse_error": True}
        
        eko_status = result.get("status")
        
        if eko_status == 0:
            return {
                "success": True,
                "message": f"Service {service_code} activated successfully!",
                "service_code": service_code,
                "user_code": USER_CODE,
                "raw_response": result
            }
        elif eko_status == 24:
            # Already activated
            return {
                "success": True,
                "message": f"Service {service_code} is already activated",
                "service_code": service_code,
                "user_code": USER_CODE,
                "already_active": True,
                "raw_response": result
            }
        else:
            return {
                "success": False,
                "message": result.get("message", "Activation failed"),
                "error_code": str(eko_status),
                "service_code": service_code,
                "raw_response": result
            }
            
    except Exception as e:
        logging.error(f"[BBPS] Service activation error: {e}")
        return {
            "success": False,
            "message": f"Activation failed: {str(e)}",
            "service_code": service_code
        }


@router.get("/service-status/{service_code}")
async def check_service_status(service_code: int = 53):
    """
    Check if a service is activated for the user
    """
    try:
        # Generate headers
        timestamp = str(round(time.time() * 1000))
        encoded_key = base64.b64encode(AUTH_KEY.encode()).decode()
        secret_key = base64.b64encode(
            hmac.new(encoded_key.encode(), timestamp.encode(), hashlib.sha256).digest()
        ).decode()
        
        headers = {
            "developer_key": DEVELOPER_KEY,
            "secret-key": secret_key,
            "secret-key-timestamp": timestamp
        }
        
        # Check specific service status
        url = f"{BASE_URL}/v1/user/services?initiator_id={INITIATOR_ID}&user_code={USER_CODE}"
        
        logging.info(f"[BBPS] Checking service status: {url}")
        
        # FIXED: Use httpx async client instead of sync requests
        async with httpx.AsyncClient() as client:
            response = await client.get(url, headers=headers, timeout=30)
        result = response.json()
        
        logging.info(f"[BBPS] Service status response: {result}")
        
        return {
            "success": True,
            "service_code": service_code,
            "user_code": USER_CODE,
            "initiator_id": INITIATOR_ID,
            "services": result.get("data", {}).get("service_list", result.get("data", {}).get("services", [])),
            "raw_response": result
        }
        
    except Exception as e:
        logging.error(f"[BBPS] Service status check error: {e}")
        return {
            "success": False,
            "message": str(e),
            "service_code": service_code
        }


@router.post("/activate-all")
async def activate_all_services():
    """
    Activate all required services (BBPS + DMT)
    """
    results = {
        "bbps_53": await activate_service(53),
        "dmt_45": await activate_service(45)
    }
    
    all_success = all(r.get("success") or r.get("already_active") for r in results.values())
    
    return {
        "success": all_success,
        "message": "All services activated" if all_success else "Some services failed",
        "results": results
    }



# ==================== TRANSACTION STATUS CALLBACK (WEBHOOK) ====================
# EKO pushes status updates for pending/processing transactions to this endpoint.
# Configure this URL in EKO dashboard: https://<domain>/api/bbps/callback/status
# Ref: https://developers.eko.in/docs/callback-setup

@router.post("/callback/status")
async def eko_transaction_callback(request: Request):
    """
    Webhook endpoint: EKO pushes transaction status updates here.
    
    EKO calls this when tx_status changes for pending/processing transactions.
    Expected payload from EKO:
      {
        "tid": "12345",
        "client_ref_id": "MOB1712345678",
        "tx_status": 0,     # 0=Success,1=Failed,2=Pending,3=RefundPending,4=Refunded
        "amount": "500",
        "bbpstrxnrefid": "BBP123456",
        "operator_ref": "...",
        "timestamp": "1712345678"
      }
    """
    if db is None:
        return {"success": False, "error": "Database not initialized"}
    
    try:
        body = await request.json()
    except Exception:
        body = dict(await request.form())
    
    logging.info(f"[EKO CALLBACK] Received: {json.dumps(body, default=str)[:1000]}")
    
    eko_tid = body.get("tid") or body.get("txn_id") or body.get("eko_tid")
    client_ref = body.get("client_ref_id")
    tx_status_raw = body.get("tx_status")
    amount = body.get("amount")
    bbps_ref = body.get("bbpstrxnrefid") or body.get("bbps_ref")
    utr = body.get("bank_ref_num") or body.get("utr")
    
    try:
        tx_status = int(tx_status_raw) if tx_status_raw is not None else None
    except (ValueError, TypeError):
        tx_status = None
    
    STATUS_MAP = {
        0: "completed",
        1: "failed",
        2: "processing",
        3: "refund_pending",
        4: "refunded",
        5: "on_hold"
    }
    new_status = STATUS_MAP.get(tx_status, "unknown")
    
    # Save callback log for audit (always)
    callback_log = {
        "callback_id": f"CB-{int(time.time()*1000)}",
        "eko_tid": eko_tid,
        "client_ref_id": client_ref,
        "tx_status": tx_status,
        "mapped_status": new_status,
        "amount": amount,
        "bbps_ref": bbps_ref,
        "utr": utr,
        "raw_payload": body,
        "received_at": datetime.now(timezone.utc).isoformat(),
        "processed": False,
        "process_result": None
    }
    await db.eko_callbacks.insert_one(callback_log)
    
    # Find matching request in DB
    request_doc = None
    collection_name = None
    
    # Strategy: try eko_tid first (most reliable), then client_ref_id
    search_queries = []
    if eko_tid:
        search_queries.append({"eko_tid": str(eko_tid)})
    if client_ref:
        search_queries.append({"client_ref_id": str(client_ref)})
        search_queries.append({"eko_client_ref_id": str(client_ref)})
    
    for query in search_queries:
        # Check redeem_requests first (unified flow)
        request_doc = await db.redeem_requests.find_one(query, {"_id": 0})
        if request_doc:
            collection_name = "redeem_requests"
            break
        # Check bill_payment_requests (legacy flow)
        request_doc = await db.bill_payment_requests.find_one(query, {"_id": 0})
        if request_doc:
            collection_name = "bill_payment_requests"
            break
    
    if not request_doc:
        logging.warning(f"[EKO CALLBACK] No matching request found for TID={eko_tid}, client_ref={client_ref}")
        await db.eko_callbacks.update_one(
            {"callback_id": callback_log["callback_id"]},
            {"$set": {"processed": True, "process_result": "no_match_found"}}
        )
        return {"success": True, "message": "Callback logged (no matching request found)"}
    
    request_id = request_doc.get("request_id")
    current_status = request_doc.get("status", "")
    user_id = request_doc.get("user_id")
    
    logging.info(f"[EKO CALLBACK] Matched request: {request_id} in {collection_name}, current_status={current_status}, new_status={new_status}")
    
    # Determine what update to apply
    update_fields = {
        "eko_callback_status": new_status,
        "eko_callback_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if eko_tid:
        update_fields["eko_tid"] = str(eko_tid)
    if bbps_ref:
        update_fields["utr_number"] = bbps_ref
    if utr:
        update_fields["utr_number"] = utr
    
    status_update_note = f"EKO Callback: tx_status={tx_status} ({new_status})"
    action_taken = "logged"
    
    # === Handle status transitions ===
    
    # CASE 1: Transaction SUCCESS (was pending/processing → now success)
    if tx_status == 0 and current_status in ["pending", "processing", "submitted"]:
        update_fields["status"] = "completed"
        update_fields["eko_status"] = "SUCCESS"
        update_fields["completed_at"] = datetime.now(timezone.utc).isoformat()
        action_taken = "marked_completed"
        logging.info(f"[EKO CALLBACK] {request_id}: {current_status} → completed (EKO confirmed success)")
    
    # CASE 2: Transaction FAILED (was pending/processing → now failed) → Refund PRC
    elif tx_status == 1 and current_status in ["pending", "processing", "submitted"]:
        update_fields["status"] = "failed"
        update_fields["eko_status"] = "FAILED"
        update_fields["eko_message"] = body.get("message", "Transaction failed (EKO callback)")
        update_fields["failed_at"] = datetime.now(timezone.utc).isoformat()
        action_taken = "marked_failed"
        
        # Auto-refund PRC
        if user_id and not request_doc.get("prc_refunded"):
            try:
                refund_amount = request_doc.get("total_prc_deducted") or request_doc.get("prc_used") or 0
                if refund_amount > 0:
                    from app.services.wallet_service_v2 import WalletServiceV2 as WalletService
                    WalletService.credit(
                        user_id=user_id,
                        amount=refund_amount,
                        txn_type="refund",
                        description=f"Auto-refund: EKO callback confirmed failure for {request_id}",
                        reference=request_id
                    )
                    update_fields["prc_refunded"] = True
                    update_fields["refund_amount"] = refund_amount
                    action_taken = f"marked_failed_and_refunded_{refund_amount}_PRC"
                    logging.info(f"[EKO CALLBACK] Auto-refunded {refund_amount} PRC to {user_id}")
            except Exception as refund_err:
                logging.error(f"[EKO CALLBACK] Refund failed for {request_id}: {refund_err}")
                update_fields["refund_error"] = str(refund_err)
                action_taken = "marked_failed_refund_error"
    
    # CASE 3: EKO Refunded (failed→refunded on EKO side, update our records)
    elif tx_status in [3, 4]:
        update_fields["eko_refund_status"] = "REFUNDED" if tx_status == 4 else "REFUND_PENDING"
        action_taken = f"eko_refund_status_updated_{STATUS_MAP.get(tx_status)}"
        logging.info(f"[EKO CALLBACK] {request_id}: EKO refund status = {STATUS_MAP.get(tx_status)}")
    
    # CASE 4: Still pending/on hold
    elif tx_status in [2, 5]:
        update_fields["eko_status"] = "PENDING" if tx_status == 2 else "ON_HOLD"
        action_taken = "still_pending"
    
    # Apply update to DB
    update_query = {"request_id": request_id}
    await db[collection_name].update_one(
        update_query,
        {
            "$set": update_fields,
            "$push": {
                "status_history": {
                    "status": new_status,
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                    "note": status_update_note,
                    "source": "eko_callback"
                }
            }
        }
    )
    
    # Update callback log with result
    await db.eko_callbacks.update_one(
        {"callback_id": callback_log["callback_id"]},
        {"$set": {
            "processed": True,
            "process_result": action_taken,
            "matched_request_id": request_id,
            "matched_collection": collection_name
        }}
    )
    
    logging.info(f"[EKO CALLBACK] Done: request={request_id}, action={action_taken}")
    
    return {
        "success": True,
        "message": f"Callback processed: {action_taken}",
        "request_id": request_id
    }


@router.get("/callback/logs")
async def get_callback_logs(limit: int = 50, skip: int = 0):
    """
    Admin endpoint: View recent EKO callback logs for debugging/audit.
    Shows all callbacks received from EKO with processing results.
    """
    if db is None:
        return {"success": False, "error": "Database not initialized"}
    
    try:
        logs = await db.eko_callbacks.find(
            {}, {"_id": 0}
        ).sort("received_at", -1).skip(skip).limit(limit).to_list(limit)
        
        total = await db.eko_callbacks.count_documents({})
        
        return {
            "success": True,
            "logs": logs,
            "total": total,
            "showing": len(logs)
        }
    except Exception as e:
        logging.error(f"[CALLBACK LOGS] Error: {e}")
        return {"success": False, "error": str(e)}



# ==================== FAILED TRANSACTIONS EXPORT (EXCEL) ====================

@router.get("/admin/export-failed")
async def export_failed_transactions(date: str = None):
    """
    Admin endpoint: Export failed transactions as Excel file.
    
    Args:
        date: Date in YYYY-MM-DD format (defaults to today)
    
    Returns: Excel file download with columns:
        Request ID, Client Ref ID, EKO TID, User Name, Mobile,
        Service Type, Amount (INR), Consumer/Account, Operator,
        Status, EKO Message, PRC Refunded, Date/Time
    """
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    
    if not date:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Query failed transactions from both collections
    failed_statuses = ["failed", "FAILED", "retry_failed"]
    
    # redeem_requests (unified flow - primary)
    redeem_docs = await db.redeem_requests.find(
        {
            "status": {"$in": failed_statuses},
            "created_at": {"$regex": f"^{date}"}
        },
        {"_id": 0, "status_history": 0, "eko_response": 0}
    ).sort("created_at", -1).to_list(1000)
    
    # bill_payment_requests (legacy flow)
    bill_docs = await db.bill_payment_requests.find(
        {
            "status": {"$in": failed_statuses},
            "created_at": {"$regex": f"^{date}"}
        },
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    
    all_docs = []
    for doc in redeem_docs:
        details = doc.get("details", {})
        all_docs.append({
            "request_id": doc.get("request_id", ""),
            "client_ref_id": doc.get("client_ref_id", ""),
            "eko_tid": doc.get("eko_tid") or "N/A",
            "user_name": doc.get("user_name", ""),
            "user_mobile": doc.get("user_mobile", ""),
            "service_type": doc.get("service_type", ""),
            "amount_inr": doc.get("amount_inr") or doc.get("amount", 0),
            "consumer_number": details.get("consumer_number") or details.get("mobile_number") or details.get("loan_account") or "",
            "operator": details.get("operator") or details.get("operator_id") or "",
            "cycle_number": details.get("cycle_number") or "",
            "status": doc.get("status", ""),
            "eko_message": doc.get("eko_message") or doc.get("error_message") or "",
            "prc_deducted": doc.get("total_prc_deducted", 0),
            "prc_refunded": "Yes" if doc.get("prc_refunded") else "No",
            "refund_amount": doc.get("refund_amount", 0),
            "created_at": doc.get("created_at", ""),
            "source": "redeem_requests"
        })
    
    for doc in bill_docs:
        details = doc.get("details", {})
        all_docs.append({
            "request_id": doc.get("request_id", ""),
            "client_ref_id": doc.get("client_ref_id", ""),
            "eko_tid": doc.get("eko_tid") or "N/A",
            "user_name": doc.get("user_name", ""),
            "user_mobile": doc.get("user_mobile", ""),
            "service_type": doc.get("request_type", ""),
            "amount_inr": doc.get("amount_inr", 0),
            "consumer_number": details.get("consumer_number") or details.get("phone_number") or "",
            "operator": details.get("operator") or "",
            "cycle_number": details.get("cycle_number") or "",
            "status": doc.get("status", ""),
            "eko_message": doc.get("eko_message") or doc.get("admin_notes") or "",
            "prc_deducted": doc.get("total_prc_deducted", 0),
            "prc_refunded": "Yes" if doc.get("prc_refunded") else "No",
            "refund_amount": doc.get("refund_amount", 0),
            "created_at": doc.get("created_at", ""),
            "source": "bill_payment_requests"
        })
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Failed Transactions {date}"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    headers = [
        "Sr No", "Request ID", "Client Ref ID", "EKO TID",
        "User Name", "Mobile", "Service Type",
        "Amount (INR)", "Consumer/Account No", "Operator", "Cycle No",
        "Status", "EKO Error Message",
        "PRC Deducted", "PRC Refunded", "Refund Amount",
        "Date & Time"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    data_alignment = Alignment(vertical="center", wrap_text=True)
    for idx, doc in enumerate(all_docs, 1):
        row = idx + 1
        values = [
            idx,
            doc["request_id"],
            doc["client_ref_id"],
            doc["eko_tid"],
            doc["user_name"],
            doc["user_mobile"],
            doc["service_type"],
            doc["amount_inr"],
            doc["consumer_number"],
            doc["operator"],
            doc["cycle_number"],
            doc["status"],
            doc["eko_message"],
            doc["prc_deducted"],
            doc["prc_refunded"],
            doc["refund_amount"],
            doc["created_at"]
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)
    
    # Summary row
    summary_row = len(all_docs) + 3
    ws.cell(row=summary_row, column=1, value="Summary:").font = Font(bold=True, size=12)
    ws.cell(row=summary_row + 1, column=1, value=f"Total Failed: {len(all_docs)}")
    total_inr = sum(float(d.get("amount_inr", 0) or 0) for d in all_docs)
    ws.cell(row=summary_row + 2, column=1, value=f"Total Amount (INR): ₹{total_inr:,.2f}")
    total_prc = sum(float(d.get("prc_deducted", 0) or 0) for d in all_docs)
    ws.cell(row=summary_row + 3, column=1, value=f"Total PRC Deducted: {total_prc:,.2f}")
    refunded_count = sum(1 for d in all_docs if d["prc_refunded"] == "Yes")
    ws.cell(row=summary_row + 4, column=1, value=f"PRC Refunded Count: {refunded_count}/{len(all_docs)}")
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"failed_transactions_{date}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== EKO REFUND API (Wallet Refund) ====================
# Ref: https://developers.eko.in/v1/reference/initiate-refund
# Flow: 1) Resend OTP → 2) Customer shares OTP → 3) Verify OTP & Refund

@router.post("/refund/resend-otp/{tid}")
async def resend_refund_otp(tid: str):
    """
    Step 1: Resend refund OTP to customer.
    
    When a transaction fails, EKO auto-sends OTP to customer.
    Use this endpoint to resend if customer didn't receive it.
    
    Args:
        tid: EKO Transaction ID (numeric)
    
    Ref: https://developers.eko.in/v1/reference/resend-refund-otp-1
    """
    try:
        timestamp = str(round(time.time() * 1000))
        encoded_key = base64.b64encode(AUTH_KEY.encode())
        secret_key = base64.b64encode(
            hmac.new(encoded_key, timestamp.encode(), hashlib.sha256).digest()
        ).decode()
        
        headers = {
            "developer_key": DEVELOPER_KEY,
            "secret-key": secret_key,
            "secret-key-timestamp": timestamp,
            "Content-Type": "application/x-www-form-urlencoded",
        }
        
        # Production uses /ekoicici/ base path (same as other APIs)
        url = f"{BASE_URL}/v1/transactions/{tid}/refund/otp"
        body = {"initiator_id": INITIATOR_ID}
        
        logging.info(f"[EKO REFUND] Resending OTP for TID: {tid}, URL: {url}")
        
        import requests as sync_requests
        response = sync_requests.post(url, headers=headers, data=body, timeout=30)
        
        logging.info(f"[EKO REFUND] OTP HTTP: {response.status_code} | Body: {response.text[:500]}")
        
        if response.status_code == 404:
            return {"success": False, "tid": tid, "message": "EKO endpoint not found. Check TID format.", "http_status": 404}
        
        try:
            result = response.json()
        except Exception:
            return {"success": False, "tid": tid, "message": f"EKO returned non-JSON: HTTP {response.status_code}", "raw": response.text[:300]}
        
        logging.info(f"[EKO REFUND] OTP Response: {json.dumps(result)[:300]}")
        
        return {
            "success": result.get("status") == 0,
            "tid": tid,
            "message": result.get("message", "OTP sent to customer"),
            "otp_ref": result.get("data", {}).get("otp"),
            "raw_response": result
        }
        
    except Exception as e:
        logging.error(f"[EKO REFUND] OTP Error for TID {tid}: {e}")
        return {"success": False, "error": str(e)}


@router.post("/refund/verify/{tid}")
async def verify_refund_otp(tid: str, otp: str, state: int = 1):
    """
    Step 2: Verify OTP and complete refund.
    
    Customer provides OTP → call this to confirm refund.
    EKO will refund eValue back to merchant wallet.
    
    Args:
        tid: EKO Transaction ID
        otp: OTP received by customer
        state: Default 1 (always pass 1)
    
    Ref: https://developers.eko.in/v1/reference/initiate-refund
    """
    try:
        timestamp = str(round(time.time() * 1000))
        encoded_key = base64.b64encode(AUTH_KEY.encode())
        secret_key = base64.b64encode(
            hmac.new(encoded_key, timestamp.encode(), hashlib.sha256).digest()
        ).decode()
        
        headers = {
            "developer_key": DEVELOPER_KEY,
            "secret-key": secret_key,
            "secret-key-timestamp": timestamp,
            "Content-Type": "application/x-www-form-urlencoded",
        }
        
        # Production uses /ekoicici/ base path (same as other APIs)
        url = f"{BASE_URL}/v2/transactions/{tid}/refund"
        body = {
            "initiator_id": INITIATOR_ID,
            "otp": str(otp),
            "state": str(state),
            "user_code": USER_CODE
        }
        
        logging.info(f"[EKO REFUND] Verifying OTP for TID: {tid}, URL: {url}")
        
        import requests as sync_requests
        response = sync_requests.post(url, headers=headers, data=body, timeout=30)
        
        logging.info(f"[EKO REFUND] Verify HTTP: {response.status_code} | Body: {response.text[:500]}")
        
        if response.status_code == 404:
            return {"success": False, "tid": tid, "message": "EKO endpoint not found.", "http_status": 404}
        
        try:
            result = response.json()
        except Exception:
            return {"success": False, "tid": tid, "message": f"EKO returned non-JSON: HTTP {response.status_code}", "raw": response.text[:300]}
        
        logging.info(f"[EKO REFUND] Verify Response: {json.dumps(result)[:500]}")
        
        refund_data = result.get("data", {})
        success = result.get("status") == 0
        
        # Log refund in DB
        if db is not None:
            await db.eko_refund_logs.insert_one({
                "tid": tid,
                "otp_verified": success,
                "refund_tid": refund_data.get("refund_tid"),
                "refunded_amount": refund_data.get("refunded_amount") or refund_data.get("amount"),
                "new_balance": refund_data.get("balance"),
                "message": result.get("message"),
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "raw_response": result
            })
        
        return {
            "success": success,
            "tid": tid,
            "refund_tid": refund_data.get("refund_tid"),
            "refunded_amount": refund_data.get("refunded_amount") or refund_data.get("amount"),
            "new_balance": refund_data.get("balance"),
            "commission_reversed": refund_data.get("commission_reverse"),
            "message": result.get("message", "Refund processed" if success else "Refund failed"),
            "raw_response": result
        }
        
    except Exception as e:
        logging.error(f"[EKO REFUND] Verify Error for TID {tid}: {e}")
        return {"success": False, "error": str(e)}



# ==================== EKO RECONCILIATION API ====================

@router.post("/reconcile/upload")
async def reconcile_eko_excel(file: UploadFile = File(...)):
    """
    Upload Eko Excel report to reconcile with internal DB.
    
    Flow:
    1. Parse Excel for Client Reference Id, Transaction Id, Amount, Status
    2. Match with redeem_requests + bill_payment_requests by amount, consumer, date
    3. Return reconciliation report with PRC refund status
    """
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    from openpyxl import load_workbook
    import io
    
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=False)
        ws = wb.active
        
        # Find header row
        headers = []
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=False), 1):
            row_vals = [str(cell.value or "").strip().lower() for cell in row]
            # Match on any of these known Eko header patterns
            if any(kw in " ".join(row_vals) for kw in ["eko transaction id", "transaction id", "client reference id"]):
                headers = row_vals
                header_row = row_idx
                break
        
        if not headers:
            # Try first row as header
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(v or "").strip().lower() for v in first_row]
            header_row = 1
        
        # Map column indices - MATCH actual Eko Excel headers
        col_map = {}
        for idx, h in enumerate(headers):
            h_clean = h.strip().lower()
            if h_clean in ["eko transaction id", "transaction id"] and "client" not in h_clean:
                col_map["eko_tid"] = idx
            elif "client" in h_clean and "ref" in h_clean:
                col_map["client_ref_id"] = idx
            elif h_clean in ["amount", "amount(rs.)", "amount (rs.)", "amount(rs)"]:
                col_map["amount"] = idx
            elif h_clean == "status":
                col_map["status"] = idx
            elif h_clean in ["cellnumber", "cell number", "customer id", "customer_id", "mobile"]:
                col_map["customer_id"] = idx
            elif h_clean in ["operator", "operator name", "operator_name"]:
                col_map["operator"] = idx
            elif h_clean in ["date", "timestamp", "transaction date", "transaction_date"]:
                col_map["date"] = idx
            elif h_clean in ["description"]:
                col_map["description"] = idx
            elif h_clean in ["activity"]:
                col_map["activity"] = idx
            elif h_clean in ["debit/credit", "debit_credit", "type"]:
                col_map["debit_credit"] = idx
            elif h_clean in ["fee"]:
                col_map["fee"] = idx
            elif h_clean in ["commission(rs.)", "commission", "commission(rs)"]:
                col_map["commission"] = idx
            elif h_clean in ["running balance(rs.)", "running balance", "balance"]:
                col_map["running_balance"] = idx
            elif h_clean in ["name"]:
                col_map["name"] = idx
            elif h_clean in ["bankrefno", "bank_ref_no", "utr"]:
                col_map["bank_ref"] = idx
        
        logging.info(f"[RECONCILE] Excel headers: {headers}")
        logging.info(f"[RECONCILE] Column mapping: {col_map}")
        
        # Parse rows
        excel_entries = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not any(row):
                continue
            
            entry = {}
            for key, idx in col_map.items():
                if idx < len(row):
                    entry[key] = str(row[idx] or "").strip()
            
            # Skip empty entries
            if not entry.get("eko_tid") and not entry.get("client_ref_id"):
                continue
            
            excel_entries.append(entry)
        
        wb.close()
        
        if not excel_entries:
            return {"success": False, "message": "No valid entries found in Excel"}
        
        logging.info(f"[RECONCILE] Parsed {len(excel_entries)} entries from Excel")
        
        # BATCH DB LOOKUP - much faster than individual queries
        # Collect all eko_tids and client_ref_ids for batch matching
        all_eko_tids = [e.get("eko_tid", "") for e in excel_entries if e.get("eko_tid") and e.get("eko_tid") not in ("N/A", "None", "")]
        all_client_refs = [e.get("client_ref_id", "") for e in excel_entries if e.get("client_ref_id") and e.get("client_ref_id") not in ("N/A", "None", "")]
        
        # Batch fetch from both collections by eko_tid and client_ref_id
        redeem_by_tid = {}
        redeem_by_ref = {}
        bill_by_tid = {}
        bill_by_ref = {}
        
        if all_eko_tids:
            cursor = db.redeem_requests.find({"eko_tid": {"$in": all_eko_tids}}, {"_id": 0})
            async for doc in cursor:
                if doc.get("eko_tid"):
                    redeem_by_tid[doc["eko_tid"]] = doc
            
            cursor = db.bill_payment_requests.find({"eko_tid": {"$in": all_eko_tids}}, {"_id": 0})
            async for doc in cursor:
                if doc.get("eko_tid"):
                    bill_by_tid[doc["eko_tid"]] = doc
        
        if all_client_refs:
            cursor = db.redeem_requests.find({"client_ref_id": {"$in": all_client_refs}}, {"_id": 0})
            async for doc in cursor:
                if doc.get("client_ref_id"):
                    redeem_by_ref[doc["client_ref_id"]] = doc
            
            cursor = db.bill_payment_requests.find({"client_ref_id": {"$in": all_client_refs}}, {"_id": 0})
            async for doc in cursor:
                if doc.get("client_ref_id"):
                    bill_by_ref[doc["client_ref_id"]] = doc
        
        # FALLBACK: Build lookup indexes by amount+mobile for failed transactions
        # These failed records don't have eko_tid or client_ref_id stored
        # Key format: "amount|mobile" -> list of docs
        redeem_by_amt_mobile = {}
        bill_by_amt_mobile = {}
        
        # Get all unique amounts and mobiles from Excel
        all_amounts = set()
        all_mobiles = set()
        for e in excel_entries:
            try:
                amt = float(e.get("amount", "0").replace(",", ""))
                if amt > 0:
                    all_amounts.add(amt)
            except:
                pass
            mob = e.get("customer_id", "")
            if mob and mob not in ("N/A", "None", ""):
                all_mobiles.add(mob)
        
        if all_amounts and all_mobiles:
            # Fetch all redeem_requests that match ANY of these amounts and mobiles
            cursor = db.redeem_requests.find(
                {
                    "amount_inr": {"$in": list(all_amounts)},
                    "$or": [
                        {"details.mobile_number": {"$in": list(all_mobiles)}},
                        {"details.consumer_number": {"$in": list(all_mobiles)}},
                        {"user_mobile": {"$in": list(all_mobiles)}}
                    ]
                },
                {"_id": 0}
            )
            async for doc in cursor:
                amt = doc.get("amount_inr", 0)
                mob = doc.get("details", {}).get("mobile_number") or doc.get("details", {}).get("consumer_number") or doc.get("user_mobile", "")
                if amt and mob:
                    key = f"{float(amt)}|{mob}"
                    if key not in redeem_by_amt_mobile:
                        redeem_by_amt_mobile[key] = []
                    redeem_by_amt_mobile[key].append(doc)
            
            # Same for bill_payment_requests
            cursor = db.bill_payment_requests.find(
                {
                    "amount_inr": {"$in": list(all_amounts)},
                    "$or": [
                        {"details.phone_number": {"$in": list(all_mobiles)}},
                        {"details.consumer_number": {"$in": list(all_mobiles)}}
                    ]
                },
                {"_id": 0}
            )
            async for doc in cursor:
                amt = doc.get("amount_inr", 0)
                mob = doc.get("details", {}).get("phone_number") or doc.get("details", {}).get("consumer_number", "")
                if amt and mob:
                    key = f"{float(amt)}|{mob}"
                    if key not in bill_by_amt_mobile:
                        bill_by_amt_mobile[key] = []
                    bill_by_amt_mobile[key].append(doc)
        
        logging.info(f"[RECONCILE] Batch DB: tid={len(redeem_by_tid)}/{len(bill_by_tid)}, ref={len(redeem_by_ref)}/{len(bill_by_ref)}, amt_mob={len(redeem_by_amt_mobile)}/{len(bill_by_amt_mobile)}")
        
        # Now cross-reference with DB
        results = []
        stats = {
            "total_excel": len(excel_entries),
            "matched": 0,
            "unmatched": 0,
            "eko_success_internal_failed": 0,
            "eko_fail": 0,
            "eko_refunded": 0,
            "prc_refunded_count": 0,
            "prc_not_refunded_count": 0,
            "needs_prc_reclaim": 0,
            "total_prc_to_reclaim": 0,
            "needs_eko_refund": 0,
            "eko_success_count": 0,
            "eko_fail_count": 0,
            "eko_refunded_count": 0,
            "total_amount": 0
        }
        
        for entry in excel_entries:
            eko_tid = entry.get("eko_tid", "")
            client_ref_id = entry.get("client_ref_id", "")
            eko_amount = entry.get("amount", "0")
            eko_status = entry.get("status", "").lower()
            customer_id = entry.get("customer_id", "")
            
            try:
                eko_amount_num = float(eko_amount.replace(",", ""))
            except:
                eko_amount_num = 0
            
            # Fast lookup from pre-fetched batch data
            db_match = None
            match_source = None
            
            # Strategy 1: Match by eko_tid
            if eko_tid and eko_tid not in ("N/A", "None", ""):
                if eko_tid in redeem_by_tid:
                    db_match = redeem_by_tid[eko_tid]
                    match_source = "redeem_requests"
                elif eko_tid in bill_by_tid:
                    db_match = bill_by_tid[eko_tid]
                    match_source = "bill_payment_requests"
            
            # Strategy 2: Match by client_ref_id
            if not db_match and client_ref_id:
                if client_ref_id in redeem_by_ref:
                    db_match = redeem_by_ref[client_ref_id]
                    match_source = "redeem_requests"
                elif client_ref_id in bill_by_ref:
                    db_match = bill_by_ref[client_ref_id]
                    match_source = "bill_payment_requests"
            
            # Strategy 3: Match by amount + mobile number (for failed txns without eko_tid/client_ref_id)
            if not db_match and eko_amount_num > 0 and customer_id:
                lookup_key = f"{eko_amount_num}|{customer_id}"
                
                # Parse Eko date for time-based filtering
                eko_date_str = entry.get("date", "")
                
                # Try redeem_requests first
                candidates = redeem_by_amt_mobile.get(lookup_key, [])
                if candidates:
                    # Pick the best match - prefer closest date, unmatched records
                    best = None
                    for c in candidates:
                        if c.get("_reconcile_used"):
                            continue
                        if best is None:
                            best = c
                        else:
                            # Prefer failed status (those are likely the phantom failures)
                            if c.get("status") in ("failed", "retry_failed", "eko_failed") and best.get("status") not in ("failed", "retry_failed", "eko_failed"):
                                best = c
                    if best:
                        db_match = best
                        best["_reconcile_used"] = True
                        match_source = "redeem_requests"
                
                # Try bill_payment_requests
                if not db_match:
                    candidates = bill_by_amt_mobile.get(lookup_key, [])
                    if candidates:
                        best = None
                        for c in candidates:
                            if c.get("_reconcile_used"):
                                continue
                            if best is None:
                                best = c
                            elif c.get("status") in ("failed", "retry_failed", "eko_failed") and best.get("status") not in ("failed", "retry_failed", "eko_failed"):
                                best = c
                        if best:
                            db_match = best
                            best["_reconcile_used"] = True
                            match_source = "bill_payment_requests"
            
            # Track status counts
            if eko_status in ["success", "0"]:
                stats["eko_success_count"] += 1
            elif eko_status in ["fail", "failed", "1"]:
                stats["eko_fail_count"] += 1
            elif eko_status in ["refunded", "3"]:
                stats["eko_refunded_count"] += 1
            stats["total_amount"] += eko_amount_num
            
            # Build result entry with ALL Excel fields
            result_entry = {
                "eko_tid": eko_tid,
                "client_ref_id": client_ref_id,
                "eko_amount": eko_amount,
                "eko_status": entry.get("status", ""),
                "customer_id": customer_id,
                "operator": entry.get("operator", ""),
                "date": entry.get("date", ""),
                "activity": entry.get("activity", ""),
                "debit_credit": entry.get("debit_credit", ""),
                "fee": entry.get("fee", ""),
                "commission": entry.get("commission", ""),
                "running_balance": entry.get("running_balance", ""),
                "name": entry.get("name", ""),
                "bank_ref": entry.get("bank_ref", ""),
                "matched": db_match is not None,
                "match_source": match_source
            }
            
            if db_match:
                stats["matched"] += 1
                internal_status = db_match.get("status", "unknown")
                prc_refunded = db_match.get("prc_refunded", False)
                prc_amount = db_match.get("total_prc_deducted", 0) or db_match.get("refund_amount", 0)
                
                result_entry.update({
                    "request_id": db_match.get("request_id", ""),
                    "internal_status": internal_status,
                    "user_id": db_match.get("user_id", ""),
                    "user_name": db_match.get("user_name", ""),
                    "prc_refunded": prc_refunded,
                    "prc_amount": prc_amount,
                    "internal_eko_tid": db_match.get("eko_tid"),
                    "created_at": db_match.get("created_at", "")
                })
                
                # Determine action needed
                if eko_status in ["success", "0"]:
                    if internal_status in ["failed", "retry_failed", "eko_failed"]:
                        stats["eko_success_internal_failed"] += 1
                        result_entry["action"] = "FIX_STATUS"
                        result_entry["action_detail"] = "Eko Success but internally Failed"
                        
                        if prc_refunded:
                            stats["needs_prc_reclaim"] += 1
                            stats["total_prc_to_reclaim"] += prc_amount
                            result_entry["action"] = "FIX_STATUS_RECLAIM_PRC"
                            result_entry["action_detail"] = f"Eko Success, Internal Failed, PRC {prc_amount} was refunded - needs re-deduction"
                            stats["prc_refunded_count"] += 1
                        else:
                            stats["prc_not_refunded_count"] += 1
                    else:
                        result_entry["action"] = "OK"
                        result_entry["action_detail"] = "Status matches"
                
                elif eko_status in ["fail", "failed", "1"]:
                    stats["eko_fail"] += 1
                    if not prc_refunded:
                        stats["needs_eko_refund"] += 1
                        result_entry["action"] = "NEEDS_REFUND"
                        result_entry["action_detail"] = "Eko Failed but PRC not refunded"
                    else:
                        result_entry["action"] = "OK"
                        result_entry["action_detail"] = "Both failed, PRC refunded"
                
                elif eko_status in ["refunded", "3"]:
                    stats["eko_refunded"] += 1
                    result_entry["action"] = "OK"
                    result_entry["action_detail"] = "Already refunded on Eko side"
                
                else:
                    result_entry["action"] = "REVIEW"
                    result_entry["action_detail"] = f"Unknown Eko status: {eko_status}"
            else:
                stats["unmatched"] += 1
                # For unmatched Eko Success - these need new DB records
                if eko_status in ["success", "0"]:
                    result_entry["action"] = "CREATE_COMPLETED"
                    result_entry["action_detail"] = "Eko Success but NO record in DB - create as completed"
                    stats["eko_success_internal_failed"] += 1
                else:
                    result_entry["action"] = "UNMATCHED"
                    result_entry["action_detail"] = "No matching record found in DB"
            
            results.append(result_entry)
        
        # Sort: action items first
        action_order = {"FIX_STATUS_RECLAIM_PRC": 0, "CREATE_COMPLETED": 1, "FIX_STATUS": 2, "NEEDS_REFUND": 3, "REVIEW": 4, "UNMATCHED": 5, "OK": 6}
        results.sort(key=lambda x: action_order.get(x.get("action", "OK"), 99))
        
        return {
            "success": True,
            "stats": stats,
            "results": results,
            "message": f"Reconciled {len(excel_entries)} entries. {stats['matched']} matched, {stats['unmatched']} unmatched. {stats['eko_success_internal_failed']} need status fix."
        }
        
    except Exception as e:
        logging.error(f"[RECONCILE] Error: {e}")
        import traceback
        logging.error(f"[RECONCILE] Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Reconciliation failed: {str(e)}")


@router.post("/reconcile/fix")
async def apply_reconciliation_fixes(request: Request):
    """
    Apply fixes for reconciled transactions.
    
    Accepts a list of fixes to apply:
    - FIX_STATUS: Update internal status to "completed" + store eko_tid
    - FIX_STATUS_RECLAIM_PRC: Same + re-deduct PRC from user
    - REFUND_PRC: Refund PRC for Eko-failed transactions
    """
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    data = await request.json()
    fixes = data.get("fixes", [])
    
    if not fixes:
        return {"success": False, "message": "No fixes provided"}
    
    results = []
    fixed_count = 0
    error_count = 0
    total_prc_reclaimed = 0
    total_prc_refunded = 0
    
    for fix in fixes:
        request_id = fix.get("request_id")
        action = fix.get("action")
        eko_tid = fix.get("eko_tid")
        match_source = fix.get("match_source", "redeem_requests")
        
        if not action:
            results.append({"request_id": request_id, "status": "skipped", "reason": "Missing action"})
            continue
        
        try:
            now_str = datetime.now(timezone.utc).isoformat()
            
            # CREATE_COMPLETED: No DB record exists - create new one
            if action == "CREATE_COMPLETED":
                eko_amount = fix.get("eko_amount", 0)
                customer_id = fix.get("customer_id", "")
                eko_date = fix.get("date", "")
                
                try:
                    eko_amount_num = float(str(eko_amount).replace(",", ""))
                except:
                    eko_amount_num = 0
                
                new_request_id = f"RECON-{eko_tid or int(datetime.now(timezone.utc).timestamp())}"
                
                new_doc = {
                    "request_id": new_request_id,
                    "service_type": "mobile_recharge",
                    "amount_inr": eko_amount_num,
                    "amount": eko_amount_num,
                    "details": {"mobile_number": customer_id},
                    "status": "completed",
                    "eko_tid": eko_tid,
                    "eko_status": "SUCCESS",
                    "client_ref_id": fix.get("client_ref_id"),
                    "reconciled": True,
                    "reconciled_at": now_str,
                    "reconcile_note": "Created via Eko Excel reconciliation - Eko Success but no internal record",
                    "completed_at": eko_date or now_str,
                    "created_at": eko_date or now_str,
                    "updated_at": now_str,
                    "user_id": "UNKNOWN",
                    "user_name": "Reconciled Entry",
                    "total_prc_deducted": 0,
                    "prc_refunded": False,
                    "status_history": [{
                        "status": "completed",
                        "timestamp": now_str,
                        "note": f"Reconciled: Eko Success (TID: {eko_tid}), no internal record found"
                    }]
                }
                
                await db.redeem_requests.insert_one(new_doc)
                del new_doc["_id"]  # Remove MongoDB added _id
                
                fixed_count += 1
                results.append({
                    "request_id": new_request_id,
                    "status": "created",
                    "action": "CREATE_COMPLETED",
                    "eko_tid": eko_tid,
                    "amount": eko_amount_num
                })
                logging.info(f"[RECONCILE FIX] Created new record {new_request_id} for Eko TID {eko_tid}")
                continue
            
            # For all other actions, record must exist in DB
            if not request_id:
                results.append({"request_id": request_id, "status": "skipped", "reason": "Missing request_id"})
                continue
            
            collection = db.redeem_requests if match_source == "redeem_requests" else db.bill_payment_requests
        
            doc = await collection.find_one({"request_id": request_id}, {"_id": 0})
            if not doc:
                results.append({"request_id": request_id, "status": "error", "reason": "Record not found"})
                error_count += 1
                continue
            
            now_str = datetime.now(timezone.utc).isoformat()
            
            if action in ["FIX_STATUS", "FIX_STATUS_RECLAIM_PRC"]:
                # Update status to completed
                update_fields = {
                    "status": "completed",
                    "eko_tid": eko_tid or doc.get("eko_tid"),
                    "eko_status": "SUCCESS",
                    "reconciled": True,
                    "reconciled_at": now_str,
                    "reconcile_note": "Fixed via Eko Excel reconciliation",
                    "completed_at": now_str,
                    "updated_at": now_str
                }
                
                status_update = {
                    "status": "completed",
                    "timestamp": now_str,
                    "note": f"Reconciled: Eko confirmed SUCCESS (TID: {eko_tid})"
                }
                
                await collection.update_one(
                    {"request_id": request_id},
                    {
                        "$set": update_fields,
                        "$push": {"status_history": status_update}
                    }
                )
                
                # Reclaim PRC if it was wrongly refunded
                if action == "FIX_STATUS_RECLAIM_PRC":
                    user_id = doc.get("user_id")
                    prc_amount = doc.get("total_prc_deducted", 0) or doc.get("refund_amount", 0)
                    
                    if user_id and prc_amount > 0:
                        # Re-deduct PRC from user
                        user = await db.users.find_one({"uid": user_id}, {"_id": 0, "prc_balance": 1})
                        if user:
                            old_balance = user.get("prc_balance", 0)
                            new_balance = old_balance - prc_amount
                            
                            await db.users.update_one(
                                {"uid": user_id},
                                {"$set": {"prc_balance": new_balance}}
                            )
                            
                            # Log the reclaim transaction
                            await db.transactions.insert_one({
                                "user_id": user_id,
                                "type": "prc_reclaim",
                                "amount": -prc_amount,
                                "balance_before": old_balance,
                                "balance_after": new_balance,
                                "description": f"PRC Reclaim: Eko payment was successful (TID: {eko_tid}), wrongly refunded PRC reversed",
                                "reference": request_id,
                                "created_at": now_str
                            })
                            
                            # Update the record to reflect PRC reclaim
                            await collection.update_one(
                                {"request_id": request_id},
                                {"$set": {
                                    "prc_reclaimed": True,
                                    "prc_reclaim_amount": prc_amount,
                                    "prc_reclaim_at": now_str,
                                    "prc_balance_before_reclaim": old_balance,
                                    "prc_balance_after_reclaim": new_balance
                                }}
                            )
                            
                            total_prc_reclaimed += prc_amount
                            logging.info(f"[RECONCILE FIX] Reclaimed {prc_amount} PRC from user {user_id}. Balance: {old_balance} → {new_balance}")
                
                fixed_count += 1
                results.append({
                    "request_id": request_id,
                    "status": "fixed",
                    "action": action,
                    "eko_tid": eko_tid,
                    "prc_reclaimed": total_prc_reclaimed if action == "FIX_STATUS_RECLAIM_PRC" else 0
                })
            
            elif action == "REFUND_PRC":
                # Refund PRC for failed transactions where PRC wasn't refunded
                user_id = doc.get("user_id")
                prc_amount = doc.get("total_prc_deducted", 0)
                
                if user_id and prc_amount > 0 and not doc.get("prc_refunded"):
                    user = await db.users.find_one({"uid": user_id}, {"_id": 0, "prc_balance": 1})
                    if user:
                        old_balance = user.get("prc_balance", 0)
                        new_balance = old_balance + prc_amount
                        
                        await db.users.update_one(
                            {"uid": user_id},
                            {"$set": {"prc_balance": new_balance}}
                        )
                        
                        await db.transactions.insert_one({
                            "user_id": user_id,
                            "type": "prc_refund",
                            "amount": prc_amount,
                            "balance_before": old_balance,
                            "balance_after": new_balance,
                            "description": f"PRC Refund: Eko confirmed FAILED (TID: {eko_tid})",
                            "reference": request_id,
                            "created_at": now_str
                        })
                        
                        await collection.update_one(
                            {"request_id": request_id},
                            {"$set": {
                                "prc_refunded": True,
                                "refund_amount": prc_amount,
                                "reconciled": True,
                                "reconciled_at": now_str
                            }}
                        )
                        
                        total_prc_refunded += prc_amount
                
                fixed_count += 1
                results.append({
                    "request_id": request_id,
                    "status": "fixed",
                    "action": "REFUND_PRC",
                    "prc_refunded": prc_amount
                })
            
            else:
                results.append({"request_id": request_id, "status": "skipped", "reason": f"Unknown action: {action}"})
            
        except Exception as e:
            logging.error(f"[RECONCILE FIX] Error fixing {request_id}: {e}")
            results.append({"request_id": request_id, "status": "error", "reason": str(e)})
            error_count += 1
    
    return {
        "success": True,
        "fixed": fixed_count,
        "errors": error_count,
        "total_prc_reclaimed": total_prc_reclaimed,
        "total_prc_refunded": total_prc_refunded,
        "results": results,
        "message": f"Applied {fixed_count} fixes. PRC Reclaimed: {total_prc_reclaimed}, PRC Refunded: {total_prc_refunded}"
    }


@router.get("/reconcile/prc-check/{request_id}")
async def check_prc_refund_status(request_id: str):
    """Check if PRC was refunded for a specific request."""
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    # Check both collections
    doc = await db.redeem_requests.find_one({"request_id": request_id}, {"_id": 0})
    source = "redeem_requests"
    if not doc:
        doc = await db.bill_payment_requests.find_one({"request_id": request_id}, {"_id": 0})
        source = "bill_payment_requests"
    
    if not doc:
        raise HTTPException(status_code=404, detail="Request not found")
    
    user_id = doc.get("user_id")
    user = await db.users.find_one({"uid": user_id}, {"_id": 0, "prc_balance": 1, "name": 1})
    
    return {
        "request_id": request_id,
        "source": source,
        "status": doc.get("status"),
        "prc_refunded": doc.get("prc_refunded", False),
        "prc_deducted": doc.get("total_prc_deducted", 0),
        "refund_amount": doc.get("refund_amount", 0),
        "eko_tid": doc.get("eko_tid"),
        "user_id": user_id,
        "user_name": doc.get("user_name", ""),
        "user_current_balance": user.get("prc_balance", 0) if user else None,
        "created_at": doc.get("created_at", ""),
        "reconciled": doc.get("reconciled", False)
    }
