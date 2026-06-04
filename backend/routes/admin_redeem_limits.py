"""
Admin Redeem Limits dashboard + Direct Redeem (June 2026)
=========================================================

Endpoints
---------
GET  /api/admin/redeem-limits/users
        List users with Total PRC, Redeem Limit, Used, Balance Redeemable.
        Supports search (name/mobile), active/inactive filter, sort, pagination.
GET  /api/admin/redeem-limits/users/export-excel
        Stream the same dataset as an Excel (.xlsx) download.
POST /api/admin/redeem-limits/direct-redeem
        Two-step direct-redeem (admin debits user's PRC and creates an instant
        "paid" bank_transfer_request). Honours progressive minimum.
"""
import logging
import math
from datetime import datetime, timezone
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

router = APIRouter()

db = None  # injected via set_db()
# Optional dependency injectors — server.py exposes these helpers globally.
_calculate_redeem_limit_func = None  # async (uid) -> redeem-limit dict
_create_success_story_post_func = None  # async kwargs
_compute_progressive_min_func = None  # async (uid) -> progressive-min dict


def set_db(database):
    global db
    db = database


def set_helpers(*, calc_redeem_limit=None, create_success_story=None, compute_progressive_min=None):
    """Inject server-side helpers so we avoid circular imports with server.py."""
    global _calculate_redeem_limit_func, _create_success_story_post_func, _compute_progressive_min_func
    if calc_redeem_limit is not None:
        _calculate_redeem_limit_func = calc_redeem_limit
    if create_success_story is not None:
        _create_success_story_post_func = create_success_story
    if compute_progressive_min is not None:
        _compute_progressive_min_func = compute_progressive_min


# ==================== HELPERS ====================

def _is_active_elite(user: dict) -> bool:
    """Active = subscription_plan==elite AND subscription not expired."""
    plan = (user.get("subscription_plan") or "explorer").lower()
    if plan != "elite":
        return False
    if user.get("subscription_expired", False):
        return False
    expiry = user.get("subscription_expiry") or user.get("subscription_expires") or user.get("vip_expiry")
    if not expiry:
        return False
    try:
        if isinstance(expiry, datetime):
            exp_dt = expiry if expiry.tzinfo else expiry.replace(tzinfo=timezone.utc)
        else:
            exp_dt = datetime.fromisoformat(str(expiry).replace("Z", "+00:00"))
            if exp_dt.tzinfo is None:
                exp_dt = exp_dt.replace(tzinfo=timezone.utc)
        return exp_dt > datetime.now(timezone.utc)
    except Exception:
        return False


async def _row_for_user(user: dict) -> dict:
    """Compute a single user's redeem-limit row (used by both list + Excel export)."""
    uid = user.get("uid")
    rl = await _calculate_redeem_limit_func(uid) if _calculate_redeem_limit_func else {}
    total_prc = float(user.get("prc_balance", 0) or 0)
    redeem_limit_prc = float(rl.get("redeemable", 0) or 0)  # absolute cap
    used_prc = float(rl.get("total_redeemed", 0) or 0)
    balance_redeemable_prc = float(rl.get("effective_available", 0) or 0)
    return {
        "uid": uid,
        "name": user.get("name") or "—",
        "mobile": user.get("mobile") or user.get("phone") or "—",
        "subscription_plan": (user.get("subscription_plan") or "explorer").lower(),
        "is_active_elite": _is_active_elite(user),
        "total_prc": round(total_prc, 2),
        "redeem_limit_prc": round(redeem_limit_prc, 2),
        "used_prc": round(used_prc, 2),
        "balance_redeemable_prc": round(balance_redeemable_prc, 2),
        "unlock_percent": rl.get("unlock_percent", 0),
        "network_size": rl.get("network_size", 0),
        "bank": {
            "account_holder_name": user.get("bank_account_holder_name") or user.get("name"),
            "account_number": user.get("bank_account_number"),
            "ifsc_code": user.get("bank_ifsc_code"),
            "bank_name": user.get("bank_name"),
            "upi_id": user.get("upi_id"),
            "phonepe_gpay_number": user.get("phonepe_gpay_number"),
        },
    }


# ==================== USERS LIST ====================

@router.get("/users")
async def list_redeem_limit_users(
    search: Optional[str] = Query(None, description="Search by name or mobile"),
    status: str = Query("all", regex="^(all|active|inactive)$"),
    sort_by: str = Query("balance_redeemable", regex="^(name|mobile|total_prc|redeem_limit|used|balance_redeemable)$"),
    sort_order: str = Query("desc", regex="^(asc|desc)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
):
    """Paginated list with filters/sort/search."""
    if db is None:
        raise HTTPException(status_code=500, detail="DB not initialised")

    base_query: dict = {}
    if search:
        s = search.strip()
        # Case-insensitive partial match on name OR mobile
        base_query["$or"] = [
            {"name": {"$regex": s, "$options": "i"}},
            {"mobile": {"$regex": s}},
            {"phone": {"$regex": s}},
        ]

    # Fetch all matching users (we compute redeem-limit + sort in-app since
    # those fields are derived, not stored on the user doc).
    projection = {
        "_id": 0, "uid": 1, "name": 1, "mobile": 1, "phone": 1,
        "prc_balance": 1, "subscription_plan": 1, "subscription_expiry": 1,
        "subscription_expires": 1, "vip_expiry": 1, "subscription_expired": 1,
        "bank_account_holder_name": 1, "bank_account_number": 1,
        "bank_ifsc_code": 1, "bank_name": 1, "upi_id": 1,
        "phonepe_gpay_number": 1,
    }
    cursor = db.users.find(base_query, projection)
    users = await cursor.to_list(50000)

    rows = []
    for u in users:
        row = await _row_for_user(u)
        if status == "active" and not row["is_active_elite"]:
            continue
        if status == "inactive" and row["is_active_elite"]:
            continue
        rows.append(row)

    sort_field_map = {
        "name": lambda r: (r["name"] or "").lower(),
        "mobile": lambda r: r["mobile"] or "",
        "total_prc": lambda r: r["total_prc"],
        "redeem_limit": lambda r: r["redeem_limit_prc"],
        "used": lambda r: r["used_prc"],
        "balance_redeemable": lambda r: r["balance_redeemable_prc"],
    }
    rows.sort(key=sort_field_map[sort_by], reverse=(sort_order == "desc"))

    total = len(rows)
    start = (page - 1) * page_size
    end = start + page_size
    return {
        "success": True,
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": math.ceil(total / page_size) if page_size else 1,
        "filters": {"search": search, "status": status, "sort_by": sort_by, "sort_order": sort_order},
        "rows": rows[start:end],
    }


# ==================== EXCEL EXPORT ====================

@router.get("/users/export-excel")
async def export_redeem_limits_excel(
    search: Optional[str] = Query(None),
    status: str = Query("all", regex="^(all|active|inactive)$"),
    sort_by: str = Query("balance_redeemable", regex="^(name|mobile|total_prc|redeem_limit|used|balance_redeemable)$"),
    sort_order: str = Query("desc", regex="^(asc|desc)$"),
):
    """Stream the full (un-paginated) dataset as an Excel file."""
    # Reuse the list endpoint with a huge page_size to get all rows
    listing = await list_redeem_limit_users(
        search=search, status=status, sort_by=sort_by, sort_order=sort_order,
        page=1, page_size=200,
    )
    # We still need everything — re-fetch without pagination
    all_rows = []
    for p in range(1, listing["pages"] + 1):
        sub = await list_redeem_limit_users(
            search=search, status=status, sort_by=sort_by, sort_order=sort_order,
            page=p, page_size=200,
        )
        all_rows.extend(sub["rows"])

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"openpyxl not installed: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Redeem Limits"

    headers = [
        "User Name", "Mobile", "Plan", "Active Elite",
        "Total PRC", "Redeem Limit (PRC)", "Used PRC", "Balance Redeemable (PRC)",
        "Unlock %", "Network Size",
        "Bank A/C Holder", "Bank A/C Number", "IFSC", "Bank Name",
        "UPI ID", "PhonePe/GPay Number",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="7C3AED")
    for col_idx, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, r in enumerate(all_rows, 2):
        ws.cell(row=row_idx, column=1, value=r["name"])
        ws.cell(row=row_idx, column=2, value=r["mobile"])
        ws.cell(row=row_idx, column=3, value=r["subscription_plan"])
        ws.cell(row=row_idx, column=4, value="Yes" if r["is_active_elite"] else "No")
        ws.cell(row=row_idx, column=5, value=r["total_prc"])
        ws.cell(row=row_idx, column=6, value=r["redeem_limit_prc"])
        ws.cell(row=row_idx, column=7, value=r["used_prc"])
        ws.cell(row=row_idx, column=8, value=r["balance_redeemable_prc"])
        ws.cell(row=row_idx, column=9, value=r["unlock_percent"])
        ws.cell(row=row_idx, column=10, value=r["network_size"])
        b = r["bank"] or {}
        ws.cell(row=row_idx, column=11, value=b.get("account_holder_name") or "")
        ws.cell(row=row_idx, column=12, value=str(b.get("account_number") or ""))
        ws.cell(row=row_idx, column=13, value=b.get("ifsc_code") or "")
        ws.cell(row=row_idx, column=14, value=b.get("bank_name") or "")
        ws.cell(row=row_idx, column=15, value=b.get("upi_id") or "")
        ws.cell(row=row_idx, column=16, value=str(b.get("phonepe_gpay_number") or ""))

    # Auto-size key columns
    widths = [22, 14, 10, 12, 14, 16, 14, 18, 10, 12, 22, 22, 14, 22, 24, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="redeem-limits-{ts}.xlsx"'},
    )


# ==================== ADMIN DIRECT REDEEM ====================

class DirectRedeemRequest(BaseModel):
    admin_id: str
    user_id: str
    amount_inr: int = Field(..., gt=0, description="Amount in INR to debit")
    utr_number: str = Field(..., min_length=4, description="Real UTR / transaction reference (mandatory)")
    remark: Optional[str] = Field(None, max_length=200)
    confirm: bool = Field(False, description="Must be true for the actual debit; false runs as preview only")


@router.post("/direct-redeem")
async def admin_direct_redeem(request: DirectRedeemRequest):
    """Two-step admin direct-redeem.

    Step 1: confirm=False → returns a PREVIEW (fees, post-deduction balance,
            next progressive minimum) so the admin can review before committing.
    Step 2: confirm=True  → actually debits PRC, creates a "paid" bank_transfer_request
            with the supplied UTR, posts a Community success story, raises the user's
            progressive minimum, and writes a transaction record.
    """
    if db is None:
        raise HTTPException(status_code=500, detail="DB not initialised")

    user = await db.users.find_one({"uid": request.user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found.")

    # Fees + PRC math
    from utils.helpers import get_prc_rate
    prc_rate = await get_prc_rate(db)
    TRANSACTION_FEE = 10
    ADMIN_FEE_PERCENT = 20
    admin_fee = int(request.amount_inr * ADMIN_FEE_PERCENT / 100)
    total_inr = request.amount_inr + admin_fee + TRANSACTION_FEE
    total_prc = total_inr * prc_rate

    # Progressive minimum (Q4 = applies even for admin direct)
    progressive_min = 0
    progressive_max = None
    if _compute_progressive_min_func:
        prog = await _compute_progressive_min_func(request.user_id)
        progressive_min = prog["minimum"]
        progressive_max = prog["maximum"]

    rl = await _calculate_redeem_limit_func(request.user_id) if _calculate_redeem_limit_func else {}
    effective_available = float(rl.get("effective_available", 0) or 0)
    current_balance = float(user.get("prc_balance", 0) or 0)

    # Validations (same gate for preview and commit so the admin sees the same blockers)
    blockers = []
    if request.amount_inr < progressive_min:
        blockers.append(
            f"Amount ₹{request.amount_inr:,} is below this user's minimum ₹{progressive_min:,}."
        )
    if progressive_max and request.amount_inr > progressive_max:
        blockers.append(
            f"Amount ₹{request.amount_inr:,} exceeds this user's maximum ₹{progressive_max:,}."
        )
    if total_prc > current_balance:
        blockers.append(
            f"User's PRC balance ({current_balance:,.0f}) is short for {total_prc:,.0f} PRC."
        )
    if total_prc > effective_available:
        blockers.append(
            f"User's redeem limit ({effective_available:,.0f} PRC available) is short."
        )

    # Bank details snapshot for UI
    bank_snapshot = {
        "account_holder_name": user.get("bank_account_holder_name") or user.get("name"),
        "account_number": user.get("bank_account_number"),
        "ifsc_code": user.get("bank_ifsc_code"),
        "bank_name": user.get("bank_name"),
        "upi_id": user.get("upi_id"),
        "phonepe_gpay_number": user.get("phonepe_gpay_number"),
    }

    preview_payload = {
        "preview": True,
        "user": {
            "uid": user["uid"],
            "name": user.get("name"),
            "mobile": user.get("mobile") or user.get("phone"),
            "current_balance_prc": current_balance,
            "effective_available_prc": effective_available,
        },
        "bank": bank_snapshot,
        "fees": {
            "withdrawal_inr": request.amount_inr,
            "admin_fee_inr": admin_fee,
            "transaction_fee_inr": TRANSACTION_FEE,
            "total_inr_debited": total_inr,
            "total_prc_debited": round(total_prc, 2),
            "prc_rate": prc_rate,
        },
        "progressive_min_after": int(math.ceil(request.amount_inr * 1.5)),
        "blockers": blockers,
        "can_proceed": len(blockers) == 0,
    }

    if not request.confirm:
        return {"success": True, **preview_payload}

    if blockers:
        raise HTTPException(status_code=400, detail="; ".join(blockers))

    # --------- Commit phase ---------
    if not request.utr_number:
        raise HTTPException(status_code=400, detail="UTR / transaction reference is required.")

    # 0) Idempotency guard — same admin+user+UTR must never debit twice.
    # We use a tiny dedup collection that asserts a unique compound key.
    dedup_key = f"adm:{request.admin_id}:{request.user_id}:{request.utr_number}"
    try:
        await db.admin_direct_redeem_dedup.create_index(
            "key", unique=True, name="adm_direct_dedup_uniq"
        )
    except Exception:
        pass  # Index already exists
    try:
        await db.admin_direct_redeem_dedup.insert_one({
            "key": dedup_key,
            "admin_id": request.admin_id,
            "user_id": request.user_id,
            "utr_number": request.utr_number,
            "amount_inr": request.amount_inr,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception as dup_err:
        # Duplicate key → this {admin,user,UTR} was already processed
        if "duplicate key" in str(dup_err).lower() or "e11000" in str(dup_err).lower():
            existing = await db.bank_transfer_requests.find_one(
                {"user_id": request.user_id, "utr_number": request.utr_number,
                 "channel": "admin_direct_redeem"},
                {"_id": 0, "request_id": 1, "withdrawal_amount": 1, "processed_at": 1}
            )
            raise HTTPException(
                status_code=409,
                detail={
                    "message": "This UTR was already processed for this user.",
                    "existing_request": existing,
                },
            )
        raise

    # 1) Atomic CAS debit on PRC balance (race-safe vs concurrent user spends)
    debit_res = await db.users.update_one(
        {"uid": user["uid"], "prc_balance": {"$gte": total_prc}},
        {"$inc": {"prc_balance": -total_prc, "total_redeemed_prc": total_prc},
         "$set": {"last_redeem_at": datetime.now(timezone.utc).isoformat(),
                  "last_redeem_amount_inr": request.amount_inr}},
    )
    if debit_res.modified_count == 0:
        # Roll back the dedup key so the admin can retry with a different UTR
        await db.admin_direct_redeem_dedup.delete_one({"key": dedup_key})
        raise HTTPException(status_code=409, detail="Concurrent state change — user balance moved. Retry.")

    # 2) Create a "paid" bank_transfer_request (instant — no pending queue)
    request_id = f"ADM-DIR-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}-{user['uid'][:6]}"
    now_iso = datetime.now(timezone.utc).isoformat()
    bank_req_doc = {
        "request_id": request_id,
        "user_id": user["uid"],
        "user_name": user.get("name"),
        "user_mobile": user.get("mobile") or user.get("phone"),
        "withdrawal_amount": request.amount_inr,
        "amount_inr": request.amount_inr,
        "total_inr": total_inr,
        "total_prc_deducted": round(total_prc, 2),
        "admin_fee_inr": admin_fee,
        "transaction_fee_inr": TRANSACTION_FEE,
        "bank_details": bank_snapshot,
        "status": "paid",
        "utr_number": request.utr_number,
        "admin_remark": request.remark or "Admin direct redeem",
        "processed_by": request.admin_id,
        "processed_at": now_iso,
        "created_at": now_iso,
        "updated_at": now_iso,
        "channel": "admin_direct_redeem",
    }
    try:
        await db.bank_transfer_requests.insert_one(bank_req_doc)
    except Exception as e:
        # COMPENSATING ACTION: re-credit PRC and clean up dedup key — never
        # leave a user debited without an audit row.
        logging.error(f"[ADMIN-DIRECT-REDEEM] bank_transfer insert failed, rolling back PRC: {e}")
        await db.users.update_one(
            {"uid": user["uid"]},
            {"$inc": {"prc_balance": total_prc, "total_redeemed_prc": -total_prc}}
        )
        await db.admin_direct_redeem_dedup.delete_one({"key": dedup_key})
        raise HTTPException(status_code=500, detail=f"Failed to record bank request — PRC re-credited: {e}")

    # 3) Transaction record (PRC statement)
    try:
        await db.transactions.insert_one({
            "user_id": user["uid"],
            "type": "bank_redeem_admin_direct",
            "amount_prc": -round(total_prc, 2),
            "amount_inr": -request.amount_inr,
            "balance_after": round(current_balance - total_prc, 2),
            "description": f"Admin direct redeem ₹{request.amount_inr:,} (UTR {request.utr_number})",
            "ref_id": request_id,
            "created_at": now_iso,
            "metadata": {
                "admin_id": request.admin_id,
                "utr_number": request.utr_number,
                "bank_account_last4": (bank_snapshot.get("account_number") or "")[-4:],
            },
        })
    except Exception as e:
        # PRC statement row is best-effort; bank_transfer doc is the source of
        # truth for accounting. Log and continue.
        logging.warning(f"[ADMIN-DIRECT-REDEEM] transactions insert failed (non-fatal): {e}")

    # 4) Raise progressive minimum
    try:
        new_floor = max(100, int(math.ceil(request.amount_inr * 1.5)))
        await db.users.update_one(
            {"uid": user["uid"]},
            {"$max": {"next_min_withdrawal_inr": new_floor}}
        )
    except Exception as e:
        logging.warning(f"[ADMIN-DIRECT-REDEEM] progressive-min update failed: {e}")

    # 5) Community Success Story post
    try:
        if _create_success_story_post_func:
            await _create_success_story_post_func(
                user_id=user["uid"],
                service_type="bank_redeem",
                amount_inr=float(request.amount_inr),
                ref_id=f"bank_redeem:{request_id}",
            )
    except Exception as e:
        logging.warning(f"[ADMIN-DIRECT-REDEEM] community post hook failed: {e}")

    return {
        "success": True,
        "request_id": request_id,
        "utr_number": request.utr_number,
        "amount_inr": request.amount_inr,
        "prc_deducted": round(total_prc, 2),
        "new_balance": round(current_balance - total_prc, 2),
        "next_min_withdrawal_inr": int(math.ceil(request.amount_inr * 1.5)),
    }


# ==================== USER BANK DETAILS (PROFILE) ====================

class UserBankDetailsUpdate(BaseModel):
    account_number: Optional[str] = Field(None, min_length=8, max_length=20)
    ifsc_code: Optional[str] = Field(None, min_length=11, max_length=11)
    bank_name: Optional[str] = Field(None, max_length=80)
    account_holder_name: Optional[str] = Field(None, max_length=80)
    upi_id: Optional[str] = Field(None, max_length=80)
    phonepe_gpay_number: Optional[str] = Field(None, min_length=10, max_length=15)


@router.put("/users/{uid}/bank-details")
async def upsert_user_bank_details(uid: str, payload: UserBankDetailsUpdate):
    """Allow a user (or admin) to save bank/UPI details on their profile."""
    if db is None:
        raise HTTPException(status_code=500, detail="DB not initialised")
    user = await db.users.find_one({"uid": uid}, {"_id": 0, "uid": 1})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    update_doc = {}
    if payload.account_number is not None:
        if not payload.account_number.isdigit():
            raise HTTPException(status_code=400, detail="Account number must contain digits only.")
        update_doc["bank_account_number"] = payload.account_number
    if payload.ifsc_code is not None:
        import re
        ifsc = payload.ifsc_code.upper().strip()
        if not re.match(r"^[A-Z]{4}0[A-Z0-9]{6}$", ifsc):
            raise HTTPException(status_code=400, detail="Invalid IFSC format.")
        update_doc["bank_ifsc_code"] = ifsc
    if payload.bank_name is not None:
        update_doc["bank_name"] = payload.bank_name.strip()
    if payload.account_holder_name is not None:
        update_doc["bank_account_holder_name"] = payload.account_holder_name.strip()
    if payload.upi_id is not None:
        upi = payload.upi_id.strip()
        if upi and "@" not in upi:
            raise HTTPException(status_code=400, detail="UPI ID must contain '@'.")
        update_doc["upi_id"] = upi
    if payload.phonepe_gpay_number is not None:
        pn = payload.phonepe_gpay_number.strip()
        if pn and not pn.isdigit():
            raise HTTPException(status_code=400, detail="PhonePe/GPay number must be digits only.")
        update_doc["phonepe_gpay_number"] = pn

    if not update_doc:
        return {"success": True, "updated": 0}

    update_doc["bank_details_updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.users.update_one({"uid": uid}, {"$set": update_doc})
    return {"success": True, "updated": len(update_doc) - 1, "fields": list(update_doc.keys())}


@router.get("/users/{uid}/bank-details")
async def get_user_bank_details(uid: str):
    """Return current saved bank/UPI details for the user."""
    if db is None:
        raise HTTPException(status_code=500, detail="DB not initialised")
    user = await db.users.find_one(
        {"uid": uid},
        {
            "_id": 0,
            "bank_account_holder_name": 1,
            "bank_account_number": 1,
            "bank_ifsc_code": 1,
            "bank_name": 1,
            "upi_id": 1,
            "phonepe_gpay_number": 1,
            "bank_details_updated_at": 1,
            "name": 1,
        },
    )
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return {
        "success": True,
        "account_holder_name": user.get("bank_account_holder_name") or user.get("name"),
        "account_number": user.get("bank_account_number") or "",
        "ifsc_code": user.get("bank_ifsc_code") or "",
        "bank_name": user.get("bank_name") or "",
        "upi_id": user.get("upi_id") or "",
        "phonepe_gpay_number": user.get("phonepe_gpay_number") or "",
        "updated_at": user.get("bank_details_updated_at"),
    }
