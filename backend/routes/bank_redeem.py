"""
Bank Account Redeem Routes - PRC to Bank withdrawal
Users can save bank details and request PRC withdrawal to bank account
- Cooldown: 24 hours from last request
- Fixed denominations: 100, 500, 1000, 5000, 10000, 25000
- Processing fees + 20% admin charges
- Eko DMT Integration for instant bank transfers
"""

from fastapi import APIRouter, HTTPException, Request
from datetime import datetime, timezone, timedelta
import uuid
import logging

# Create router
router = APIRouter(tags=["Bank Redeem"])

# Database reference
db = None

# Cache reference  
cache = None

# Helpers
log_transaction = None
create_notification = None

# Eko Service
eko_service = None

def set_db(database):
    global db, eko_service
    db = database
    # Initialize Eko service with database
    try:
        from services.eko_service import EkoService
        eko_service = EkoService(db)
    except ImportError:
        logging.warning("Eko service not available")

def set_cache(cache_manager):
    global cache
    cache = cache_manager

def set_helpers(helpers: dict):
    global log_transaction, create_notification
    log_transaction = helpers.get('log_transaction')
    create_notification = helpers.get('create_notification')


# ==================== EKO DMT FUNCTIONS (Using New Service) ====================

async def get_eko_service():
    """Get Eko service instance"""
    global eko_service
    if eko_service is None:
        try:
            from services.eko_service import EkoService
            eko_service = EkoService(db)
        except ImportError:
            return None
    return eko_service


def get_dynamic_prc_rate() -> int:
    """Get PRC to INR rate - delegates to single source of truth."""
    from utils.helpers import get_prc_rate_sync
    return get_prc_rate_sync()


async def eko_verify_bank_account(ifsc: str, account_number: str):
    """Verify bank account using Eko API"""
    service = await get_eko_service()
    if not service:
        return {"success": False, "error": "Eko service not available"}
    
    response = await service.verify_bank_account(ifsc, account_number)
    
    if response.success:
        return {
            "success": True,
            "account_holder": response.data.get("account_holder", ""),
            "verified": True,
            "raw_response": response.raw_response
        }
    else:
        return {
            "success": False,
            "error": response.message,
            "error_code": response.error_code
        }


async def eko_initiate_transfer(
    recipient_mobile: str,
    account_number: str,
    ifsc: str,
    amount: float,
    recipient_name: str,
    client_ref_id: str
):
    """
    Initiate bank transfer using Eko DMT
    Returns: {success: bool, txn_id: str, message: str}
    """
    service = await get_eko_service()
    if not service:
        return {"success": False, "error": "Eko service not available", "txn_id": None}
    
    try:
        from services.eko_service import EkoChannel
        
        response = await service.direct_transfer(
            account_number=account_number,
            ifsc=ifsc,
            amount=int(amount),
            recipient_name=recipient_name,
            recipient_mobile=recipient_mobile or "9999999999",
            client_ref_id=client_ref_id,
            channel=EkoChannel.IMPS
        )
        
        if response.success:
            return {
                "success": True,
                "txn_id": response.eko_tid,
                "utr_number": response.utr_number,
                "message": response.message,
                "tx_status": response.tx_status.value if response.tx_status else None,
                "tx_status_desc": response.tx_status.description if response.tx_status else None,
                "eko_response": response.raw_response
            }
        else:
            return {
                "success": False,
                "txn_id": None,
                "error": response.message,
                "error_code": response.error_code,
                "eko_response": response.raw_response
            }
    except Exception as e:
        logging.error(f"Eko transfer error: {e}")
        return {"success": False, "txn_id": None, "error": str(e)}


async def eko_check_balance():
    """Check Eko settlement account balance"""
    service = await get_eko_service()
    if not service:
        return {"balance": 0, "error": "Eko service not available"}
    
    response = await service.check_wallet_balance()
    
    if response.success:
        return {
            "balance": response.data.get("balance", 0),
            "success": True
        }
    return {"balance": 0, "error": response.message}


async def eko_check_transaction_status(transaction_id: str, use_client_ref: bool = False):
    """Check transaction status from Eko"""
    service = await get_eko_service()
    if not service:
        return {"success": False, "error": "Eko service not available"}
    
    response = await service.check_transaction_status(transaction_id, use_client_ref)
    
    return {
        "success": response.success,
        "tx_status": response.tx_status.value if response.tx_status else None,
        "tx_status_desc": response.tx_status.description if response.tx_status else None,
        "utr_number": response.utr_number,
        "eko_tid": response.eko_tid,
        "message": response.message,
        "data": response.data
    }


def get_current_week_monday():
    """
    DEPRECATED: Now using 7-day rolling window from last request
    Kept for backward compatibility
    """
    now = datetime.now(timezone.utc)
    days_since_monday = now.weekday()
    monday = (now - timedelta(days=days_since_monday)).replace(hour=0, minute=0, second=0, microsecond=0)
    next_monday = monday + timedelta(days=7)
    return monday, next_monday


def get_7_days_ago():
    """Get datetime from 7 days ago (rolling window)"""
    now = datetime.now(timezone.utc)
    seven_days_ago = now - timedelta(days=7)
    return seven_days_ago


async def get_last_request_date(user_id: str, collection: str, request_type: str = None) -> dict:
    """
    Get user's last request date and calculate next eligible date
    
    Returns:
    - last_request: datetime of last request
    - next_eligible: datetime when user can request again (last + 24 hours)
    - days_remaining: days until next eligible
    - can_request: True if 24 hours have passed
    
    NOTE: Only counts COMPLETED/APPROVED requests. Failed/pending/processing requests don't start cooldown.
    """
    # Only count COMPLETED/APPROVED requests for cooldown - anything else should NOT block users
    query = {
        "user_id": user_id, 
        "status": {"$in": ["completed", "success", "approved", "paid", "delivered", "COMPLETED", "SUCCESS", "APPROVED", "PAID", "Paid", "DELIVERED"]}
    }
    if request_type:
        query["request_type"] = request_type
    
    last_request = await db[collection].find_one(
        query,
        sort=[("created_at", -1)]
    )
    
    if not last_request:
        return {
            "has_request": False,
            "can_request": True,
            "last_request": None,
            "next_eligible": None,
            "days_remaining": 0
        }
    
    # Parse last request date
    last_date_str = last_request.get("created_at")
    if isinstance(last_date_str, str):
        last_date = datetime.fromisoformat(last_date_str.replace('Z', '+00:00'))
        if last_date.tzinfo is None:
            last_date = last_date.replace(tzinfo=timezone.utc)
    else:
        last_date = last_date_str
    
    # Calculate next eligible date (24 hours from last request)
    next_eligible = last_date + timedelta(hours=24)
    now = datetime.now(timezone.utc)
    
    remaining_seconds = (next_eligible - now).total_seconds()
    can_request = now >= next_eligible
    
    hours_remaining = max(0, int(remaining_seconds // 3600)) if not can_request else 0
    
    return {
        "has_request": True,
        "can_request": can_request,
        "last_request": last_date.isoformat(),
        "next_eligible": next_eligible.isoformat(),
        "hours_remaining": hours_remaining,
        "days_remaining": 0,
        "request_id": last_request.get("request_id")
    }


async def check_loan_emi_this_week(user_id: str) -> dict:
    """Check if user has done loan_emi in last 24 hours"""
    result = await get_last_request_date(user_id, "bill_payment_requests", "loan_emi")
    
    return {
        "has_loan_emi": result["has_request"] and not result["can_request"],
        "loan_emi_request": result.get("request_id"),
        "next_eligible": result.get("next_eligible", ""),
        "days_remaining": result.get("days_remaining", 0),
        # Backward compatibility
        "next_monday": result.get("next_eligible", "")
    }

# Processing fees - EMI style
# <= ₹499: 50% of amount
# > ₹499: Flat ₹10
def get_processing_fee(amount_inr: int) -> int:
    """Calculate processing fee like EMI"""
    if amount_inr <= 499:
        return int(amount_inr * 0.5)  # 50% of amount
    else:
        return 10  # Flat ₹10

# Admin charge percentage - 20%
ADMIN_CHARGE_PERCENT = 20

# Slider allows any amount from 100 to max
MIN_AMOUNT = 100
MAX_AMOUNT = 25000


async def get_user_burn_rate_bank(user_id: str = None) -> dict:
    """
    Get burn rate for bank redeem based on subscription_payment_type.
    - Cash users: 1% burn
    - PRC users: 5% burn
    """
    if not user_id or db is None:
        return {"burn_rate_percent": 1, "payment_type": "unknown"}
    
    user = await db.users.find_one({"uid": user_id}, {"_id": 0, "subscription_payment_type": 1, "subscription_plan": 1})
    if not user:
        return {"burn_rate_percent": 1, "payment_type": "unknown"}
    
    payment_type = user.get("subscription_payment_type", "cash")
    plan = user.get("subscription_plan", "explorer")
    
    if plan in ["elite", "vip", "startup", "growth", "pro"] and payment_type == "prc":
        return {"burn_rate_percent": 5, "payment_type": "prc"}
    else:
        return {"burn_rate_percent": 1, "payment_type": payment_type or "cash"}


def calculate_total_prc(amount_inr: int, burn_rate_percent: int = 1) -> dict:
    """Calculate total PRC needed for withdrawal - fees + admin + burn"""
    if amount_inr < MIN_AMOUNT:
        return None
    
    processing_fee = get_processing_fee(amount_inr)
    admin_charge = int(amount_inr * (ADMIN_CHARGE_PERCENT / 100))
    subtotal_inr = amount_inr + processing_fee + admin_charge
    
    # Burn calculation: burn_rate% of subtotal
    burn_inr = round(subtotal_inr * burn_rate_percent / 100, 2)
    total_inr = subtotal_inr + burn_inr
    
    # PRC rate: Dynamic from database
    prc_rate = get_dynamic_prc_rate()
    
    return {
        "amount_inr": amount_inr,
        "processing_fee_inr": processing_fee,
        "admin_charge_inr": admin_charge,
        "admin_charge_percent": ADMIN_CHARGE_PERCENT,
        "burn_inr": burn_inr,
        "burn_rate_percent": burn_rate_percent,
        "subtotal_inr": subtotal_inr,
        "total_inr": round(total_inr, 2),
        "amount_prc": amount_inr * prc_rate,
        "processing_fee_prc": processing_fee * prc_rate,
        "admin_charge_prc": admin_charge * prc_rate,
        "burn_prc": round(burn_inr * prc_rate, 2),
        "total_prc": int(total_inr * prc_rate),
        "prc_rate": prc_rate
    }


def calculate_charges(amount_inr: int, burn_rate_percent: int = 1) -> dict:
    """Calculate charges breakdown for a given amount including burn"""
    processing_fee = get_processing_fee(amount_inr)
    admin_charge = int(amount_inr * (ADMIN_CHARGE_PERCENT / 100))
    subtotal_inr = amount_inr + processing_fee + admin_charge
    burn_inr = round(subtotal_inr * burn_rate_percent / 100, 2)
    total_inr = subtotal_inr + burn_inr
    
    return {
        "processing_fee_inr": processing_fee,
        "admin_charge_inr": admin_charge,
        "burn_inr": burn_inr,
        "burn_rate_percent": burn_rate_percent,
        "total_inr": round(total_inr, 2)
    }


# ========== BANK DETAILS MANAGEMENT ==========

@router.get("/bank-details/{user_id}")
async def get_bank_details(user_id: str):
    """Get user's saved bank details"""
    user = await db.users.find_one({"uid": user_id}, {"_id": 0, "bank_details": 1})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    bank_details = user.get("bank_details")
    if not bank_details:
        return {"has_bank_details": False, "bank_details": None}
    
    # Mask account number for security (show only last 4 digits)
    masked_account = "XXXX" + bank_details.get("account_number", "")[-4:] if bank_details.get("account_number") else None
    
    return {
        "has_bank_details": True,
        "bank_details": {
            "account_holder_name": bank_details.get("account_holder_name"),
            "account_number_masked": masked_account,
            "ifsc_code": bank_details.get("ifsc_code"),
            "bank_name": bank_details.get("bank_name"),
            "updated_at": bank_details.get("updated_at")
        }
    }


@router.post("/bank-details/{user_id}")
async def save_bank_details(user_id: str, request: Request):
    """Save or update user's bank details"""
    user = await db.users.find_one({"uid": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    data = await request.json()
    
    # Validate required fields
    required_fields = ["account_holder_name", "account_number", "ifsc_code", "bank_name"]
    for field in required_fields:
        if not data.get(field):
            raise HTTPException(status_code=400, detail=f"{field.replace('_', ' ').title()} is required")
    
    # Normalize names for comparison
    account_holder_name = data["account_holder_name"].strip().upper()
    user_profile_name = user.get("name", "").strip().upper()
    
    # Validate: Account Holder Name must match Profile Name
    # Remove extra spaces and compare
    def normalize_name(name):
        return " ".join(name.split())
    
    normalized_holder = normalize_name(account_holder_name)
    normalized_profile = normalize_name(user_profile_name)
    
    if normalized_holder != normalized_profile:
        raise HTTPException(
            status_code=400, 
            detail=f"Account holder name must match your profile name: '{user_profile_name}'"
        )
    
    # Validate IFSC format (11 characters: 4 letters + 0 + 6 alphanumeric)
    ifsc = data["ifsc_code"].upper().strip()
    if len(ifsc) != 11:
        raise HTTPException(status_code=400, detail="IFSC code must be 11 characters")
    
    # IFSC format: First 4 chars = Bank code (letters), 5th char = 0, Last 6 = Branch code
    import re
    ifsc_pattern = r'^[A-Z]{4}0[A-Z0-9]{6}$'
    if not re.match(ifsc_pattern, ifsc):
        raise HTTPException(
            status_code=400, 
            detail="Invalid IFSC format. Must be: 4 letters + 0 + 6 alphanumeric (e.g., SBIN0001234)"
        )
    
    # Validate account number (8-18 digits)
    account_number = data["account_number"].strip()
    if not account_number.isdigit() or len(account_number) < 8 or len(account_number) > 18:
        raise HTTPException(status_code=400, detail="Account number must be 8-18 digits")
    
    bank_details = {
        "account_holder_name": data["account_holder_name"].strip().upper(),
        "account_number": account_number,
        "ifsc_code": ifsc,
        "bank_name": data["bank_name"].strip().upper(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.update_one(
        {"uid": user_id},
        {"$set": {"bank_details": bank_details}}
    )
    
    return {"success": True, "message": "Bank details saved successfully"}


@router.delete("/bank-details/{user_id}")
async def delete_bank_details(user_id: str):
    """Delete user's bank details"""
    user = await db.users.find_one({"uid": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one(
        {"uid": user_id},
        {"$unset": {"bank_details": ""}}
    )
    
    return {"success": True, "message": "Bank details deleted"}


# ========== WITHDRAWAL REQUESTS ==========

@router.get("/bank-redeem/denominations")
async def get_denominations():
    """Get fee info for slider - EMI style fees"""
    # Sample calculations for common amounts
    sample_amounts = [100, 200, 300, 400, 500, 1000, 2000, 5000]
    samples = []
    for amount in sample_amounts:
        calc = calculate_total_prc(amount)
        if calc:
            samples.append(calc)
    
    return {
        "min_amount": MIN_AMOUNT,
        "max_amount": MAX_AMOUNT,
        "fee_structure": {
            "below_500": "50% of amount",
            "above_500": "Flat ₹10"
        },
        "samples": samples,
        "admin_charge_percent": ADMIN_CHARGE_PERCENT,
        "daily_limit": 1,
        "note": "One withdrawal request allowed per 24 hours"
    }


@router.get("/bank-redeem/check-eligibility/{user_id}")
async def check_withdrawal_eligibility(user_id: str):
    """
    Check if user can make a withdrawal request
    
    24-hour cooldown from last request time
    
    Rule: Only 1 of (Bank Redeem / Pay EMI / PRC Vault Redeem) allowed per 24 hours
    """
    user = await db.users.find_one({"uid": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Check if bank details exist
    if not user.get("bank_details"):
        return {
            "eligible": False,
            "reason": "no_bank_details",
            "message": "Please add your bank details first"
        }
    
    # Check KYC status - accept both "verified" and "approved"
    if user.get("kyc_status") not in ["verified", "approved"]:
        return {
            "eligible": False,
            "reason": "kyc_pending",
            "message": "KYC verification required for bank withdrawals"
        }
    
    # Check 1: Loan EMI in last 24 hours
    emi_check = await get_last_request_date(user_id, "bill_payment_requests", "loan_emi")
    if emi_check["has_request"] and not emi_check["can_request"]:
        return {
            "eligible": False,
            "reason": "emi_done_recently",
            "message": f"24-hour limit: Only ONE of Pay EMI / Bank Redeem / PRC Vault Redeem allowed per 24 hours. You did Pay EMI recently. Try again after {emi_check.get('hours_remaining', 0)} hours.",
            "next_eligible_date": emi_check["next_eligible"],
            "hours_remaining": emi_check.get("hours_remaining", 0)
        }
    
    # Check 2: Bank Withdrawal in last 24 hours
    bank_check = await get_last_request_date(user_id, "bank_withdrawal_requests")
    if bank_check["has_request"] and not bank_check["can_request"]:
        return {
            "eligible": False,
            "reason": "bank_redeem_recently",
            "message": f"24-hour limit: Only 1 Bank Redeem allowed per 24 hours. Try again after {bank_check.get('hours_remaining', 0)} hours.",
            "next_eligible_date": bank_check["next_eligible"],
            "hours_remaining": bank_check.get("hours_remaining", 0),
            "existing_request": {
                "request_id": bank_check.get("request_id")
            }
        }
    
    # Check 3: PRC Vault (RD) Redeem in last 24 hours
    rd_check = await get_last_request_date(user_id, "bank_redeem_requests", "rd_redeem")
    if rd_check["has_request"] and not rd_check["can_request"]:
        return {
            "eligible": False,
            "reason": "rd_redeem_recently",
            "message": f"24-hour limit: Only ONE of Pay EMI / Bank Redeem / PRC Vault Redeem allowed per 24 hours. You did PRC Vault Redeem recently. Try again after {rd_check.get('hours_remaining', 0)} hours."
        }
    
    # NOTE: Referral requirement removed as per user request (March 2026)
    
    return {
        "eligible": True,
        "prc_balance": user.get("prc_balance", 0),
        "bank_details": {
            "bank_name": user["bank_details"].get("bank_name"),
            "account_masked": "XXXX" + user["bank_details"].get("account_number", "")[-4:]
        }
    }


@router.post("/bank-redeem/request/{user_id}")
async def create_withdrawal_request(user_id: str, request: Request):
    """Create a new bank withdrawal request"""
    user = await db.users.find_one({"uid": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # ===== CRITICAL: CHECK SUBSCRIPTION FIRST =====
    subscription_plan = (user.get("subscription_plan") or "explorer").lower()
    
    # Free/Explorer users cannot redeem
    if subscription_plan in ["explorer", "free", "", None]:
        raise HTTPException(
            status_code=403,
            detail="Paid subscription required for bank withdrawal. Please upgrade to Startup, Growth or Elite plan."
        )
    
    # Check if subscription is expired
    expiry = user.get("subscription_expiry") or user.get("subscription_expires") or user.get("vip_expiry")
    if expiry:
        try:
            if isinstance(expiry, str):
                expiry_dt = datetime.fromisoformat(expiry.replace('Z', '+00:00'))
            else:
                expiry_dt = expiry
            
            if expiry_dt.tzinfo is None:
                expiry_dt = expiry_dt.replace(tzinfo=timezone.utc)
            
            now = datetime.now(timezone.utc)
            
            if expiry_dt < now:
                days_expired = (now - expiry_dt).days
                raise HTTPException(
                    status_code=403,
                    detail=f"Your subscription expired {days_expired} days ago. Please renew to use bank withdrawal."
                )
        except HTTPException:
            raise
        except Exception as e:
            logging.warning(f"[BANK-REDEEM] Expiry parse error for {user_id}: {e}")
    # ================================================
    
    data = await request.json()
    amount_inr = data.get("amount_inr")
    
    # ===== CHECK IF BANK REDEEM SERVICE IS ENABLED =====
    settings = await db.settings.find_one({}, {"_id": 0, "service_toggles": 1})
    toggles = settings.get("service_toggles", {}) if settings else {}
    if not toggles.get("bank_redeem", True):
        raise HTTPException(
            status_code=503, 
            detail="Service temporarily down. Please try again later."
        )
    # ===================================================
    
    # Validate amount - slider allows any amount from MIN to MAX
    if not amount_inr or amount_inr < MIN_AMOUNT:
        raise HTTPException(
            status_code=400, 
            detail=f"Minimum withdrawal amount is ₹{MIN_AMOUNT}"
        )
    
    if amount_inr > MAX_AMOUNT:
        raise HTTPException(
            status_code=400, 
            detail=f"Maximum withdrawal amount is ₹{MAX_AMOUNT}"
        )
    
    # Check bank details
    if not user.get("bank_details"):
        raise HTTPException(status_code=400, detail="Please add your bank details first")
    
    # Check KYC - accept both "verified" and "approved"
    if user.get("kyc_status") not in ["verified", "approved"]:
        raise HTTPException(status_code=400, detail="KYC verification required for withdrawals")
    
    # Check 1: Loan EMI in last 24 hours
    emi_check = await get_last_request_date(user_id, "bill_payment_requests", "loan_emi")
    if emi_check["has_request"] and not emi_check["can_request"]:
        raise HTTPException(
            status_code=429, 
            detail=f"24-hour limit: Only ONE of Pay EMI / Bank Redeem / PRC Vault allowed per 24 hours. You did Pay EMI recently. Try again after {emi_check.get('hours_remaining', 0)} hours."
        )
    
    # Check 2: Bank Withdrawal in last 24 hours
    bank_check = await get_last_request_date(user_id, "bank_withdrawal_requests")
    if bank_check["has_request"] and not bank_check["can_request"]:
        raise HTTPException(
            status_code=429, 
            detail=f"24-hour limit: Only 1 Bank Redeem allowed per 24 hours. Try again after {bank_check.get('hours_remaining', 0)} hours."
        )
    
    # Check 3: PRC Vault (RD) Redeem in last 24 hours
    rd_check = await get_last_request_date(user_id, "bank_redeem_requests", "rd_redeem")
    if rd_check["has_request"] and not rd_check["can_request"]:
        raise HTTPException(
            status_code=429, 
            detail=f"24-hour limit: Only ONE of Pay EMI / Bank Redeem / PRC Vault allowed per 24 hours. You did PRC Vault Redeem recently. Try again after {rd_check.get('hours_remaining', 0)} hours."
        )
    
    # NOTE: Referral requirement removed as per user request (March 2026)
    
    # Get user's burn rate based on subscription payment type
    burn_info = await get_user_burn_rate_bank(user_id)
    burn_rate = burn_info["burn_rate_percent"]
    
    # Calculate charges with burn rate
    charges = calculate_total_prc(amount_inr, burn_rate_percent=burn_rate)
    if not charges:
        raise HTTPException(status_code=400, detail="Invalid amount")
    total_prc = charges["total_prc"]
    
    # Check PRC balance
    prc_balance = user.get("prc_balance", 0)
    if prc_balance < total_prc:
        raise HTTPException(
            status_code=400, 
            detail=f"Insufficient PRC balance. Required: {total_prc} PRC, Available: {prc_balance} PRC"
        )
    
    # User can redeem up to 100% of their balance
    # (Balance check already done above - no additional limit needed)
    
    # Create request
    now = datetime.now(timezone.utc)
    request_id = f"BWR{now.strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"
    
    withdrawal_request = {
        "request_id": request_id,
        "user_id": user_id,
        "amount_inr": amount_inr,
        "processing_fee_inr": charges["processing_fee_inr"],
        "admin_charge_inr": charges["admin_charge_inr"],
        "burn_inr": charges.get("burn_inr", 0),
        "burn_rate_percent": charges.get("burn_rate_percent", 1),
        "total_inr": charges["total_inr"],
        "total_prc_deducted": total_prc,
        "bank_details": user["bank_details"].copy(),
        "status": "pending",
        "created_at": now.isoformat(),
        "user_name": user.get("name", ""),
        "user_email": user.get("email", ""),
        "user_mobile": user.get("mobile", "")
    }
    
    # Deduct PRC from user balance
    await db.users.update_one(
        {"uid": user_id},
        {"$inc": {"prc_balance": -total_prc}}
    )
    
    # Create the request
    await db.bank_withdrawal_requests.insert_one(withdrawal_request)
    
    # Log transaction
    transaction = {
        "transaction_id": f"TXN{now.strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:4]}",
        "user_id": user_id,
        "type": "bank_withdrawal_request",
        "amount": -total_prc,
        "description": f"Bank withdrawal request - ₹{amount_inr}",
        "reference_id": request_id,
        "created_at": now.isoformat()
    }
    await db.transactions.insert_one(transaction)
    
    # Create notification
    if create_notification:
        try:
            await create_notification(
                user_id=user_id,
                title="Bank Withdrawal Request",
                message=f"Your withdrawal request of ₹{amount_inr} has been submitted. {total_prc} PRC deducted.",
                notification_type="withdrawal"
            )
        except Exception as e:
            logging.error(f"Notification error: {e}")
    
    return {
        "success": True,
        "message": "Withdrawal request submitted successfully",
        "request_id": request_id,
        "amount_inr": amount_inr,
        "total_prc_deducted": total_prc,
        "charges": charges
    }


@router.get("/bank-redeem/history/{user_id}")
async def get_withdrawal_history(user_id: str, page: int = 1, limit: int = 10):
    """Get user's withdrawal request history"""
    skip = (page - 1) * limit
    
    requests = await db.bank_withdrawal_requests.find(
        {"user_id": user_id},
        {"_id": 0, "bank_details.account_number": 0}  # Don't expose full account number
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    total = await db.bank_withdrawal_requests.count_documents({"user_id": user_id})
    
    # Mask account numbers in response
    for req in requests:
        if req.get("bank_details"):
            req["bank_details"]["account_masked"] = "XXXX" + req["bank_details"].get("account_number", "")[-4:] if req["bank_details"].get("account_number") else "XXXX"
            req["bank_details"].pop("account_number", None)
    
    return {
        "requests": requests,
        "total": total,
        "page": page,
        "total_pages": (total + limit - 1) // limit
    }



# ========== USER EDIT PENDING REQUEST ==========

@router.put("/bank-redeem/request/{user_id}/{request_id}")
async def edit_pending_request(user_id: str, request_id: str, request: Request):
    """
    Allow user to edit their pending bank redeem request
    - Only pending requests can be edited
    - Can change: amount, bank details
    - PRC will be adjusted accordingly
    """
    # Get user
    user = await db.users.find_one({"uid": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get the existing request
    existing_request = await db.bank_withdrawal_requests.find_one({
        "request_id": request_id,
        "user_id": user_id
    })
    
    if not existing_request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Only pending requests can be edited
    if existing_request.get("status") != "pending":
        raise HTTPException(
            status_code=400, 
            detail=f"Only pending requests can be edited. Current status: {existing_request.get('status')}"
        )
    
    data = await request.json()
    new_amount_inr = data.get("amount_inr")
    new_bank_details = data.get("bank_details")
    
    if not new_amount_inr and not new_bank_details:
        raise HTTPException(status_code=400, detail="No changes provided")
    
    old_amount_inr = existing_request.get("amount_inr", 0)
    old_total_prc = existing_request.get("total_prc_deducted", 0)
    
    updates = {
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    prc_adjustment = 0
    
    # Handle amount change
    if new_amount_inr and new_amount_inr != old_amount_inr:
        # Validate new amount
        if new_amount_inr < MIN_AMOUNT or new_amount_inr > MAX_AMOUNT:
            raise HTTPException(
                status_code=400, 
                detail=f"Amount must be between ₹{MIN_AMOUNT} and ₹{MAX_AMOUNT}"
            )
        
        # Get user's burn rate for charge recalculation
        burn_info = await get_user_burn_rate_bank(user_id)
        burn_rate = burn_info["burn_rate_percent"]
        
        # Calculate new charges with burn
        new_charges = calculate_charges(new_amount_inr, burn_rate_percent=burn_rate)
        new_total_prc = calculate_total_prc(new_amount_inr, burn_rate_percent=burn_rate)
        
        if not new_total_prc:
            raise HTTPException(status_code=400, detail="Invalid amount")
        
        new_total_prc = new_total_prc["total_prc"]
        prc_adjustment = old_total_prc - new_total_prc  # Positive = refund, Negative = deduct more
        
        # Check if user has enough balance for additional deduction
        if prc_adjustment < 0:
            current_balance = user.get("prc_balance", 0)
            if current_balance < abs(prc_adjustment):
                raise HTTPException(
                    status_code=400, 
                    detail=f"Insufficient PRC balance. Need {abs(prc_adjustment)} more PRC."
                )
        
        updates.update({
            "amount_inr": new_amount_inr,
            "processing_fee_inr": new_charges["processing_fee_inr"],
            "admin_charge_inr": new_charges["admin_charge_inr"],
            "burn_inr": new_charges.get("burn_inr", 0),
            "burn_rate_percent": new_charges.get("burn_rate_percent", 1),
            "total_inr": new_charges["total_inr"],
            "total_prc_deducted": new_total_prc
        })
    
    # Handle bank details change
    if new_bank_details:
        # Validate bank details
        if new_bank_details.get("account_number"):
            acc_num = new_bank_details["account_number"].replace(" ", "")
            if len(acc_num) < 9 or len(acc_num) > 18:
                raise HTTPException(status_code=400, detail="Account number must be 9-18 digits")
            new_bank_details["account_number"] = acc_num
        
        if new_bank_details.get("ifsc_code"):
            ifsc = new_bank_details["ifsc_code"].upper()
            import re
            if not re.match(r'^[A-Z]{4}0[A-Z0-9]{6}$', ifsc):
                raise HTTPException(status_code=400, detail="Invalid IFSC format")
            new_bank_details["ifsc_code"] = ifsc
        
        # Merge with existing bank details
        current_bank = existing_request.get("bank_details", {})
        current_bank.update(new_bank_details)
        updates["bank_details"] = current_bank
    
    # Apply updates
    await db.bank_withdrawal_requests.update_one(
        {"request_id": request_id},
        {"$set": updates}
    )
    
    # Adjust PRC balance if amount changed
    if prc_adjustment != 0:
        await db.users.update_one(
            {"uid": user_id},
            {"$inc": {"prc_balance": prc_adjustment}}
        )
        
        # Log the adjustment transaction
        now = datetime.now(timezone.utc)
        adjustment_type = "prc_refund" if prc_adjustment > 0 else "prc_deduction"
        await db.transactions.insert_one({
            "transaction_id": f"TXN{now.strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:4]}",
            "user_id": user_id,
            "type": adjustment_type,
            "amount": prc_adjustment,
            "description": f"Bank redeem request edit: ₹{old_amount_inr} → ₹{new_amount_inr}",
            "reference_id": request_id,
            "created_at": now.isoformat()
        })
    
    return {
        "success": True,
        "message": "Request updated successfully",
        "request_id": request_id,
        "changes": {
            "amount_changed": new_amount_inr != old_amount_inr if new_amount_inr else False,
            "bank_details_changed": new_bank_details is not None,
            "prc_adjustment": prc_adjustment
        }
    }


@router.delete("/bank-redeem/request/{user_id}/{request_id}")
async def cancel_pending_request(user_id: str, request_id: str):
    """
    Allow user to cancel their pending bank redeem request
    - Full PRC refund on cancellation
    """
    # Get user
    user = await db.users.find_one({"uid": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get the existing request
    existing_request = await db.bank_withdrawal_requests.find_one({
        "request_id": request_id,
        "user_id": user_id
    })
    
    if not existing_request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Only pending requests can be cancelled
    if existing_request.get("status") != "pending":
        raise HTTPException(
            status_code=400, 
            detail=f"Only pending requests can be cancelled. Current status: {existing_request.get('status')}"
        )
    
    prc_to_refund = existing_request.get("total_prc_deducted", 0)
    
    # Update request status
    now = datetime.now(timezone.utc)
    await db.bank_withdrawal_requests.update_one(
        {"request_id": request_id},
        {"$set": {
            "status": "cancelled",
            "cancelled_at": now.isoformat(),
            "cancelled_by": "user"
        }}
    )
    
    # Refund PRC
    if prc_to_refund > 0:
        await db.users.update_one(
            {"uid": user_id},
            {"$inc": {"prc_balance": prc_to_refund}}
        )
        
        # Log refund transaction
        await db.transactions.insert_one({
            "transaction_id": f"TXN{now.strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:4]}",
            "user_id": user_id,
            "type": "prc_refund",
            "amount": prc_to_refund,
            "description": f"Bank redeem request cancelled - ₹{existing_request.get('amount_inr', 0)}",
            "reference_id": request_id,
            "created_at": now.isoformat()
        })
    
    return {
        "success": True,
        "message": "Request cancelled successfully",
        "request_id": request_id,
        "prc_refunded": prc_to_refund
    }



# ========== ADMIN ENDPOINTS ==========

@router.get("/admin/bank-redeem/requests")
async def get_admin_withdrawal_requests(
    status: str = None,
    page: int = 1,
    limit: int = 50,
    search: str = None,
    date_from: str = None,
    date_to: str = None,
    sort_order: str = "desc"
):
    """Get all bank withdrawal requests for admin with search and date filters
    Note: Fetches from BOTH bank_withdrawal_requests AND bank_redeem_requests collections
    to handle legacy data migration issues
    """
    query = {}
    if status:
        query["status"] = status
    
    # Date range filtering
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = f"{date_from}T00:00:00+00:00"
        if date_to:
            date_query["$lte"] = f"{date_to}T23:59:59+00:00"
        if date_query:
            query["created_at"] = date_query
    
    # If search provided, find matching user IDs first
    user_ids_to_search = []
    if search:
        search_users = await db.users.find({
            "$or": [
                {"name": {"$regex": search, "$options": "i"}},
                {"email": {"$regex": search, "$options": "i"}},
                {"phone": {"$regex": search, "$options": "i"}},
                {"mobile": {"$regex": search, "$options": "i"}}
            ]
        }, {"_id": 0, "uid": 1}).to_list(100)
        user_ids_to_search = [u["uid"] for u in search_users]
        
        # Search by user_id or request_id or bank name or mobile
        search_query = [
            {"request_id": {"$regex": search, "$options": "i"}},
            {"bank_details.bank_name": {"$regex": search, "$options": "i"}},
            {"user_mobile": {"$regex": search, "$options": "i"}},
            {"mobile": {"$regex": search, "$options": "i"}}
        ]
        if user_ids_to_search:
            search_query.append({"user_id": {"$in": user_ids_to_search}})
        query["$or"] = search_query
    
    skip = (page - 1) * limit
    sort_direction = 1 if sort_order == "asc" else -1
    
    # Fetch from BOTH collections and merge
    requests_from_withdrawal = await db.bank_withdrawal_requests.find(
        query, {"_id": 0}
    ).sort("created_at", sort_direction).to_list(2000)
    
    # Also fetch from bank_redeem_requests (legacy collection)
    requests_from_redeem = await db.bank_redeem_requests.find(
        query, {"_id": 0}
    ).sort("created_at", sort_direction).to_list(2000)
    
    # Merge and deduplicate by request_id
    all_requests = []
    seen_ids = set()
    
    for req in requests_from_withdrawal + requests_from_redeem:
        req_id = req.get("request_id") or req.get("_id")
        if req_id and req_id not in seen_ids:
            seen_ids.add(req_id)
            all_requests.append(req)
    
    # Sort merged results
    all_requests.sort(
        key=lambda x: x.get("created_at", ""), 
        reverse=(sort_direction == -1)
    )
    
    # Apply pagination
    total = len(all_requests)
    paginated = all_requests[skip:skip + limit]
    
    # Get combined stats
    all_for_stats = requests_from_withdrawal + requests_from_redeem
    stats_dict = {}
    for req in all_for_stats:
        s = req.get("status", "unknown")
        if s not in stats_dict:
            stats_dict[s] = {"count": 0, "total": 0}
        stats_dict[s]["count"] += 1
        stats_dict[s]["total"] += req.get("amount_inr", 0)
    
    return {
        "requests": paginated,
        "total": total,
        "page": page,
        "stats": stats_dict
    }


@router.post("/admin/bank-redeem/{request_id}/approve")
async def approve_withdrawal(request_id: str, request: Request):
    """Approve a bank withdrawal request and initiate Eko DMT transfer"""
    data = await request.json()
    admin_id = data.get("admin_id")
    transaction_ref = data.get("transaction_ref", "")
    admin_notes = data.get("admin_notes", "")
    use_eko_transfer = data.get("use_eko_transfer", True)  # Default to Eko transfer
    
    withdrawal = await db.bank_withdrawal_requests.find_one({"request_id": request_id})
    if not withdrawal:
        raise HTTPException(status_code=404, detail="Request not found")
    
    if withdrawal.get("status") != "pending":
        raise HTTPException(status_code=400, detail=f"Request already {withdrawal.get('status')}")
    
    now = datetime.now(timezone.utc)
    
    # Get admin details for tracking
    admin_user = await db.users.find_one({"uid": admin_id}, {"_id": 0, "name": 1, "email": 1})
    admin_name = admin_user.get("name", "Admin") if admin_user else "Admin"
    
    # Get user's bank details
    bank_account = withdrawal.get("bank_account_number")
    bank_ifsc = withdrawal.get("bank_ifsc")
    bank_holder = withdrawal.get("bank_account_holder_name", withdrawal.get("user_name", ""))
    user_mobile = withdrawal.get("user_mobile", "9999999999")
    amount_inr = withdrawal.get("amount_inr", 0)
    
    eko_transfer_result = None
    eko_txn_id = None
    transfer_status = "manual"
    
    # Try Eko DMT transfer if enabled and configured
    if use_eko_transfer and bank_account and bank_ifsc and amount_inr > 0:
        # Check Eko balance first
        balance_check = await eko_check_balance()
        
        if balance_check.get("success") and balance_check.get("balance", 0) >= amount_inr:
            # Initiate Eko DMT transfer
            client_ref = f"PARAS-{request_id[-8:]}"
            
            eko_transfer_result = await eko_initiate_transfer(
                recipient_mobile=user_mobile,
                account_number=bank_account,
                ifsc=bank_ifsc,
                amount=amount_inr,
                recipient_name=bank_holder,
                client_ref_id=client_ref
            )
            
            if eko_transfer_result.get("success"):
                eko_txn_id = eko_transfer_result.get("txn_id")
                transaction_ref = eko_txn_id or client_ref
                transfer_status = "eko_dmt"
                logging.info(f"Eko DMT transfer successful: {eko_txn_id} for {request_id}")
            else:
                # Eko transfer failed - SET eko_failed status (NOT approved)
                # This keeps the request visible for admin to manually complete
                eko_error = eko_transfer_result.get('error', 'Unknown error')
                logging.warning(f"Eko DMT transfer failed: {eko_error} for {request_id}")
                transfer_status = "eko_failed"
                
                # Parse error for user-friendly reason
                eko_fail_reason = "Eko API Error"
                error_str = str(eko_error).lower()
                if "403" in error_str or "forbidden" in error_str:
                    eko_fail_reason = "Eko API: Server IP not whitelisted (403)"
                elif "401" in error_str or "unauthorized" in error_str:
                    eko_fail_reason = "Eko API: Authentication failed (401)"
                elif "400" in error_str:
                    eko_fail_reason = "Eko API: Invalid request parameters (400)"
                elif "500" in error_str:
                    eko_fail_reason = "Eko API: Server error (500)"
                elif "timeout" in error_str or "timed out" in error_str:
                    eko_fail_reason = "Eko API: Request timeout"
                elif "connection" in error_str:
                    eko_fail_reason = "Eko API: Connection error"
                elif "insufficient" in error_str or "balance" in error_str:
                    eko_fail_reason = "Eko API: Insufficient wallet balance"
                else:
                    eko_fail_reason = f"Eko API Error: {str(eko_error)[:100]}"
                
                admin_notes = f"Eko transfer failed: {eko_fail_reason}. {admin_notes}"
                
                # Update to eko_failed status and return early
                await db.bank_withdrawal_requests.update_one(
                    {"request_id": request_id},
                    {"$set": {
                        "status": "eko_failed",
                        "eko_fail_reason": eko_fail_reason,
                        "eko_error": str(eko_error),
                        "transfer_status": "eko_failed",
                        "admin_notes": admin_notes,
                        "last_eko_attempt": now.isoformat(),
                        "processed_by": admin_name,
                        "processed_by_uid": admin_id
                    }}
                )
                
                return {
                    "success": False,
                    "message": "⚠️ Eko API failed. Request pending for manual action.",
                    "status": "eko_failed",
                    "eko_fail_reason": eko_fail_reason,
                    "admin_options": ["manual_complete", "retry", "reject"],
                    "next_steps": "Admin can: 1) Complete Manually with UTR, 2) Retry Eko, 3) Reject with PRC refund"
                }
        else:
            # Insufficient Eko balance - SET eko_failed status
            balance_available = balance_check.get('balance', 0)
            logging.warning(f"Insufficient Eko balance for transfer. Required: {amount_inr}, Available: {balance_available}")
            transfer_status = "eko_failed"
            eko_fail_reason = f"Eko wallet balance insufficient. Required: ₹{amount_inr}, Available: ₹{balance_available}"
            admin_notes = f"{eko_fail_reason}. {admin_notes}"
            
            # Update to eko_failed status and return early
            await db.bank_withdrawal_requests.update_one(
                {"request_id": request_id},
                {"$set": {
                    "status": "eko_failed",
                    "eko_fail_reason": eko_fail_reason,
                    "transfer_status": "eko_failed",
                    "admin_notes": admin_notes,
                    "last_eko_attempt": now.isoformat(),
                    "processed_by": admin_name,
                    "processed_by_uid": admin_id
                }}
            )
            
            return {
                "success": False,
                "message": "⚠️ Insufficient Eko balance. Request pending for manual action.",
                "status": "eko_failed",
                "eko_fail_reason": eko_fail_reason,
                "admin_options": ["manual_complete", "reject"],
                "next_steps": "Admin can: 1) Complete Manually with UTR, 2) Reject with PRC refund"
            }
    
    # If we reach here, Eko transfer was successful
    # Update withdrawal request with approved status
    await db.bank_withdrawal_requests.update_one(
        {"request_id": request_id},
        {"$set": {
            "status": "approved",
            "approved_at": now.isoformat(),
            "approved_by": admin_id,
            "approved_by_name": admin_name,
            "processed_by": admin_name,
            "processed_by_uid": admin_id,
            "processed_at": now.isoformat(),
            "transaction_ref": transaction_ref,
            "admin_notes": admin_notes,
            "transfer_status": transfer_status,
            "manually_approved": False,
            "eko_txn_id": eko_txn_id,
            "eko_response": eko_transfer_result
        }}
    )
    
    # Notify user
    if create_notification:
        try:
            message = f"Your bank withdrawal of ₹{withdrawal['amount_inr']:,.0f} has been approved"
            if transfer_status == "eko_dmt":
                message += " and transferred to your bank account instantly!"
            else:
                message += ". Transfer will be processed within 24-48 hours."
            
            await create_notification(
                user_id=withdrawal["user_id"],
                title="Payment Approved ✓",
                message=message,
                notification_type="payment_approved",
                data={"amount_inr": withdrawal['amount_inr'], "payment_type": "bank_withdrawal", "transfer_status": transfer_status}
            )
        except Exception:
            pass
    
    return {
        "success": True, 
        "message": "Withdrawal approved and transferred via Eko DMT!",
        "transfer_status": transfer_status,
        "eko_txn_id": eko_txn_id,
        "eko_transfer_success": True,
        "manually_approved": False
    }


@router.post("/admin/bank-redeem/{request_id}/manual-complete")
async def manual_complete_withdrawal(request_id: str, request: Request):
    """
    Manually complete a bank withdrawal request without Eko API.
    Use when Eko is down or admin has processed payment offline.
    
    Required: txn_reference (UTR/Reference number)
    """
    data = await request.json()
    admin_id = data.get("admin_id") or data.get("admin_uid")
    txn_reference = data.get("txn_reference", "")
    admin_notes = data.get("admin_notes", "Manually completed by admin")
    
    if not txn_reference:
        raise HTTPException(status_code=400, detail="UTR/Reference number is required for manual completion")
    
    withdrawal = await db.bank_withdrawal_requests.find_one({"request_id": request_id})
    if not withdrawal:
        raise HTTPException(status_code=404, detail="Request not found")
    
    current_status = withdrawal.get("status")
    if current_status not in ["pending", "approved", "eko_failed", "processing"]:
        raise HTTPException(status_code=400, detail=f"Cannot manually complete request with status: {current_status}")
    
    now = datetime.now(timezone.utc)
    
    # Get admin details
    admin_user = await db.users.find_one({"uid": admin_id}, {"_id": 0, "name": 1})
    admin_name = admin_user.get("name", "Admin") if admin_user else "Admin"
    
    # Calculate processing time
    processing_time_str = None
    created_at = withdrawal.get("created_at")
    if created_at:
        try:
            if isinstance(created_at, str):
                created_dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            else:
                created_dt = created_at
            if created_dt.tzinfo is None:
                created_dt = created_dt.replace(tzinfo=timezone.utc)
            time_diff = now - created_dt
            hours = int(time_diff.total_seconds() // 3600)
            minutes = int((time_diff.total_seconds() % 3600) // 60)
            if hours > 24:
                days = hours // 24
                processing_time_str = f"{days}d {hours % 24}h {minutes}m"
            elif hours > 0:
                processing_time_str = f"{hours}h {minutes}m"
            else:
                processing_time_str = f"{minutes}m"
        except Exception:
            pass
    
    # Update request as manually completed
    await db.bank_withdrawal_requests.update_one(
        {"request_id": request_id},
        {"$set": {
            "status": "completed",
            "completed_at": now.isoformat(),
            "processed_at": now.isoformat(),
            "processed_by": admin_name,
            "processed_by_uid": admin_id,
            "processing_time": processing_time_str,
            "transaction_ref": txn_reference,
            "transfer_status": "manual",
            "manually_approved": True,
            "manual_txn_reference": txn_reference,
            "admin_notes": admin_notes
        }}
    )
    
    # Notify user
    if create_notification:
        try:
            await create_notification(
                user_id=withdrawal["user_id"],
                title="✅ Bank Withdrawal Completed",
                message=f"Your withdrawal of ₹{withdrawal['amount_inr']:,.0f} has been transferred to your bank account.\n🧾 Reference: {txn_reference}",
                notification_type="payment_completed",
                data={
                    "amount_inr": withdrawal['amount_inr'],
                    "payment_type": "bank_withdrawal",
                    "txn_reference": txn_reference
                }
            )
        except Exception:
            pass
    
    return {
        "success": True,
        "message": "✅ Withdrawal manually completed",
        "status": "completed",
        "txn_reference": txn_reference,
        "processing_time": processing_time_str,
        "manually_approved": True
    }


@router.post("/admin/bank-redeem/{request_id}/reject")
async def reject_withdrawal(request_id: str, request: Request):
    """Reject a bank withdrawal request and refund PRC"""
    data = await request.json()
    admin_id = data.get("admin_id")
    reason = data.get("reason", "Rejected by admin")
    
    withdrawal = await db.bank_withdrawal_requests.find_one({"request_id": request_id})
    if not withdrawal:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Allow rejection for pending and eko_failed statuses
    current_status = withdrawal.get("status")
    if current_status not in ["pending", "eko_failed"]:
        raise HTTPException(status_code=400, detail=f"Cannot reject request with status: {current_status}")
    
    now = datetime.now(timezone.utc)
    user_id = withdrawal["user_id"]
    refund_amount = withdrawal.get("total_prc_deducted", 0)
    
    # Get admin details for tracking
    admin_user = await db.users.find_one({"uid": admin_id}, {"_id": 0, "name": 1, "email": 1})
    admin_name = admin_user.get("name", "Admin") if admin_user else "Admin"
    
    # Refund PRC
    await db.users.update_one(
        {"uid": user_id},
        {"$inc": {"prc_balance": refund_amount}}
    )
    
    # Log refund transaction
    await db.transactions.insert_one({
        "transaction_id": f"REF{now.strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:4]}",
        "user_id": user_id,
        "type": "bank_withdrawal_refund",
        "amount": refund_amount,
        "description": f"Bank withdrawal rejected - ₹{withdrawal['amount_inr']} refunded",
        "reference_id": request_id,
        "created_at": now.isoformat()
    })
    
    # Update request status
    await db.bank_withdrawal_requests.update_one(
        {"request_id": request_id},
        {"$set": {
            "status": "rejected",
            "rejected_at": now.isoformat(),
            "rejected_by": admin_id,
            "rejected_by_name": admin_name,
            "processed_by": admin_name,
            "processed_by_uid": admin_id,
            "processed_at": now.isoformat(),
            "rejection_reason": reason
        }}
    )
    
    # Notify user
    if create_notification:
        try:
            await create_notification(
                user_id=user_id,
                title="Payment Rejected",
                message=f"Your withdrawal request of ₹{withdrawal['amount_inr']:,.0f} was rejected. {refund_amount:,.0f} PRC refunded. Reason: {reason}",
                notification_type="payment_rejected",
                data={"amount_inr": withdrawal['amount_inr'], "reason": reason, "payment_type": "bank_withdrawal"}
            )
        except Exception:
            pass
    
    return {"success": True, "message": "Withdrawal rejected and PRC refunded"}


@router.get("/admin/bank-redeem/pending-count")
async def get_pending_count():
    """Get count of pending bank withdrawal requests"""
    count = await db.bank_withdrawal_requests.count_documents({"status": "pending"})
    return {"pending_count": count}



@router.get("/admin/bank-redeem/eko-balance")
async def get_eko_balance_for_admin():
    """Get Eko settlement account balance for admin dashboard"""
    result = await eko_check_balance()
    
    # Check if service is configured
    service = await get_eko_service()
    is_configured = service is not None
    
    return {
        "balance": result.get("balance", 0),
        "currency": "INR",
        "configured": is_configured,
        "error": result.get("error")
    }


@router.post("/admin/bank-redeem/verify-account")
async def verify_bank_account_admin(request: Request):
    """Verify bank account before transfer"""
    data = await request.json()
    ifsc = data.get("ifsc")
    account_number = data.get("account_number")
    
    if not ifsc or not account_number:
        raise HTTPException(status_code=400, detail="IFSC and account number required")
    
    result = await eko_verify_bank_account(ifsc, account_number)
    return result



@router.post("/admin/eko/dmt-transfer")
async def admin_direct_dmt_transfer(request: Request):
    """
    Admin Direct DMT Transfer via Eko API
    No approval required - instant processing
    """
    data = await request.json()
    recipient_mobile = data.get("recipient_mobile")
    account_number = data.get("account_number")
    ifsc = data.get("ifsc")
    amount = data.get("amount")
    recipient_name = data.get("recipient_name")
    admin_id = data.get("admin_id")
    
    if not all([recipient_mobile, account_number, ifsc, amount, recipient_name]):
        raise HTTPException(status_code=400, detail="All fields are required")
    
    if float(amount) < 100:
        raise HTTPException(status_code=400, detail="Minimum transfer amount is ₹100")
    
    try:
        # Check Eko balance first
        balance_result = await eko_check_balance()
        if not balance_result.get("success"):
            raise HTTPException(status_code=400, detail="Failed to check Eko balance")
        
        current_balance = float(balance_result.get("balance", 0))
        if current_balance < float(amount):
            raise HTTPException(
                status_code=400, 
                detail=f"Insufficient Eko balance. Available: ₹{current_balance}, Required: ₹{amount}"
            )
        
        # Create transaction reference
        client_ref = f"ADMT{datetime.now().strftime('%Y%m%d%H%M%S')}{account_number[-4:]}"
        
        # Initiate transfer
        result = await eko_initiate_transfer(
            recipient_mobile=recipient_mobile,
            account_number=account_number,
            ifsc=ifsc,
            amount=float(amount),
            recipient_name=recipient_name,
            client_ref_id=client_ref
        )
        
        # Log transaction
        if db is not None:
            await db.admin_eko_transactions.insert_one({
                "type": "admin_dmt_transfer",
                "recipient_mobile": recipient_mobile,
                "account_number": account_number[-4:],  # Store only last 4 digits
                "ifsc": ifsc,
                "amount": float(amount),
                "recipient_name": recipient_name,
                "admin_id": admin_id,
                "client_ref": client_ref,
                "eko_response": result,
                "success": result.get("success", False),
                "timestamp": datetime.now(timezone.utc).isoformat()
            })
        
        if result.get("success"):
            return {
                "success": True,
                "message": "Transfer successful!",
                "txn_id": result.get("txn_id"),
                "eko_txn_id": result.get("txn_id"),
                "client_ref": client_ref,
                "amount": amount
            }
        else:
            return {
                "success": False,
                "message": result.get("error") or "Transfer failed",
                "error": result.get("error"),
                "client_ref": client_ref
            }
            
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Admin DMT transfer failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))



@router.get("/admin/bank-redeem/export-pending-excel")
async def export_pending_bank_redeem_excel():
    """
    Download pending bank redeem requests as an Excel file.
    Columns: Sr No, Name, Account Number, IFSC, Amount, Active/Inactive, Date of Request.
    """
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    # Pull pending from both collections (new + legacy)
    pending_new = await db.bank_withdrawal_requests.find(
        {"status": "pending"}, {"_id": 0}
    ).sort("created_at", 1).to_list(5000)
    pending_legacy = await db.bank_transfer_requests.find(
        {"status": "pending"}, {"_id": 0}
    ).sort("created_at", 1).to_list(5000)

    # Merge + dedupe by request_id
    merged = []
    seen = set()
    for r in pending_new + pending_legacy:
        rid = r.get("request_id") or ""
        if rid and rid in seen:
            continue
        if rid:
            seen.add(rid)
        merged.append(r)
    # Oldest-first (so admin pays earliest requests first)
    merged.sort(key=lambda x: x.get("created_at", ""))

    # Batch-resolve user details (name, status)
    user_ids = list({r.get("user_id") for r in merged if r.get("user_id")})
    uid_to_user = {}
    if user_ids:
        async for u in db.users.find(
            {"uid": {"$in": user_ids}},
            {"_id": 0, "uid": 1, "name": 1, "is_banned": 1, "is_blocked": 1, "is_active": 1},
        ):
            uid_to_user[u["uid"]] = u

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pending Bank Redeems"

    # Title row
    title = f"Pending Bank Redeem Requests — {datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')}"
    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="4F46E5")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Header row
    headers = ["Sr. No", "Name", "Account Number", "IFSC Code", "Amount (₹)", "Active / Inactive", "Date of Request"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6366F1")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[2].height = 22

    # Data rows
    inactive_fill = PatternFill("solid", fgColor="FEE2E2")
    active_fill = PatternFill("solid", fgColor="DCFCE7")
    row = 3
    total_amount = 0.0

    for idx, r in enumerate(merged, start=1):
        user = uid_to_user.get(r.get("user_id") or "", {})
        bank = r.get("bank_details") or {}

        # Name priority: request snapshot > user profile > "—"
        name = (r.get("user_name") or user.get("name") or bank.get("account_holder_name") or "—")
        account_number = bank.get("account_number") or "—"
        ifsc = (bank.get("ifsc_code") or "").upper() or "—"
        amount = float(r.get("amount_inr") or r.get("amount") or 0)
        total_amount += amount

        # Active/Inactive from user flags (is_banned / is_blocked / is_active)
        if user:
            is_banned = bool(user.get("is_banned") or user.get("is_blocked"))
            is_active_flag = user.get("is_active", True)
            active_status = "Inactive" if (is_banned or not is_active_flag) else "Active"
        else:
            active_status = "Unknown"

        # Date of Request (ISO → dd-mm-yyyy hh:mm)
        created = r.get("created_at", "")
        try:
            dt = datetime.fromisoformat(str(created).replace("Z", "+00:00"))
            # Convert to IST (UTC+5:30)
            ist = dt.astimezone(timezone(timedelta(hours=5, minutes=30)))
            date_str = ist.strftime("%d-%m-%Y %H:%M IST")
        except Exception:
            date_str = str(created)[:19]

        row_data = [idx, name, str(account_number), ifsc, amount, active_status, date_str]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal=("right" if col_idx in (1, 5) else "left"), vertical="center")
            if col_idx == 5:
                cell.number_format = '₹ #,##0.00'
            if col_idx == 6:
                if active_status == "Inactive":
                    cell.fill = inactive_fill
                    cell.font = Font(color="991B1B", bold=True)
                elif active_status == "Active":
                    cell.fill = active_fill
                    cell.font = Font(color="166534", bold=True)
        row += 1

    # Totals row
    if merged:
        totals_fill = PatternFill("solid", fgColor="FEF3C7")
        ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=row, column=4, value="Sum").alignment = Alignment(horizontal="right")
        total_cell = ws.cell(row=row, column=5, value=total_amount)
        total_cell.number_format = '₹ #,##0.00'
        total_cell.font = Font(bold=True)
        for c in range(1, 8):
            cell = ws.cell(row=row, column=c)
            cell.border = border
            cell.fill = totals_fill
    else:
        ws.cell(row=row, column=1, value="No pending requests").font = Font(italic=True, color="64748B")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)

    # Column widths
    widths = [8, 26, 22, 15, 14, 18, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A3"

    # Save to buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"pending-bank-redeems-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Total-Pending": str(len(merged)),
            "X-Total-Amount": f"{total_amount:.2f}",
        },
    )
