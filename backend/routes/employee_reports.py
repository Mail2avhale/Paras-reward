"""
EMPLOYEE MANAGEMENT - REPORTS MODULE
=====================================
Phase A (Core):
  1. Salary Register (monthly Excel)
  2. Salary Slip PDF (individual)
  3. Attendance Report (monthly Excel)
  4. Pool Distribution Report (date-wise Excel)

Phase B (Statutory):
  5. PF Return (ECR Excel)
  6. ESI Return (Excel)
  7. TDS Report (Excel)
  8. Form 16 PDF

Phase C (Analytics):
  9. Leave Balance Report
  10. HR Analytics Dashboard
  11. YTD Earnings
"""
import io
import calendar
import logging
from datetime import datetime, timezone, timedelta
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

router = APIRouter(prefix="/employees/reports", tags=["Employee Reports"])

db = None


def set_db(database):
    global db
    db = database


# ==================== COMPANY / CONSTANTS ====================
COMPANY_NAME = "PARAS REWARD TECHNOLOGIES PRIVATE LIMITED"
COMPANY_ADDRESS = "B-18, Bizz Tower, Chatrapati Sambhaji Nagar, Maharashtra - 431001"
COMPANY_WEBSITE = "www.parasreward.com"
COMPANY_EMAIL = "info@parasreward.com"
COMPANY_CIN = ""  # Add if available
COMPANY_PF_CODE = ""
COMPANY_ESI_CODE = ""


# ==================== STYLE HELPERS ====================
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True, color="334155")
CELL_FONT = Font(name="Calibri", size=10)
BORDER_THIN = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0")
)


def apply_header_row(ws, row_idx, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
    ws.row_dimensions[row_idx].height = 30


def autosize_columns(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        # Skip merged cells (first cell may be MergedCell)
        first = col_cells[0]
        if not hasattr(first, 'column_letter'):
            # Find first real cell with column_letter
            col_letter = None
            for cell in col_cells:
                if hasattr(cell, 'column_letter'):
                    col_letter = cell.column_letter
                    break
            if not col_letter:
                continue
        else:
            col_letter = first.column_letter
        length = 10
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, length + 2))


def write_company_header(ws, row, merge_cols, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_cols)
    c = ws.cell(row=row, column=1, value=COMPANY_NAME)
    c.font = Font(name="Calibri", size=14, bold=True, color="0F172A")
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=merge_cols)
    c = ws.cell(row=row + 1, column=1, value=COMPANY_ADDRESS)
    c.font = Font(name="Calibri", size=9, color="64748B")
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2, end_column=merge_cols)
    c = ws.cell(row=row + 2, column=1, value=title)
    c.font = Font(name="Calibri", size=12, bold=True, color="1E293B")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row + 2].height = 24
    return row + 3  # next free row


# ==================== SALARY BREAKDOWN (Indian compliance) ====================
def calculate_salary_breakdown(monthly_salary: float) -> dict:
    """Same logic as employee_management.py for consistency."""
    basic = round(monthly_salary * 0.50, 2)
    hra = round(basic * 0.40, 2)
    conveyance = 1600
    medical = 1250
    special = round(monthly_salary - basic - hra - conveyance - medical, 2)
    if special < 0:
        special = 0
    total_earnings = round(basic + hra + conveyance + medical + special, 2)
    return {
        "earnings": {
            "basic_salary": basic,
            "hra": hra,
            "conveyance": conveyance,
            "medical": medical,
            "special_allowance": special,
            "total_earnings": total_earnings
        }
    }


def calculate_deductions(gross: float, basic: float, full_monthly_for_tds: float = None) -> dict:
    pf_basic = min(basic, 15000)
    pf_employee = round(pf_basic * 0.12, 2)
    esi_employee = round(gross * 0.0075, 2) if gross <= 21000 else 0
    professional_tax = 200 if gross > 10000 else 0

    # Use FULL monthly salary for TDS projection when available (not ratio-adjusted)
    # This prevents TDS under-projection on months with LOP
    tds_basis = full_monthly_for_tds if full_monthly_for_tds is not None else gross
    annual = tds_basis * 12
    if annual <= 300000:
        tds = 0
    elif annual <= 600000:
        tds = round((annual - 300000) * 0.05 / 12, 2)
    elif annual <= 900000:
        tds = round(((annual - 600000) * 0.10 + 300000 * 0.05) / 12, 2)
    else:
        tds = round(((annual - 900000) * 0.15 + 300000 * 0.10 + 300000 * 0.05) / 12, 2)

    pf_employer = round(pf_basic * 0.12, 2)
    esi_employer = round(gross * 0.0325, 2) if gross <= 21000 else 0
    gratuity = round(basic * 15 / 26 / 12, 2)

    return {
        "pf_employee": pf_employee,
        "esi_employee": esi_employee,
        "professional_tax": professional_tax,
        "tds": tds,
        "total": round(pf_employee + esi_employee + professional_tax + tds, 2),
        "pf_employer": pf_employer,
        "esi_employer": esi_employer,
        "gratuity": gratuity
    }


async def get_month_attendance_summary(employee_id: str, month: int, year: int) -> dict:
    days_in_month = calendar.monthrange(year, month)[1]
    start = f"{year}-{month:02d}-01"
    end = f"{year}-{month:02d}-{days_in_month:02d}"
    attendance = await db.employee_attendance.find(
        {"employee_id": employee_id, "date": {"$gte": start, "$lte": end}},
        {"_id": 0}
    ).to_list(31)
    present = sum(1 for a in attendance if a.get("status") == "present")
    half_day = sum(1 for a in attendance if a.get("status") == "half_day")
    holiday = sum(1 for a in attendance if a.get("status") == "holiday")
    leave = sum(1 for a in attendance if a.get("status") == "leave")
    absent = sum(1 for a in attendance if a.get("status") == "absent")
    total_working_days = 26
    effective_days = present + (half_day * 0.5) + holiday + leave
    attendance_ratio = min(1, effective_days / total_working_days) if total_working_days > 0 else 1
    return {
        "total_working_days": total_working_days,
        "present": present, "half_day": half_day, "absent": absent,
        "leave": leave, "holiday": holiday,
        "effective_days": effective_days,
        "attendance_ratio": round(attendance_ratio, 4)
    }


# ==================== 1. SALARY REGISTER ====================
@router.get("/salary-register")
async def salary_register(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2100),
    fmt: str = Query("excel", pattern="^(excel|csv)$")
):
    """Monthly salary register — all active employees in one sheet."""
    try:
        employees = await db.employees.find({"status": "active"}, {"_id": 0}).to_list(10000)
        if not employees:
            raise HTTPException(status_code=404, detail="No active employees")

        month_name = calendar.month_name[month]

        rows = []
        totals = {"gross": 0, "pf_e": 0, "esi_e": 0, "pt": 0, "tds": 0, "lop": 0, "net": 0,
                  "pf_emp": 0, "esi_emp": 0, "grat": 0}

        for emp in employees:
            att = await get_month_attendance_summary(emp["employee_id"], month, year)
            monthly = emp.get("monthly_salary", 0)
            bd = calculate_salary_breakdown(monthly)

            # Adjust for attendance
            adj_earn = {k: round(v * att["attendance_ratio"], 2) for k, v in bd["earnings"].items() if k != "total_earnings"}
            adj_earn["total_earnings"] = round(sum(adj_earn.values()), 2)

            basic = adj_earn["basic_salary"]
            gross = adj_earn["total_earnings"]
            lop = round(monthly - (monthly * att["attendance_ratio"]), 2) if att["attendance_ratio"] < 1 else 0
            ded = calculate_deductions(gross, basic)

            total_ded = ded["total"] + lop
            net = round(gross - total_ded, 2)

            rows.append({
                "emp_id": emp["employee_id"],
                "name": emp.get("name", ""),
                "dept": emp.get("department", ""),
                "designation": emp.get("designation", ""),
                "doj": emp.get("joining_date", ""),
                "days_paid": att["effective_days"],
                "gross_ctc": monthly,
                "basic": basic,
                "hra": adj_earn["hra"],
                "conv": adj_earn["conveyance"],
                "med": adj_earn["medical"],
                "special": adj_earn["special_allowance"],
                "gross": gross,
                "pf_emp": ded["pf_employee"],
                "esi_emp": ded["esi_employee"],
                "pt": ded["professional_tax"],
                "tds": ded["tds"],
                "lop": lop,
                "total_ded": total_ded,
                "net": net,
                "pf_empl": ded["pf_employer"],
                "esi_empl": ded["esi_employer"],
                "grat": ded["gratuity"],
                "ctc": round(gross + ded["pf_employer"] + ded["esi_employer"] + ded["gratuity"], 2)
            })

            totals["gross"] += gross
            totals["pf_e"] += ded["pf_employee"]
            totals["esi_e"] += ded["esi_employee"]
            totals["pt"] += ded["professional_tax"]
            totals["tds"] += ded["tds"]
            totals["lop"] += lop
            totals["net"] += net
            totals["pf_emp"] += ded["pf_employer"]
            totals["esi_emp"] += ded["esi_employer"]
            totals["grat"] += ded["gratuity"]

        # Build Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Salary Register"

        header_row = write_company_header(ws, 1, 24, f"SALARY REGISTER — {month_name} {year}")

        headers = [
            "Emp ID", "Name", "Department", "Designation", "DOJ",
            "Days Paid", "CTC (Monthly)", "Basic", "HRA", "Conveyance", "Medical", "Special",
            "Gross", "PF (E)", "ESI (E)", "PT", "TDS", "LOP", "Total Ded", "Net Salary",
            "PF (Empr)", "ESI (Empr)", "Gratuity", "CTC Total"
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=header_row + 1, column=col, value=h)
        apply_header_row(ws, header_row + 1, len(headers))

        data_start = header_row + 2
        for i, r in enumerate(rows):
            vals = [
                r["emp_id"], r["name"], r["dept"], r["designation"], r["doj"],
                r["days_paid"], r["gross_ctc"], r["basic"], r["hra"], r["conv"], r["med"], r["special"],
                r["gross"], r["pf_emp"], r["esi_emp"], r["pt"], r["tds"], r["lop"], r["total_ded"], r["net"],
                r["pf_empl"], r["esi_empl"], r["grat"], r["ctc"]
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=data_start + i, column=col, value=v)
                c.font = CELL_FONT
                c.border = BORDER_THIN
                if col >= 6:
                    c.alignment = Alignment(horizontal="right")

        # Totals row
        total_row = data_start + len(rows)
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
        for col, key in [(13, "gross"), (14, "pf_e"), (15, "esi_e"), (16, "pt"), (17, "tds"),
                          (18, "lop"), (20, "net"), (21, "pf_emp"), (22, "esi_emp"), (23, "grat")]:
            c = ws.cell(row=total_row, column=col, value=round(totals[key], 2))
            c.font = Font(bold=True, color="0F172A")
            c.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        autosize_columns(ws)

        # Return stream
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"Salary_Register_{month_name}_{year}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"[REPORTS] Salary register error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 2. SALARY SLIP PDF ====================
@router.get("/salary-slip-pdf/{employee_id}")
async def salary_slip_pdf(
    employee_id: str,
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2100)
):
    """PDF salary slip for an employee for given month."""
    try:
        emp = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0})
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")

        att = await get_month_attendance_summary(employee_id, month, year)
        monthly = emp.get("monthly_salary", 0)
        bd = calculate_salary_breakdown(monthly)
        adj_earn = {k: round(v * att["attendance_ratio"], 2) for k, v in bd["earnings"].items() if k != "total_earnings"}
        adj_earn["total_earnings"] = round(sum(adj_earn.values()), 2)

        basic = adj_earn["basic_salary"]
        gross = adj_earn["total_earnings"]
        lop = round(monthly - (monthly * att["attendance_ratio"]), 2) if att["attendance_ratio"] < 1 else 0
        ded = calculate_deductions(gross, basic)
        total_ded = round(ded["total"] + lop, 2)
        net = round(gross - total_ded, 2)

        # Build PDF
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm,
                                 topMargin=12 * mm, bottomMargin=12 * mm, title=f"Salary Slip {month}/{year}")
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=14, textColor=colors.HexColor("#0F172A"), alignment=1)
        subtitle_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor("#64748B"), alignment=1)
        section_style = ParagraphStyle('sec', parent=styles['Heading3'], fontSize=11, textColor=colors.HexColor("#1E293B"), alignment=1, spaceBefore=4, spaceAfter=4)

        story.append(Paragraph(COMPANY_NAME, title_style))
        story.append(Paragraph(COMPANY_ADDRESS, subtitle_style))
        story.append(Paragraph(f"Website: {COMPANY_WEBSITE}  |  Email: {COMPANY_EMAIL}", subtitle_style))
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"Payslip for {calendar.month_name[month]} {year}", section_style))
        story.append(Spacer(1, 4))

        # Employee info table
        info_data = [
            ["Employee ID", emp.get("employee_id", ""), "Employee Name", emp.get("name", "")],
            ["Designation", emp.get("designation", ""), "Department", emp.get("department", "")],
            ["Date of Joining", str(emp.get("joining_date", ""))[:10], "PAN", emp.get("pan", "—")],
            ["Bank A/c", emp.get("bank_account", "—"), "IFSC", emp.get("bank_ifsc", "—")],
            ["Days in Month", calendar.monthrange(year, month)[1],
             "Days Paid", f"{att['effective_days']:.1f} / {att['total_working_days']}"],
        ]
        info_table = Table(info_data, colWidths=[38 * mm, 52 * mm, 38 * mm, 52 * mm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor("#F1F5F9")),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 10))

        # Earnings + Deductions side by side
        earn_data = [
            ["EARNINGS", "Amount (INR)"],
            ["Basic Salary", f"{adj_earn['basic_salary']:,.2f}"],
            ["HRA", f"{adj_earn['hra']:,.2f}"],
            ["Conveyance", f"{adj_earn['conveyance']:,.2f}"],
            ["Medical", f"{adj_earn['medical']:,.2f}"],
            ["Special Allowance", f"{adj_earn['special_allowance']:,.2f}"],
            ["Gross Earnings", f"{gross:,.2f}"],
        ]
        ded_data = [
            ["DEDUCTIONS", "Amount (INR)"],
            ["PF (Employee)", f"{ded['pf_employee']:,.2f}"],
            ["ESI (Employee)", f"{ded['esi_employee']:,.2f}"],
            ["Professional Tax", f"{ded['professional_tax']:,.2f}"],
            ["TDS", f"{ded['tds']:,.2f}"],
            ["Loss of Pay", f"{lop:,.2f}"],
            ["Total Deductions", f"{total_ded:,.2f}"],
        ]

        combined_data = []
        for i in range(max(len(earn_data), len(ded_data))):
            e = earn_data[i] if i < len(earn_data) else ["", ""]
            d = ded_data[i] if i < len(ded_data) else ["", ""]
            combined_data.append(e + d)

        combined = Table(combined_data, colWidths=[50 * mm, 40 * mm, 50 * mm, 40 * mm])
        combined.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (1, 0), colors.HexColor("#DBEAFE")),
            ('BACKGROUND', (2, 0), (3, 0), colors.HexColor("#FEE2E2")),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(combined)
        story.append(Spacer(1, 10))

        # Net pay
        net_table = Table([["NET SALARY", f"INR {net:,.2f}"]], colWidths=[120 * mm, 60 * mm])
        net_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#10B981")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(net_table)
        story.append(Spacer(1, 6))

        words_style = ParagraphStyle('words', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor("#475569"))
        story.append(Paragraph(f"<i>In Words: {number_to_words_inr(net)}</i>", words_style))
        story.append(Spacer(1, 14))

        # Employer contributions + Attendance breakdown side by side
        emp_contrib = [
            ["EMPLOYER CONTRIBUTIONS", "INR"],
            ["PF (Employer)", f"{ded['pf_employer']:,.2f}"],
            ["ESI (Employer)", f"{ded['esi_employer']:,.2f}"],
            ["Gratuity Accrual", f"{ded['gratuity']:,.2f}"],
        ]
        att_breakdown = [
            ["ATTENDANCE BREAKDOWN", ""],
            ["Present", f"{att['present']}"],
            ["Half-day", f"{att['half_day']}"],
            ["Leave / Holiday", f"{att['leave'] + att['holiday']}"],
            ["Absent (LOP)", f"{att['absent']}"],
        ]
        max_len = max(len(emp_contrib), len(att_breakdown))
        cdata = []
        for i in range(max_len):
            e = emp_contrib[i] if i < len(emp_contrib) else ["", ""]
            a = att_breakdown[i] if i < len(att_breakdown) else ["", ""]
            cdata.append(e + a)

        bottom_tbl = Table(cdata, colWidths=[55 * mm, 35 * mm, 55 * mm, 35 * mm])
        bottom_tbl.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (1, 0), colors.HexColor("#FEF3C7")),
            ('BACKGROUND', (2, 0), (3, 0), colors.HexColor("#F1F5F9")),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(bottom_tbl)
        story.append(Spacer(1, 12))

        footer_style = ParagraphStyle('foot', parent=styles['Normal'], fontSize=7.5,
                                       textColor=colors.HexColor("#64748B"), alignment=1)
        story.append(Paragraph(
            "This is a computer-generated payslip. No signature required.",
            footer_style
        ))
        story.append(Paragraph(
            f"Generated on {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}",
            footer_style
        ))

        doc.build(story)
        buf.seek(0)
        filename = f"Payslip_{emp.get('name','employee').replace(' ','_')}_{calendar.month_name[month]}_{year}.pdf"
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"[REPORTS] Salary slip PDF error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 3. ATTENDANCE REPORT ====================
@router.get("/attendance")
async def attendance_report(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2100)
):
    """Monthly attendance matrix Excel: rows=employees, cols=days."""
    try:
        employees = await db.employees.find({"status": "active"}, {"_id": 0}).to_list(10000)
        if not employees:
            raise HTTPException(status_code=404, detail="No active employees")

        days_in_month = calendar.monthrange(year, month)[1]
        month_name = calendar.month_name[month]

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        header_row = write_company_header(ws, 1, days_in_month + 8, f"ATTENDANCE SHEET — {month_name} {year}")

        headers = ["Emp ID", "Name", "Department"] + [f"{d:02d}" for d in range(1, days_in_month + 1)] + \
                  ["Present", "Absent", "Half", "Leave", "%"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=header_row + 1, column=col, value=h)
        apply_header_row(ws, header_row + 1, len(headers))

        STATUS_MAP = {
            "present": ("P", "D1FAE5"),
            "absent": ("A", "FEE2E2"),
            "half_day": ("H", "FEF3C7"),
            "leave": ("L", "DBEAFE"),
            "holiday": ("O", "E0E7FF"),
        }

        for i, emp in enumerate(employees):
            row = header_row + 2 + i
            ws.cell(row=row, column=1, value=emp.get("employee_id", ""))
            ws.cell(row=row, column=2, value=emp.get("name", ""))
            ws.cell(row=row, column=3, value=emp.get("department", ""))

            # Fetch attendance for the month once
            start = f"{year}-{month:02d}-01"
            end = f"{year}-{month:02d}-{days_in_month:02d}"
            recs = await db.employee_attendance.find(
                {"employee_id": emp["employee_id"], "date": {"$gte": start, "$lte": end}},
                {"_id": 0, "date": 1, "status": 1}
            ).to_list(31)
            day_map = {r["date"]: r["status"] for r in recs}

            counts = {"present": 0, "absent": 0, "half_day": 0, "leave": 0, "holiday": 0}
            for d in range(1, days_in_month + 1):
                date_str = f"{year}-{month:02d}-{d:02d}"
                status = day_map.get(date_str, "")
                label, color = STATUS_MAP.get(status, ("", "FFFFFF"))
                cell = ws.cell(row=row, column=3 + d, value=label)
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(name="Calibri", size=9, bold=bool(label))
                if color and color != "FFFFFF":
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                if status in counts:
                    counts[status] += 1

            # Summary cols
            base_col = 3 + days_in_month + 1
            eff = counts["present"] + counts["half_day"] * 0.5
            ws.cell(row=row, column=base_col, value=counts["present"])
            ws.cell(row=row, column=base_col + 1, value=counts["absent"])
            ws.cell(row=row, column=base_col + 2, value=counts["half_day"])
            ws.cell(row=row, column=base_col + 3, value=counts["leave"])
            ws.cell(row=row, column=base_col + 4, value=f"{round(eff / 26 * 100, 1)}%" if eff > 0 else "0%")

        # Legend
        legend_row = header_row + 3 + len(employees)
        ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
        legends = [("P = Present", "D1FAE5"), ("A = Absent", "FEE2E2"), ("H = Half-day", "FEF3C7"), ("L = Leave", "DBEAFE"), ("O = Holiday", "E0E7FF")]
        for j, (text, col) in enumerate(legends):
            cell = ws.cell(row=legend_row, column=2 + j, value=text)
            cell.fill = PatternFill(start_color=col, end_color=col, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Narrow day columns
        for d in range(1, days_in_month + 1):
            ws.column_dimensions[get_column_letter(3 + d)].width = 4.5
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 16

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"Attendance_{month_name}_{year}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"[REPORTS] Attendance error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 4. POOL DISTRIBUTION REPORT ====================
@router.get("/pool-distribution")
async def pool_distribution_report(
    from_date: str = Query(..., description="YYYY-MM-DD"),
    to_date: str = Query(..., description="YYYY-MM-DD")
):
    """PRC pool earnings per employee across date range (Excel)."""
    try:
        # All active employees
        employees = await db.employees.find({"status": "active"}, {"_id": 0}).to_list(10000)
        emp_map = {e["employee_id"]: e for e in employees}

        # Pool transactions in range (type=distribution)
        txns = await db.employee_pool_transactions.find(
            {"type": "distribution", "timestamp": {"$gte": from_date, "$lte": to_date + "T23:59:59"}},
            {"_id": 0}
        ).to_list(100000)

        # Aggregate per employee
        per_emp = {}
        for t in txns:
            eid = t.get("employee_id") or t.get("employee_pool_employee_id")
            if not eid:
                continue
            per_emp.setdefault(eid, {"total": 0, "count": 0, "last": ""})
            per_emp[eid]["total"] += float(t.get("amount", 0))
            per_emp[eid]["count"] += 1
            ts = t.get("timestamp", "")
            if ts > per_emp[eid]["last"]:
                per_emp[eid]["last"] = ts

        # Build Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Pool Distribution"

        header_row = write_company_header(ws, 1, 7, f"EMPLOYEE POOL DISTRIBUTION — {from_date} to {to_date}")
        headers = ["Emp ID", "Name", "Department", "Monthly Salary (INR)", "Total PRC Earned", "Distribution Count", "Last Distributed"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=header_row + 1, column=col, value=h)
        apply_header_row(ws, header_row + 1, len(headers))

        row_idx = header_row + 2
        total_prc = 0
        for eid, emp in emp_map.items():
            stats = per_emp.get(eid, {"total": 0, "count": 0, "last": "—"})
            if stats["count"] == 0:
                continue
            total_prc += stats["total"]
            vals = [
                eid, emp.get("name", ""), emp.get("department", ""),
                emp.get("monthly_salary", 0),
                round(stats["total"], 4),
                stats["count"],
                stats["last"][:10] if stats["last"] != "—" else "—"
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row_idx, column=col, value=v)
                c.border = BORDER_THIN
                c.font = CELL_FONT
                if col in (4, 5, 6):
                    c.alignment = Alignment(horizontal="right")
            row_idx += 1

        if row_idx == header_row + 2:
            ws.cell(row=row_idx, column=1, value="No distributions in this date range.")
            row_idx += 1

        # Totals
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        ws.cell(row=row_idx, column=1, value="TOTAL PRC DISTRIBUTED").font = Font(bold=True)
        total_cell = ws.cell(row=row_idx, column=5, value=round(total_prc, 4))
        total_cell.font = Font(bold=True)
        total_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        autosize_columns(ws)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"Pool_Distribution_{from_date}_to_{to_date}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"[REPORTS] Pool distribution error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 5. ANALYTICS DASHBOARD (JSON) ====================
@router.get("/analytics")
async def hr_analytics():
    """HR analytics dashboard: headcount, avg salary, pool stats, department distribution."""
    try:
        # Headcount
        total_active = await db.employees.count_documents({"status": "active"})
        total_resigned = await db.employees.count_documents({"status": "resigned"})
        total_all = await db.employees.count_documents({})

        # Department distribution
        dept_agg = await db.employees.aggregate([
            {"$match": {"status": "active"}},
            {"$group": {"_id": "$department", "count": {"$sum": 1}, "total_salary": {"$sum": "$monthly_salary"}}},
            {"$sort": {"count": -1}}
        ]).to_list(50)
        departments = [{"department": d["_id"] or "Unassigned", "count": d["count"], "total_salary": d["total_salary"]} for d in dept_agg]

        # Salary stats
        salary_agg = await db.employees.aggregate([
            {"$match": {"status": "active"}},
            {"$group": {"_id": None, "total_monthly": {"$sum": "$monthly_salary"}, "avg": {"$avg": "$monthly_salary"},
                         "min": {"$min": "$monthly_salary"}, "max": {"$max": "$monthly_salary"}}}
        ]).to_list(1)
        salary_stats = salary_agg[0] if salary_agg else {"total_monthly": 0, "avg": 0, "min": 0, "max": 0}

        # Pool stats
        pool_settings = await db.employee_pool_settings.find_one({}, {"_id": 0}) or {}
        this_month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()

        pool_agg = await db.employee_pool_transactions.aggregate([
            {"$match": {"type": "distribution", "timestamp": {"$gte": this_month_start}}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}, "count": {"$sum": 1}}}
        ]).to_list(1)
        pool_this_month = pool_agg[0] if pool_agg else {"total": 0, "count": 0}

        # Top earners (this month)
        top_earners = await db.employee_pool_transactions.aggregate([
            {"$match": {"type": "distribution", "timestamp": {"$gte": this_month_start}}},
            {"$group": {"_id": "$employee_id", "total_prc": {"$sum": "$amount"}}},
            {"$sort": {"total_prc": -1}},
            {"$limit": 5}
        ]).to_list(5)
        # Enrich with name
        enriched_top = []
        for t in top_earners:
            emp = await db.employees.find_one({"employee_id": t["_id"]}, {"_id": 0, "name": 1, "department": 1})
            enriched_top.append({
                "employee_id": t["_id"],
                "name": emp.get("name", "Unknown") if emp else "Unknown",
                "department": emp.get("department", "") if emp else "",
                "total_prc": round(t["total_prc"], 4)
            })

        # Leave usage (this year)
        year_start = f"{datetime.now(timezone.utc).year}-01-01"
        leave_agg = await db.employee_leaves.aggregate([
            {"$match": {"status": "approved", "from_date": {"$gte": year_start}}},
            {"$group": {"_id": "$leave_type", "count": {"$sum": 1}, "days": {"$sum": "$days"}}}
        ]).to_list(10)
        leave_usage = [{"type": l["_id"], "count": l["count"], "days": l.get("days", 0)} for l in leave_agg]

        # Attendance this month
        month_start = f"{datetime.now(timezone.utc).year}-{datetime.now(timezone.utc).month:02d}-01"
        att_agg = await db.employee_attendance.aggregate([
            {"$match": {"date": {"$gte": month_start}}},
            {"$group": {"_id": "$status", "count": {"$sum": 1}}}
        ]).to_list(10)
        attendance_summary = {a["_id"]: a["count"] for a in att_agg}

        return {
            "success": True,
            "headcount": {
                "active": total_active,
                "resigned": total_resigned,
                "total": total_all,
                "attrition_rate": round((total_resigned / max(total_all, 1)) * 100, 2)
            },
            "departments": departments,
            "salary": {
                "total_monthly_cost": round(salary_stats.get("total_monthly", 0), 2),
                "avg_salary": round(salary_stats.get("avg", 0) or 0, 2),
                "min_salary": round(salary_stats.get("min", 0) or 0, 2),
                "max_salary": round(salary_stats.get("max", 0) or 0, 2)
            },
            "pool": {
                "current_balance": round(pool_settings.get("pool_balance", 0), 4),
                "this_month_distributed": round(pool_this_month.get("total", 0), 4),
                "this_month_distribution_count": pool_this_month.get("count", 0),
                "prc_to_inr_rate": pool_settings.get("prc_to_inr_rate", 0.10)
            },
            "top_earners_this_month": enriched_top,
            "leave_usage_this_year": leave_usage,
            "attendance_this_month": attendance_summary
        }
    except Exception as e:
        logging.error(f"[REPORTS] Analytics error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 6. PF MONTHLY ECR (Phase B) ====================
@router.get("/pf-ecr")
async def pf_monthly_ecr(month: int = Query(..., ge=1, le=12), year: int = Query(..., ge=2020, le=2100)):
    """PF Electronic Challan Return (ECR) — monthly Excel for EPFO upload."""
    try:
        employees = await db.employees.find({"status": "active"}, {"_id": 0}).to_list(10000)
        if not employees:
            raise HTTPException(status_code=404, detail="No active employees")

        month_name = calendar.month_name[month]
        wb = Workbook()
        ws = wb.active
        ws.title = "PF ECR"

        header_row = write_company_header(ws, 1, 11, f"PF MONTHLY ECR — {month_name} {year}")

        headers = [
            "UAN", "Member Name", "Gross Wages", "EPF Wages", "EPS Wages", "EDLI Wages",
            "EPF Contribution (12%)", "EPS Contribution (8.33%)", "EPF Employer Share",
            "NCP Days", "Refund of Advances"
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=header_row + 1, column=col, value=h)
        apply_header_row(ws, header_row + 1, len(headers))

        totals = {"gross": 0, "epf_wages": 0, "eps_wages": 0, "edli": 0,
                   "epf_emp": 0, "eps": 0, "epf_empr": 0}
        for i, emp in enumerate(employees):
            att = await get_month_attendance_summary(emp["employee_id"], month, year)
            monthly = emp.get("monthly_salary", 0)
            bd = calculate_salary_breakdown(monthly)
            adj_basic = round(bd["earnings"]["basic_salary"] * att["attendance_ratio"], 2)
            adj_gross = round(monthly * att["attendance_ratio"], 2)

            epf_wages = min(adj_basic, 15000)
            eps_wages = min(adj_basic, 15000)
            edli_wages = min(adj_basic, 15000)
            epf_employee = round(epf_wages * 0.12, 2)
            eps_contrib = round(eps_wages * 0.0833, 2)
            epf_employer = round(epf_wages * 0.12 - eps_contrib, 2)
            ncp_days = max(0, att["total_working_days"] - att["effective_days"])

            vals = [
                emp.get("pf_uan", ""),
                emp.get("name", ""),
                adj_gross,
                epf_wages, eps_wages, edli_wages,
                epf_employee, eps_contrib, epf_employer,
                int(ncp_days), 0
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=header_row + 2 + i, column=col, value=v)
                c.border = BORDER_THIN
                c.font = CELL_FONT
                if col >= 3:
                    c.alignment = Alignment(horizontal="right")

            totals["gross"] += adj_gross
            totals["epf_wages"] += epf_wages
            totals["eps_wages"] += eps_wages
            totals["edli"] += edli_wages
            totals["epf_emp"] += epf_employee
            totals["eps"] += eps_contrib
            totals["epf_empr"] += epf_employer

        total_row = header_row + 2 + len(employees)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        for col, key in [(3, "gross"), (4, "epf_wages"), (5, "eps_wages"), (6, "edli"),
                          (7, "epf_emp"), (8, "eps"), (9, "epf_empr")]:
            c = ws.cell(row=total_row, column=col, value=round(totals[key], 2))
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        autosize_columns(ws)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="PF_ECR_{month_name}_{year}.xlsx"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"[REPORTS] PF ECR error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 7. ESI MONTHLY RETURN (Phase B) ====================
@router.get("/esi-return")
async def esi_monthly_return(month: int = Query(..., ge=1, le=12), year: int = Query(..., ge=2020, le=2100)):
    """ESI Monthly Return Excel (covered employees only — wages <= 21000)."""
    try:
        employees = await db.employees.find({"status": "active"}, {"_id": 0}).to_list(10000)
        if not employees:
            raise HTTPException(status_code=404, detail="No active employees")

        month_name = calendar.month_name[month]
        wb = Workbook()
        ws = wb.active
        ws.title = "ESI Return"

        header_row = write_company_header(ws, 1, 8, f"ESI MONTHLY RETURN — {month_name} {year}")
        headers = [
            "ESI IP Number", "Name", "Working Days", "Total Wages",
            "ESI Wages", "Employee Contrib (0.75%)", "Employer Contrib (3.25%)", "Total ESI"
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=header_row + 1, column=col, value=h)
        apply_header_row(ws, header_row + 1, len(headers))

        totals = {"wages": 0, "emp": 0, "empr": 0}
        row_idx = header_row + 2
        covered_count = 0
        for emp in employees:
            att = await get_month_attendance_summary(emp["employee_id"], month, year)
            monthly = emp.get("monthly_salary", 0)
            adj_gross = round(monthly * att["attendance_ratio"], 2)
            if adj_gross > 21000:
                continue  # not ESI-covered
            covered_count += 1
            esi_employee = round(adj_gross * 0.0075, 2)
            esi_employer = round(adj_gross * 0.0325, 2)

            vals = [
                emp.get("esi_ip", ""),
                emp.get("name", ""),
                int(att["effective_days"]),
                adj_gross, adj_gross,
                esi_employee, esi_employer,
                round(esi_employee + esi_employer, 2)
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row_idx, column=col, value=v)
                c.border = BORDER_THIN
                c.font = CELL_FONT
                if col >= 3:
                    c.alignment = Alignment(horizontal="right")
            totals["wages"] += adj_gross
            totals["emp"] += esi_employee
            totals["empr"] += esi_employer
            row_idx += 1

        if covered_count == 0:
            ws.cell(row=row_idx, column=1, value="No employees under ESI wage ceiling (<= INR 21000)").font = Font(italic=True, color="64748B")
        else:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
            ws.cell(row=row_idx, column=1, value=f"TOTAL ({covered_count} covered)").font = Font(bold=True)
            for col, key in [(4, "wages"), (6, "emp"), (7, "empr")]:
                c = ws.cell(row=row_idx, column=col, value=round(totals[key], 2))
                c.font = Font(bold=True)
                c.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            ws.cell(row=row_idx, column=8, value=round(totals["emp"] + totals["empr"], 2)).font = Font(bold=True)

        autosize_columns(ws)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="ESI_Return_{month_name}_{year}.xlsx"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"[REPORTS] ESI return error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 8. TDS REPORT (Phase B) ====================
@router.get("/tds")
async def tds_report(fy_start_year: int = Query(..., description="Financial year start (e.g., 2025 for FY25-26)")):
    """Financial Year TDS report — all employees summed Apr-Mar."""
    try:
        employees = await db.employees.find({"status": "active"}, {"_id": 0}).to_list(10000)
        if not employees:
            raise HTTPException(status_code=404, detail="No active employees")

        months = [(fy_start_year, m) for m in range(4, 13)] + [(fy_start_year + 1, m) for m in range(1, 4)]

        wb = Workbook()
        ws = wb.active
        ws.title = "TDS Report"
        header_row = write_company_header(ws, 1, 8, f"TDS REPORT — FY {fy_start_year}-{str(fy_start_year + 1)[-2:]}")

        headers = ["Emp ID", "Name", "PAN", "Annual Gross", "Annual Basic",
                    "Total TDS (FY)", "Avg Monthly TDS", "Months Processed"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=header_row + 1, column=col, value=h)
        apply_header_row(ws, header_row + 1, len(headers))

        row_idx = header_row + 2
        grand_tds = 0
        grand_gross = 0
        for emp in employees:
            monthly_tds_total = 0
            gross_total = 0
            basic_total = 0
            months_processed = 0
            for yr, m in months:
                att = await get_month_attendance_summary(emp["employee_id"], m, yr)
                if att["effective_days"] == 0:
                    continue
                monthly = emp.get("monthly_salary", 0)
                bd = calculate_salary_breakdown(monthly)
                adj_basic = round(bd["earnings"]["basic_salary"] * att["attendance_ratio"], 2)
                adj_gross = round(monthly * att["attendance_ratio"], 2)
                ded = calculate_deductions(adj_gross, adj_basic, full_monthly_for_tds=monthly)
                monthly_tds_total += ded["tds"]
                gross_total += adj_gross
                basic_total += adj_basic
                months_processed += 1

            vals = [
                emp.get("employee_id", ""),
                emp.get("name", ""),
                emp.get("pan", "—"),
                round(gross_total, 2),
                round(basic_total, 2),
                round(monthly_tds_total, 2),
                round(monthly_tds_total / max(months_processed, 1), 2),
                months_processed
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row_idx, column=col, value=v)
                c.border = BORDER_THIN
                c.font = CELL_FONT
                if col >= 4:
                    c.alignment = Alignment(horizontal="right")
            grand_tds += monthly_tds_total
            grand_gross += gross_total
            row_idx += 1

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
        ws.cell(row=row_idx, column=1, value="GRAND TOTAL").font = Font(bold=True)
        ws.cell(row=row_idx, column=4, value=round(grand_gross, 2)).font = Font(bold=True)
        c = ws.cell(row=row_idx, column=6, value=round(grand_tds, 2))
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        autosize_columns(ws)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="TDS_FY{fy_start_year}_{str(fy_start_year+1)[-2:]}.xlsx"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"[REPORTS] TDS error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 9. FORM 16 PDF (Phase B) ====================
@router.get("/form-16/{employee_id}")
async def form_16_pdf(employee_id: str, fy_start_year: int = Query(...)):
    """Form 16 Part A + B PDF — annual TDS certificate."""
    try:
        emp = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0})
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")

        months = [(fy_start_year, m) for m in range(4, 13)] + [(fy_start_year + 1, m) for m in range(1, 4)]

        total_gross = 0
        total_basic = 0
        total_pf = 0
        total_tds = 0
        total_pt = 0
        quarterly_tds = {"Q1": 0, "Q2": 0, "Q3": 0, "Q4": 0}

        for idx, (yr, m) in enumerate(months):
            att = await get_month_attendance_summary(employee_id, m, yr)
            if att["effective_days"] == 0:
                continue
            monthly = emp.get("monthly_salary", 0)
            bd = calculate_salary_breakdown(monthly)
            adj_basic = round(bd["earnings"]["basic_salary"] * att["attendance_ratio"], 2)
            adj_gross = round(monthly * att["attendance_ratio"], 2)
            ded = calculate_deductions(adj_gross, adj_basic, full_monthly_for_tds=monthly)
            total_gross += adj_gross
            total_basic += adj_basic
            total_pf += ded["pf_employee"]
            total_tds += ded["tds"]
            total_pt += ded["professional_tax"]
            q = "Q1" if idx < 3 else "Q2" if idx < 6 else "Q3" if idx < 9 else "Q4"
            quarterly_tds[q] += ded["tds"]

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm,
                                 topMargin=12 * mm, bottomMargin=12 * mm,
                                 title=f"Form 16 FY{fy_start_year}-{str(fy_start_year+1)[-2:]}")
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle('t', parent=styles['Title'], fontSize=13, alignment=1)
        small_center = ParagraphStyle('s', parent=styles['Normal'], fontSize=9, alignment=1, textColor=colors.HexColor("#475569"))

        story.append(Paragraph(f"FORM NO. 16 [See rule 31(1)(a)]", title_style))
        story.append(Paragraph(f"Certificate under section 203 of the Income-tax Act, 1961", small_center))
        story.append(Paragraph(f"for tax deducted at source from income chargeable under the head 'Salaries'", small_center))
        story.append(Spacer(1, 8))

        story.append(Paragraph(f"<b>PART A — Details of Tax Deducted and Deposited</b>",
                                ParagraphStyle('h', parent=styles['Heading3'], fontSize=10)))

        fy = f"{fy_start_year}-{fy_start_year + 1}"
        ay = f"{fy_start_year + 1}-{fy_start_year + 2}"

        # Employer/Employee block
        header_tbl = Table([
            ["Name of the Employer", COMPANY_NAME],
            ["Address", COMPANY_ADDRESS],
            ["PAN of Deductor", emp.get("employer_pan", "—")],
            ["TAN of Deductor", emp.get("employer_tan", "—")],
            ["", ""],
            ["Name of the Employee", emp.get("name", "")],
            ["Employee PAN", emp.get("pan", "—")],
            ["Employee ID", emp.get("employee_id", "")],
            ["Financial Year", fy],
            ["Assessment Year", ay],
        ], colWidths=[60 * mm, 120 * mm])
        header_tbl.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(header_tbl)
        story.append(Spacer(1, 8))

        # Quarterly TDS
        story.append(Paragraph("<b>Quarterly TDS Summary</b>",
                                ParagraphStyle('q', parent=styles['Heading4'], fontSize=9)))
        qtbl = Table([
            ["Quarter", "Amount Deducted (INR)", "Amount Deposited (INR)"],
            ["Q1 (Apr-Jun)", f"{quarterly_tds['Q1']:,.2f}", f"{quarterly_tds['Q1']:,.2f}"],
            ["Q2 (Jul-Sep)", f"{quarterly_tds['Q2']:,.2f}", f"{quarterly_tds['Q2']:,.2f}"],
            ["Q3 (Oct-Dec)", f"{quarterly_tds['Q3']:,.2f}", f"{quarterly_tds['Q3']:,.2f}"],
            ["Q4 (Jan-Mar)", f"{quarterly_tds['Q4']:,.2f}", f"{quarterly_tds['Q4']:,.2f}"],
            ["TOTAL", f"{total_tds:,.2f}", f"{total_tds:,.2f}"],
        ], colWidths=[60 * mm, 60 * mm, 60 * mm])
        qtbl.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#DBEAFE")),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#FEF3C7")),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ]))
        story.append(qtbl)
        story.append(Spacer(1, 12))

        story.append(Paragraph("<b>PART B — Details of Salary Paid and Tax Deducted</b>",
                                ParagraphStyle('h', parent=styles['Heading3'], fontSize=10)))

        # Salary details
        income_tax_payable = total_tds
        net_taxable = max(0, total_gross - 50000 - total_pf)  # Standard deduction + PF
        sal_tbl = Table([
            ["1.", "Gross Salary", f"{total_gross:,.2f}"],
            ["2.", "Less: Allowances exempt u/s 10", "0.00"],
            ["3.", "Balance (1-2)", f"{total_gross:,.2f}"],
            ["4.", "Deductions u/s 16:", ""],
            ["", "  (a) Standard deduction u/s 16(ia)", "50,000.00"],
            ["", "  (b) Professional Tax u/s 16(iii)", f"{total_pt:,.2f}"],
            ["5.", "Income chargeable under head 'Salaries' (3-4)", f"{max(0, total_gross - 50000 - total_pt):,.2f}"],
            ["6.", "Less: Deduction u/s 80C (PF)", f"{total_pf:,.2f}"],
            ["7.", "Total Taxable Income", f"{net_taxable:,.2f}"],
            ["8.", "Tax on total income (TDS Deducted)", f"{income_tax_payable:,.2f}"],
            ["9.", "Rebate u/s 87A", "—"],
            ["10.", "Net Tax Payable (TDS)", f"{income_tax_payable:,.2f}"],
        ], colWidths=[12 * mm, 108 * mm, 60 * mm])
        sal_tbl.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(sal_tbl)
        story.append(Spacer(1, 16))

        # Declaration
        decl_style = ParagraphStyle('d', parent=styles['Normal'], fontSize=8.5,
                                     textColor=colors.HexColor("#334155"), leading=12)
        story.append(Paragraph(
            "I, <b>Authorized Signatory</b>, working in the capacity of Finance Manager, do hereby certify that "
            "the information given above is true, complete and correct and is based on the books of accounts, "
            "documents, TDS statements and other available records.",
            decl_style
        ))
        story.append(Spacer(1, 20))
        sig_tbl = Table([
            ["Place:", COMPANY_ADDRESS.split(",")[-2].strip() if "," in COMPANY_ADDRESS else "", "", "Signature of Person Responsible"],
            ["Date:", datetime.now(timezone.utc).strftime("%d %b %Y"), "", f"Designation: HR / Finance"]
        ], colWidths=[20 * mm, 70 * mm, 20 * mm, 70 * mm])
        sig_tbl.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (3, 0), (3, 0), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LINEABOVE', (3, 0), (3, 0), 0.5, colors.HexColor("#94A3B8")),
        ]))
        story.append(sig_tbl)

        doc.build(story)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="Form16_{emp.get("name","emp").replace(" ","_")}_FY{fy_start_year}-{str(fy_start_year+1)[-2:]}.pdf"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"[REPORTS] Form 16 error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 10. LEAVE BALANCE REPORT (Phase C) ====================
@router.get("/leave-balance")
async def leave_balance_report():
    """Leave balances summary Excel — all active employees with CL/SL/EL balances + used."""
    try:
        employees = await db.employees.find({"status": "active"}, {"_id": 0}).to_list(10000)
        if not employees:
            raise HTTPException(status_code=404, detail="No active employees")

        year_start = f"{datetime.now(timezone.utc).year}-01-01"
        wb = Workbook()
        ws = wb.active
        ws.title = "Leave Balance"

        header_row = write_company_header(ws, 1, 11, f"LEAVE BALANCE REPORT — as on {datetime.now(timezone.utc).strftime('%d %b %Y')}")
        headers = ["Emp ID", "Name", "Department", "DOJ",
                    "CL Total", "CL Used", "CL Balance",
                    "SL Total", "SL Used", "SL Balance",
                    "EL Balance"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=header_row + 1, column=col, value=h)
        apply_header_row(ws, header_row + 1, len(headers))

        for i, emp in enumerate(employees):
            lb = emp.get("leave_balance", {"cl": 12, "sl": 12, "el": 21})
            # Used this year
            used = await db.employee_leaves.aggregate([
                {"$match": {"employee_id": emp["employee_id"], "status": "approved", "from_date": {"$gte": year_start}}},
                {"$group": {"_id": "$leave_type", "days": {"$sum": {"$ifNull": ["$days", 1]}}}}
            ]).to_list(10)
            used_map = {u["_id"]: u.get("days", 0) for u in used}
            cl_used = used_map.get("casual", 0) + used_map.get("cl", 0)
            sl_used = used_map.get("sick", 0) + used_map.get("sl", 0)
            el_bal = lb.get("el", 0) if isinstance(lb, dict) else 0

            vals = [
                emp["employee_id"], emp.get("name", ""), emp.get("department", ""),
                str(emp.get("joining_date", ""))[:10],
                lb.get("cl", 12), cl_used, lb.get("cl", 12) - cl_used,
                lb.get("sl", 12), sl_used, lb.get("sl", 12) - sl_used,
                el_bal
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=header_row + 2 + i, column=col, value=v)
                c.border = BORDER_THIN
                c.font = CELL_FONT
                if col >= 5:
                    c.alignment = Alignment(horizontal="center")

        autosize_columns(ws)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="Leave_Balance_{datetime.now().strftime("%Y%m%d")}.xlsx"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"[REPORTS] Leave balance error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 11. YTD EARNINGS (Phase C) ====================
@router.get("/ytd-earnings/{employee_id}")
async def ytd_earnings(employee_id: str, fy_start_year: int = Query(...)):
    """Year-to-date earnings breakdown for one employee (JSON) — useful for self-service."""
    try:
        emp = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0})
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")

        months = [(fy_start_year, m) for m in range(4, 13)] + [(fy_start_year + 1, m) for m in range(1, 4)]
        monthly_data = []
        totals = {"gross": 0, "net": 0, "pf": 0, "tds": 0, "pt": 0, "esi": 0}

        for yr, m in months:
            att = await get_month_attendance_summary(employee_id, m, yr)
            if att["effective_days"] == 0:
                continue
            monthly = emp.get("monthly_salary", 0)
            bd = calculate_salary_breakdown(monthly)
            adj_basic = round(bd["earnings"]["basic_salary"] * att["attendance_ratio"], 2)
            adj_gross = round(monthly * att["attendance_ratio"], 2)
            ded = calculate_deductions(adj_gross, adj_basic, full_monthly_for_tds=monthly)
            lop = round(monthly - (monthly * att["attendance_ratio"]), 2) if att["attendance_ratio"] < 1 else 0
            total_ded = ded["total"] + lop
            net = round(adj_gross - total_ded, 2)

            monthly_data.append({
                "month": calendar.month_abbr[m],
                "year": yr,
                "gross": adj_gross, "net": net,
                "pf": ded["pf_employee"], "esi": ded["esi_employee"],
                "pt": ded["professional_tax"], "tds": ded["tds"],
                "lop": lop, "days_paid": att["effective_days"]
            })
            totals["gross"] += adj_gross
            totals["net"] += net
            totals["pf"] += ded["pf_employee"]
            totals["tds"] += ded["tds"]
            totals["pt"] += ded["professional_tax"]
            totals["esi"] += ded["esi_employee"]

        return {
            "success": True,
            "employee": {
                "employee_id": emp.get("employee_id"),
                "name": emp.get("name"),
                "designation": emp.get("designation"),
                "department": emp.get("department"),
                "pan": emp.get("pan"),
                "doj": str(emp.get("joining_date", ""))[:10]
            },
            "financial_year": f"{fy_start_year}-{str(fy_start_year + 1)[-2:]}",
            "monthly": monthly_data,
            "totals": {k: round(v, 2) for k, v in totals.items()},
            "ytd_days": sum(m["days_paid"] for m in monthly_data),
            "months_processed": len(monthly_data)
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"[REPORTS] YTD error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ==================== HELPERS ====================
def number_to_words_inr(amount: float) -> str:
    """Convert number to Indian-format words (e.g., 1,25,340 -> One Lakh Twenty Five Thousand Three Hundred Forty Rupees)."""
    try:
        amount = float(amount or 0)
    except Exception:
        return "Zero Rupees"
    rupees = int(amount)
    paise = int(round((amount - rupees) * 100))

    def _below_thousand(n):
        ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
                "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
                "Seventeen", "Eighteen", "Nineteen"]
        tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
        if n < 20:
            return ones[n]
        if n < 100:
            return (tens[n // 10] + (" " + ones[n % 10] if n % 10 else "")).strip()
        return (ones[n // 100] + " Hundred" + (" " + _below_thousand(n % 100) if n % 100 else "")).strip()

    def _indian(n):
        if n == 0:
            return "Zero"
        parts = []
        crore = n // 10000000
        n %= 10000000
        lakh = n // 100000
        n %= 100000
        thousand = n // 1000
        n %= 1000
        rest = n
        if crore:
            parts.append(_below_thousand(crore) + " Crore")
        if lakh:
            parts.append(_below_thousand(lakh) + " Lakh")
        if thousand:
            parts.append(_below_thousand(thousand) + " Thousand")
        if rest:
            parts.append(_below_thousand(rest))
        return " ".join(parts).strip()

    result = _indian(rupees) + " Rupees"
    if paise:
        result += f" and {_below_thousand(paise)} Paise"
    return result + " Only"
